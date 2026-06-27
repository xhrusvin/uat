"""
live_staff_signature.py
────────────────────────
Cron + export routes for signature handwriting validation using Gemini Vision.

Routes:
    GET/POST  /live-staffs/cron/sync-signature-validation
    GET       /live-staffs/export/signature-validation-xlsx

Env vars required:
    GEMINI_API_KEY          — Google AI / Gemini API key
    GCS_BUCKET_NAME         — GCS bucket where signatures are stored
                              (e.g. "xpress-health-prod")
    GOOGLE_APPLICATION_CREDENTIALS — path to GCS service account JSON key file
                              OR set GCS_SA_JSON with the JSON content as a string
    CRON_SECRET             — (optional) guards the cron endpoint

Register in your app factory alongside the other admin blueprint files.
"""

import os
import re
import json as _json
import base64
import datetime as _dt
import requests as _req

from flask import request, jsonify, Response

from database import db
from . import admin_bp
from admin.views import admin_required


# ── Helpers ────────────────────────────────────────────────────────────────────

def _staffs_col():
    return db.live_staffs


def _v(val):
    if val is None:
        return ''
    return str(val).strip()


def _gcs_client():
    """
    Identical to live_staffs.py — reuse same env vars and priority order:
      1. GCS_CREDENTIALS_JSON  (full SA JSON as string)
      2. GCS_KEY_FILE          (path to SA JSON file)
      3. Application Default Credentials (ADC)
    """
    from google.cloud import storage as _gcs

    creds_json = os.environ.get('GCS_CREDENTIALS_JSON', '').strip()
    if creds_json:
        from google.oauth2 import service_account
        info  = _json.loads(creds_json)
        creds = service_account.Credentials.from_service_account_info(
            info,
            scopes=["https://www.googleapis.com/auth/cloud-platform"],
        )
        return _gcs.Client(credentials=creds, project=info.get('project_id'))

    key_path = os.environ.get('GCS_KEY_FILE', '').strip()
    if key_path and os.path.exists(key_path):
        return _gcs.Client.from_service_account_json(key_path)

    return _gcs.Client()


def _gcs_download(blob_name: str) -> bytes:
    """Direct byte download from GCS — no signed URL, no IAM signing needed."""
    bucket_name = os.environ.get('GCS_BUCKET_NAME', '')
    if not bucket_name:
        raise ValueError("GCS_BUCKET_NAME env var is not set")
    bucket = _gcs_client().bucket(bucket_name)
    return bucket.blob(blob_name).download_as_bytes()


def _download_signature(staff: dict) -> tuple:
    """
    Returns (raw_bytes, content_type).
    Uses _gcs_download(blob_name) — same pattern as live_staffs.py line ~4896.
    No signed URLs, no IAM scopes needed.
    Falls back to HTTP fetch of signature_url only if no GCS blob is stored.
    """
    gcs_blob = _v(staff.get('signature_gcs_blob') or '')

    if gcs_blob:
        raw_bytes = _gcs_download(gcs_blob)
        # Infer content type from file extension
        ext = gcs_blob.rsplit('.', 1)[-1].lower() if '.' in gcs_blob else 'png'
        mime_map = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                    'png': 'image/png',  'gif': 'image/gif', 'webp': 'image/webp'}
        content_type = mime_map.get(ext, 'image/png')
        return raw_bytes, content_type

    # Fallback: HTTP fetch from signature_url (may be expired for old records)
    url = _v(staff.get('signature_url') or '')
    if not url:
        raise ValueError("No signature_gcs_blob or signature_url on record")

    resp = _req.get(url, timeout=30)
    if resp.status_code == 404:
        raise FileNotFoundError("Signature URL returned 404")
    resp.raise_for_status()

    content_type = resp.headers.get('Content-Type', 'image/png').lower().split(';')[0].strip()
    if not content_type.startswith('image/'):
        content_type = 'image/png'
    return resp.content, content_type


# ── Cron: Validate signature handwriting via Gemini ────────────────────────────

@admin_bp.route('/live-staffs/cron/sync-signature-validation', methods=['GET', 'POST'])
def live_staff_cron_sync_signature_validation():
    """
    Cron job — processes ONE staff member per call.
    Downloads the staff's signature from GCS (fresh signed URL via signature_gcs_blob),
    sends it to Gemini Vision to determine if it is a genuine handwritten signature.

    Saves to MongoDB:
        sig_valid              : bool
        sig_validation_note    : str   (Gemini's reason)
        sig_fetched            : True
        sig_fetched_at         : datetime
    """
    from google import genai as google_genai

    # ── auth ──────────────────────────────────────────────────────────────
    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    gemini_key = os.environ.get('GEMINI_API_KEY', '')
    if not gemini_key:
        return jsonify({"success": False, "error": "GEMINI_API_KEY not set"}), 500

    col = _staffs_col()

    # ── find next unprocessed staff that has a signature ──────────────────
    pending_query = {
        "$or": [
            {"signature_gcs_blob": {"$exists": True, "$ne": None, "$ne": ""}},
            {"signature_url":      {"$exists": True, "$ne": None, "$ne": ""}},
        ],
        "$and": [{
            "$or": [
                {"sig_fetched": {"$exists": False}},
                {"sig_fetched": False},
                {"sig_fetched": None},
            ]
        }],
    }
    remaining_total = col.count_documents(pending_query)
    staff           = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff signatures already validated.",
            "remaining_count": 0,
        })

    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or s1.get('email_address') or '')
    emp_code  = _v(staff.get('employee_code') or '')

    def _mark_done(fields):
        fields["sig_fetched"]    = True
        fields["sig_fetched_at"] = _dt.datetime.utcnow()
        col.update_one({"_id": staff['_id']}, {"$set": fields})

    # ── download signature (fresh signed URL from GCS blob) ───────────────
    try:
        raw_bytes, content_type = _download_signature(staff)
    except FileNotFoundError as e:
        _mark_done({"sig_valid": False, "sig_validation_note": str(e)})
        return jsonify({
            "success": True, "skipped": True, "email": email,
            "staff_name": full_name,
            "reason": str(e),
            "remaining_count": max(0, remaining_total - 1),
        })
    except Exception as e:
        _mark_done({"sig_valid": False,
                    "sig_validation_note": f"download error: {e}"})
        return jsonify({
            "success": False, "email": email, "staff_name": full_name,
            "error":   f"Failed to download signature: {e}",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── send to Gemini Vision ─────────────────────────────────────────────
    prompt_text = """You are a document verification expert reviewing a staff onboarding system.

Examine this image and determine whether it contains a genuine handwritten signature.

A VALID handwritten signature has:
- Natural, flowing pen or stylus strokes
- Slight irregularities and personal style
- Cursive, stylised, or abbreviated writing

Mark as INVALID if the image is:
- A blank or nearly blank white image
- A printed or typed name
- A rubber stamp or digital clipart
- A checkbox, tick, or unrelated graphic
- Completely illegible or corrupted

Reply ONLY with a JSON object — no markdown, no extra text:
{"valid": true, "reason": "brief one-sentence explanation"}
or
{"valid": false, "reason": "brief one-sentence explanation"}"""

    try:
        client   = google_genai.Client(api_key=gemini_key)
        b64_data = base64.standard_b64encode(raw_bytes).decode('utf-8')
        parts    = [
            {"inline_data": {"mime_type": content_type, "data": b64_data}},
            {"text": prompt_text},
        ]
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=[{"parts": parts}],
        )

        raw_out = (response.text or '').strip()
        raw_out = re.sub(r'^```(?:json)?\s*', '', raw_out, flags=re.MULTILINE)
        raw_out = re.sub(r'```\s*$',          '', raw_out, flags=re.MULTILINE).strip()

        result   = _json.loads(raw_out)
        is_valid = bool(result.get('valid', False))
        reason   = _v(result.get('reason') or '')

    except _json.JSONDecodeError:
        _mark_done({"sig_valid": False,
                    "sig_validation_note": "Gemini returned non-JSON"})
        return jsonify({
            "success": False, "email": email, "staff_name": full_name,
            "error":   "Gemini returned non-JSON",
            "remaining_count": max(0, remaining_total - 1),
        })
    except Exception as e:
        _mark_done({"sig_valid": False,
                    "sig_validation_note": f"Gemini error: {e}"})
        return jsonify({
            "success": False, "email": email, "staff_name": full_name,
            "error":   str(e),
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── persist result ────────────────────────────────────────────────────
    _mark_done({
        "sig_valid":           is_valid,
        "sig_validation_note": reason,
    })

    return jsonify({
        "success":         True,
        "email":           email,
        "staff_name":      full_name,
        "employee_code":   emp_code,
        "valid":           is_valid,
        "reason":          reason,
        "remaining_count": max(0, remaining_total - 1),
        "message": (
            f"Signature {'✔ valid' if is_valid else '✘ invalid'} for {full_name} "
            f"({email}) — {max(0, remaining_total - 1)} remaining."
        ),
    })


# ── Export: Signature validation results to Excel ──────────────────────────────

@admin_bp.route('/live-staffs/export/signature-validation-xlsx')
@admin_required
def live_staff_export_signature_validation_xlsx():
    """Export signature validation results to a formatted Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as _io

        docs = list(_staffs_col().find(
            {},
            {
                "section_1_personal_details": 1,
                "email":               1,
                "employee_code":       1,
                "sig_valid":           1,
                "sig_validation_note": 1,
                "sig_fetched":         1,
            }
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        NAVY     = '1B3A6B'
        WHITE    = 'FFFFFF'
        ALT      = 'EFF6FF'
        GREEN_BG = 'E2EFDA'
        RED_BG   = 'FFDDDD'
        WARN_BG  = 'FFF3CD'

        h_font   = Font(name='Arial', bold=True, color=WHITE, size=10)
        h_fill   = PatternFill('solid', start_color=NAVY,  end_color=NAVY)
        h_align  = Alignment(horizontal='center', vertical='center')
        b_font   = Font(name='Arial', size=10)
        l_align  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        c_align  = Alignment(horizontal='center', vertical='center')
        thin     = Side(style='thin',   color='CCCCCC')
        med_grn  = Side(style='medium', color='2E9E44')
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)
        h_border = Border(left=thin, right=thin, top=thin, bottom=med_grn)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Signature Validation'

        headers    = ['Sno', 'Staff Name', 'Email', 'Employee Code',
                      'Valid Signature', 'Reason / Note', 'Status']
        col_widths = [5, 28, 36, 18, 16, 55, 14]

        for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell            = ws.cell(row=1, column=ci, value=hdr)
            cell.font       = h_font
            cell.fill       = h_fill
            cell.alignment  = h_align
            cell.border     = h_border
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 24
        ws.freeze_panes             = 'A2'
        ws.auto_filter.ref          = f'A1:G{len(docs) + 1}'

        for ri, doc in enumerate(docs, start=2):
            s1       = doc.get('section_1_personal_details') or {}
            name     = _v(s1.get('full_name')          or '')
            email    = _v(doc.get('email')              or '')
            emp_code = _v(doc.get('employee_code')      or '')
            valid    = doc.get('sig_valid')
            note     = _v(doc.get('sig_validation_note') or '')
            fetched  = doc.get('sig_fetched', False)

            if not fetched:
                status, valid_label = 'Not Checked', '—'
                row_fill = PatternFill('solid', start_color=WARN_BG,  end_color=WARN_BG)
            elif valid is True:
                status, valid_label = 'Valid',   '✔  Valid'
                row_fill = PatternFill('solid', start_color=GREEN_BG, end_color=GREEN_BG)
            elif valid is False:
                status, valid_label = 'Invalid', '✘  Invalid'
                row_fill = PatternFill('solid', start_color=RED_BG,   end_color=RED_BG)
            else:
                status, valid_label = 'Unknown', '?'
                row_fill = PatternFill('solid', start_color=ALT, end_color=ALT)

            alt_fill = PatternFill('solid', start_color=ALT, end_color=ALT) \
                       if ri % 2 == 0 else None

            row_vals = [ri - 1, name, email, emp_code, valid_label, note, status]
            aligns   = [c_align, l_align, l_align, c_align, c_align, l_align, c_align]

            for ci, (val, align) in enumerate(zip(row_vals, aligns), start=1):
                cell           = ws.cell(row=ri, column=ci, value=val)
                cell.font      = b_font
                cell.alignment = align
                cell.border    = border
                cell.fill      = row_fill or alt_fill or PatternFill()

            ws.row_dimensions[ri].height = 17

        total   = len(docs)
        valid_c = sum(1 for d in docs if d.get('sig_valid') is True)
        inv_c   = sum(1 for d in docs if d.get('sig_valid') is False)
        pend_c  = total - valid_c - inv_c

        summary_row = total + 2
        ws.cell(row=summary_row, column=1,
                value=f'Total: {total}  |  Valid: {valid_c}  |  Invalid: {inv_c}  |  Pending: {pend_c}'
                ).font = Font(name='Arial', bold=True, size=9, color=NAVY)

        buf = _io.BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition":
                    f'attachment; filename="signature_validation_'
                    f'{_dt.datetime.utcnow().strftime("%Y%m%d")}.xlsx"'
            },
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
