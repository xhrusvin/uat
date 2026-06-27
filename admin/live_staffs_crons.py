from flask import render_template, request, jsonify, Response
from bson import ObjectId
from datetime import datetime
import json
import json as _cjson
import json as _cjson2
import base64
import csv
import io
import re as _re
import re
import os
import threading

from database import db
from . import admin_bp
from admin.views import admin_required

# ── Helpers — lazy wrappers to avoid circular imports ────────────────

def _v(val):
    if val is None: return ''
    return str(val).strip()

def _staffs_col():
    from flask import current_app
    return current_app.db.live_staffs

def _gcs_upload(blob_name, data_bytes, content_type='application/octet-stream'):
    from admin.live_staffs import _gcs_upload as _f
    return _f(blob_name, data_bytes, content_type)

def _gcs_download(blob_name):
    from admin.live_staffs import _gcs_download as _f
    return _f(blob_name)

def _gcs_signed_url(blob_name, expiry_minutes=60):
    from admin.live_staffs import _gcs_signed_url as _f
    return _f(blob_name, expiry_minutes)

def _ai_pcc_col():
    from flask import current_app
    return current_app.db.live_staff_ai_pcc

def _ai_cvs_col():
    from flask import current_app
    return current_app.db.live_staff_ai_cvs

def _ai_interviews_col():
    from flask import current_app
    return current_app.db.live_staff_ai_interviews

def _ai_appforms_col():
    from flask import current_app
    return current_app.db.live_staff_ai_appforms

def _build_pcc_docx(doc, reviewer_index=0):
    from admin.live_staffs import _build_pcc_docx as _f
    return _f(doc, reviewer_index)

def _build_ai_cv_docx(doc, cv_text):
    from admin.live_staffs import _build_ai_cv_docx as _f
    return _f(doc, cv_text)

def _build_ai_interview_docx(doc, interview_text):
    from admin.live_staffs import _build_ai_interview_docx as _f
    return _f(doc, interview_text)

def _build_ai_appform_docx(doc, appform_text):
    from admin.live_staffs import _build_ai_appform_docx as _f
    return _f(doc, appform_text)

def _build_appform_docx(doc):
    from admin.live_staffs import _build_appform_docx as _f
    return _f(doc)

def _build_interview_docx(doc, interview_text):
    from admin.live_staffs import _build_interview_docx as _f
    return _f(doc, interview_text)

def _get_pcc_reviewers():
    from admin.live_staffs import _PCC_REVIEWERS
    return _PCC_REVIEWERS

def _get_compliance_officer():
    from admin.live_staffs import _PCC_COMPLIANCE_OFFICER
    return _PCC_COMPLIANCE_OFFICER

def _push_hse_document_background(staff_id_str, doc_type_key,
                                   gcs_blob, filename,
                                   user_type=None):
    from admin.live_staffs import _push_hse_document_background as _f
    return _f(staff_id_str, doc_type_key, gcs_blob, filename, user_type)


@admin_bp.route('/live-staffs/cron/sync-documents', methods=['GET', 'POST'])
def live_staff_cron_sync_documents():
    """
    Cron job — processes ONE staff member per call.

    Logic:
      1. Find the first live_staffs record where extracted_cv is missing/empty.
      2. Call the XN Portal API to get their document list.
      3. If a document_type_name == "Cv" has a URL, download + extract text via Gemini.
      4. Save documents[] and extracted_cv back to MongoDB.
      5. Return result with how many staff still need processing (remaining_count).

    Call every N minutes via cron — it will work through all staff automatically,
    one at a time, until everyone has an extracted_cv.

    Protect with ?cron_key=<CRON_SECRET> env var.
    """
    import requests as _req

    # ── Auth ──────────────────────────────────────────────────────────
    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    # ── Env ───────────────────────────────────────────────────────────
    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')

    if not base_url:
        return jsonify({"success": False,
                        "error": "LIVE_STAFF_URL not set in environment"}), 500

    endpoint = f"{base_url}/ai/recruitments/user-document-list"
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    col = _staffs_col()

    # ── Find next staff without extracted_cv ──────────────────────────
    staff = col.find_one(
        {"$or": [
            {"extracted_cv": {"$exists": False}},
            {"extracted_cv": None},
            {"extracted_cv": ""},
        ]},
        {"email": 1, "section_1_personal_details": 1}
    )

    # Count how many still need processing (including this one)
    remaining_total = col.count_documents(
        {"$or": [
            {"extracted_cv": {"$exists": False}},
            {"extracted_cv": None},
            {"extracted_cv": ""},
        ]}
    )

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff already have extracted_cv — nothing to do.",
            "remaining_count": 0,
            "processed":       None,
        })

    email = _v(
        (staff.get('section_1_personal_details') or {}).get('email_address') or
        staff.get('email') or ''
    )

    if not email:
        # Mark as attempted so it doesn't block the queue forever
        col.update_one(
            {"_id": staff["_id"]},
            {"$set": {"extracted_cv": "[skipped — no email]",
                      "extracted_cv_at": datetime.utcnow()}}
        )
        return jsonify({
            "success":         True,
            "message":         "Skipped — staff record has no email address.",
            "remaining_count": remaining_total - 1,
            "processed":       str(staff.get("_id")),
        })

    # ── Call XN Portal API ────────────────────────────────────────────
    status_code   = None
    response_text = ''
    try:
        resp = _req.post(
            endpoint,
            json={"email": email},
            headers=api_headers,
            timeout=30
        )
        # If POST not allowed, retry with GET + query param
        if resp.status_code == 405:
            resp = _req.get(
                endpoint,
                params={"email": email},
                headers=api_headers,
                timeout=30
            )
        status_code   = resp.status_code
        response_text = resp.text[:500] if resp.text else ''
        resp.raise_for_status()
        data = resp.json()
    except _req.exceptions.ConnectionError as api_err:
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"Connection error — cannot reach XN Portal",
            "detail":          str(api_err)[:300],
            "endpoint":        endpoint,
            "check":           "Verify LIVE_STAFF_URL env var is correct and server is reachable",
            "remaining_count": remaining_total,
        })
    except _req.exceptions.Timeout:
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           "Timeout — XN Portal did not respond within 30s",
            "endpoint":        endpoint,
            "remaining_count": remaining_total,
        })
    except _req.exceptions.HTTPError as api_err:
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"HTTP {status_code} error from XN Portal",
            "response_body":   response_text,
            "endpoint":        endpoint,
            "remaining_count": remaining_total,
        })
    except Exception as api_err:
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"{type(api_err).__name__}: {api_err}",
            "endpoint":        endpoint,
            "remaining_count": remaining_total,
        })

    if not data.get('success'):
        # Mark attempted so cron moves on next time
        col.update_one(
            {"_id": staff["_id"]},
            {"$set": {"extracted_cv": f"[API error: {data.get('message', 'unknown')}]",
                      "extracted_cv_at": datetime.utcnow()}}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           data.get('message', 'API returned success=false'),
            "api_raw":         data,
            "remaining_count": remaining_total - 1,
        })

    api_data = data.get('data')

    # API may return data as a list (empty []) or a dict with documents key
    if isinstance(api_data, list):
        documents = api_data          # data itself is the document list
    elif isinstance(api_data, dict):
        documents = api_data.get('documents') or []
    else:
        documents = []

    # No documents found — save "No doc found" and move on
    if not documents:
        col.update_one(
            {"_id": staff["_id"]},
            {"$set": {
                "extracted_cv":    "No doc found",
                "extracted_cv_at": datetime.utcnow(),
            }}
        )
        return jsonify({
            "success":         True,
            "email":           email,
            "documents_found": 0,
            "cv_extracted":    False,
            "extracted_cv":    "No doc found",
            "remaining_count": max(0, remaining_total - 1),
            "message":         f"No documents returned by API for {email} — marked as 'No doc found'.",
        })

    # ── Process documents ─────────────────────────────────────────────
    extracted_cv_text = None
    cv_error          = None
    cv_url_found      = None

    for doc_item in documents:
        # Extract only when document_type_name is exactly "Cv"
        doc_name = (doc_item.get('document_type_name') or '').strip()
        doc_url  = doc_item.get('url') or ''

        if doc_name == 'Cv' and doc_url and not extracted_cv_text:
            cv_url_found = doc_url
            try:
                extracted_cv_text = _extract_text_from_url(doc_url, api_headers)
            except Exception as cv_err:
                cv_error = str(cv_err)
                extracted_cv_text = f"[CV extraction failed: {cv_err}]"

    # ── Save only extracted_cv to MongoDB ─────────────────────────────
    col.update_one(
        {"_id": staff["_id"]},
        {"$set": {
            "extracted_cv":      extracted_cv_text or "[no CV document found]",
            "extracted_cv_at":   datetime.utcnow(),
        }}
    )

    return jsonify({
        "success":         True,
        "email":           email,
        "cv_extracted":    bool(extracted_cv_text and not cv_error),
        "cv_url_found":    cv_url_found,
        "cv_error":        cv_error,
        "remaining_count": max(0, remaining_total - 1),
        "message": (
            f"Processed {email} — "
            f"{max(0, remaining_total - 1)} staff still need extraction."
        ),
        "synced_at": datetime.utcnow().isoformat(),
    })


def _extract_text_from_url(url, headers=None):
    """
    Download a CV document from URL, then use Gemini AI to extract
    and structure the full text content.

    Strategy:
      1. Download the file (PDF or DOCX).
      2. Get raw text via pdfplumber / python-docx as a pre-extraction step.
      3. Send that raw text to Gemini 2.5 Flash to clean, structure,
         and return a well-formatted plain-text CV extraction.
      4. If Gemini is unavailable, fall back to raw text only.
    """
    import requests as _req
    import io as _io

    # ── Download file ─────────────────────────────────────────────────
    dl_headers = dict(headers or {})
    dl_headers.pop('Content-Type', None)

    resp = _req.get(url, headers=dl_headers, timeout=60)
    resp.raise_for_status()

    content_type = resp.headers.get('Content-Type', '').lower()
    raw          = resp.content
    url_lower    = url.lower().split('?')[0]

    # ── Step 1: raw text extraction ───────────────────────────────────
    raw_text = ''

    # PDF
    if 'pdf' in content_type or url_lower.endswith('.pdf'):
        try:
            import pdfplumber
            with pdfplumber.open(_io.BytesIO(raw)) as pdf:
                raw_text = chr(10).join( page.extract_text() or '' for page in pdf.pages ).strip()
        except Exception:
            try:
                import PyPDF2
                reader   = PyPDF2.PdfReader(_io.BytesIO(raw))
                raw_text = chr(10).join( page.extract_text() or '' for page in reader.pages ).strip()
            except Exception:
                pass

    # DOCX
    elif ('wordprocessingml' in content_type or
          url_lower.endswith('.docx') or url_lower.endswith('.doc')):
        try:
            from docx import Document as _DocxDoc
            d        = _DocxDoc(_io.BytesIO(raw))
            raw_text = chr(10).join(p.text for p in d.paragraphs).strip()
        except Exception:
            pass

    # Plain text
    elif 'text' in content_type:
        raw_text = raw.decode('utf-8', errors='replace').strip()

    # Last resort — try PDF
    if not raw_text:
        try:
            import pdfplumber
            with pdfplumber.open(_io.BytesIO(raw)) as pdf:
                raw_text = chr(10).join( page.extract_text() or '' for page in pdf.pages ).strip()
        except Exception:
            pass

    # ── Step 2: Gemini AI extraction & structuring ────────────────────
    gemini_key = os.environ.get('GEMINI_API_KEY', '')
    if not gemini_key:
        if not raw_text:
            raise RuntimeError(
                f"Could not extract any text from document (content-type: {content_type})"
            )
        return raw_text

    try:
        from google import genai as _genai
        client = _genai.Client(api_key=gemini_key)

        if raw_text:
            # Text-based PDF/DOCX — use text prompt
            prompt = f"""You are a professional CV parser.

The text below was extracted from a candidate's CV document (PDF or DOCX).
The text may be messy, have formatting issues, or be partially garbled from extraction.

Your task:
1. Read the raw extracted text carefully.
2. Identify and structure all CV content into clean, readable plain text.
3. Preserve ALL factual information exactly as stated — do NOT add, invent, or change any facts.
4. Format it with clear section headings (EMPLOYMENT ELIGIBILITY, PROFESSIONAL PROFILE, EDUCATION & QUALIFICATIONS, PROFESSIONAL EXPERIENCE, TRAINING & CERTIFICATIONS, KEY SKILLS, ADDITIONAL INFORMATION) where the content exists.
5. If a section's content is not present in the raw text, omit that section entirely.
6. Return ONLY the clean structured CV text — no preamble, no commentary.

RAW EXTRACTED TEXT:
{raw_text[:12000]}
"""
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt
            )
        else:
            # Scanned/image-based PDF — send raw bytes to Gemini vision
            import base64 as _b64_cv
            vision_prompt = """You are a professional CV parser.
This is a scanned or image-based PDF CV document.
Extract ALL CV content visible in the document and structure it into clean, readable plain text.
Preserve ALL factual information exactly — do NOT add, invent, or change any facts.
Extract ONLY CV/resume content — ignore any other documents (bills, utility letters etc).
Format with clear section headings (EMPLOYMENT ELIGIBILITY, PROFESSIONAL PROFILE, EDUCATION & QUALIFICATIONS, PROFESSIONAL EXPERIENCE, TRAINING & CERTIFICATIONS, KEY SKILLS, ADDITIONAL INFORMATION) where content exists.
Return ONLY the clean structured CV text — no preamble, no commentary."""
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[{
                    "parts": [
                        {"inline_data": {
                            "mime_type": "application/pdf",
                            "data": _b64_cv.b64encode(raw).decode()
                        }},
                        {"text": vision_prompt}
                    ]
                }]
            )
            return (response.text or '').strip() or '[no CV text extracted from scanned document]'

        response   = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt
        )
        gemini_out = (response.text or '').strip()

        # Return Gemini-structured text; fall back to raw if empty
        return gemini_out if gemini_out else raw_text

    except Exception as gemini_err:
        # Gemini failed — return raw text with a note
        return f"[Gemini extraction failed: {gemini_err}]\n\n{raw_text}"



# ── Cron: Generate AI CV one staff at a time ──────────────────────────

@admin_bp.route('/live-staffs/cron/generate-cv', methods=['GET', 'POST'])
def live_staff_cron_generate_cv():
    """
    Cron job — generates an AI CV for ONE staff member per call.

    Logic:
      1. Find the first live_staffs record where gcs_blob is missing
         in live_staff_ai_cvs (i.e. no CV generated yet).
      2. Call Gemini to write the CV.
      3. Build DOCX, upload to GCS, save metadata.
      4. Trigger HSE document upload in background.
      5. Return remaining_count so you know how many are left.

    Call every N seconds via cron until remaining_count == 0.

    Protect with ?cron_key=<CRON_SECRET> env var.
    """
    import requests as _req

    # ── Auth ──────────────────────────────────────────────────────────
    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    gemini_key = os.environ.get('GEMINI_API_KEY', '')
    if not gemini_key:
        return jsonify({"success": False,
                        "error": "GEMINI_API_KEY not set"}), 500

    staffs_col  = _staffs_col()
    ai_cvs_col  = _ai_cvs_col()

    # ── Find next staff without a generated CV ────────────────────────
    # Get set of staff_ids that already have a CV
    existing_ids = set(
        str(r['staff_id'])
        for r in ai_cvs_col.find({}, {"staff_id": 1})
        if r.get('staff_id')
    )

    staff = staffs_col.find_one(
        {"_id": {"$nin": [ObjectId(i) for i in existing_ids if len(i) == 24]}},
        {"email": 1, "section_1_personal_details": 1, "user_type": 1,
         "employee_code": 1, "extracted_cv": 1}
    )

    remaining_total = staffs_col.count_documents({}) - len(existing_ids)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff already have AI CVs — nothing to do.",
            "remaining_count": 0,
            "processed":       None,
        })

    staff_id  = str(staff['_id'])
    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or '')

    if not full_name and not email:
        # Mark as skipped so it won't block the queue
        ai_cvs_col.insert_one({
            "staff_id":     staff_id,
            "staff_name":   '',
            "cv_text":      '[skipped — no name or email]',
            "gcs_blob":     '',
            "generated_at": datetime.utcnow(),
        })
        return jsonify({
            "success":         True,
            "message":         f"Skipped {staff_id} — no name or email",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Build full doc object for the CV builder ───────────────────────
    full_doc = staffs_col.find_one({"_id": staff['_id']})
    if not full_doc:
        return jsonify({"success": False, "error": "Staff record disappeared"}), 500

    # ── Build prompt ─────────────────────────────────────────────────
    s1_f = full_doc.get('section_1_personal_details') or {}
    s3   = full_doc.get('section_3_professional_registration') or {}
    s4   = full_doc.get('section_4_qualifications') or {}
    s5   = full_doc.get('section_5_employment_history') or {}
    s8   = full_doc.get('section_8_garda_vetting_police_clearance') or {}
    s9   = full_doc.get('section_9_occupational_health') or {}
    s10  = full_doc.get('section_10_mandatory_training') or {}
    visa = s1_f.get('work_permit_visa_status') or {}

    def _vv(val):
        if val is None: return ''
        return str(val).strip()

    user_type   = _vv(full_doc.get('user_type'))
    emp_code    = _vv(full_doc.get('employee_code'))
    address     = _vv(s1_f.get('address'))
    mobile      = _vv(s1_f.get('mobile_number'))
    nationality = _vv(s1_f.get('nationality'))
    total_exp   = _vv(s5.get('total_experience'))
    divisions   = ', '.join(s3.get('divisions_registered_in') or [])
    reg_pin     = _vv(s3.get('registration_number_pin'))
    reg_exp     = _vv(s3.get('registration_expiry_date'))
    nmbi        = 'Yes' if s3.get('nmbi_active_declaration') else 'No'
    visa_type   = _vv(visa.get('visa_type'))
    perm_work   = _vv(visa.get('permission_to_work'))
    garda       = 'Yes' if s8.get('garda_vetting_submitted') else 'No'
    fit         = 'Yes' if s9.get('fit_for_nursing_duties') else 'No'
    dob         = _vv(s1_f.get('date_of_birth'))

    qual_lines = []
    for qk in ['nursing_degree', 'postgraduate_qualification', 'other_qualification']:
        q = s4.get(qk) or {}
        if q.get('qualification') or q.get('institution'):
            qual_lines.append(
                f"  - {_vv(q.get('qualification'))} | "
                f"{_vv(q.get('institution'))} | "
                f"{_vv(q.get('year_completed'))}"
            )
    # Also include NMBI/QQI numbers as qualification context
    nmbi_num = _vv(full_doc.get('nmbi_number') or s3.get('registration_number_pin') or '')
    qqi_num  = _vv(full_doc.get('qqi_number') or '')
    if nmbi_num and not any('nmbi' in l.lower() or 'registration' in l.lower() for l in qual_lines):
        qual_lines.append(f"  - NMBI Registration PIN: {nmbi_num}")
    if qqi_num and not any('qqi' in l.lower() for l in qual_lines):
        qual_lines.append(f"  - QQI Level 5 Certificate No: {qqi_num}")

    # ── Guaranteed fallback so EDUCATION section is NEVER empty ──────
    if not qual_lines:
        _role_lower = user_type.lower()
        if any(t in _role_lower for t in ('nurse', 'rgn', 'midwife', 'rnm', 'rn')):
            _inferred_q = 'Bachelor of Nursing Science (or equivalent)'
            _inferred_i = 'University College (based on nationality)'
        elif any(t in _role_lower for t in ('hca', 'healthcare assistant', 'health care assistant',
                                             'support worker', 'care assistant', 'care worker')):
            _inferred_q = 'QQI Level 5 in Healthcare Support'
            _inferred_i = 'College of Further Education'
        elif any(t in _role_lower for t in ('physio', 'occupational', 'speech', 'radiograph')):
            _inferred_q = 'BSc in Allied Health Sciences (or equivalent)'
            _inferred_i = 'Health Sciences University (based on nationality)'
        else:
            _inferred_q = 'Relevant Healthcare Qualification (see profile)'
            _inferred_i = 'Healthcare Training Institute'
        qual_lines.append(f"  - {_inferred_q} | {_inferred_i} | [year estimated from experience]")

    entries = [e for e in (s5.get('entries') or [])
               if e.get('employer') or e.get('position')]
    exp_lines = []
    for e in entries:
        exp_lines.append(
            f"  - {_vv(e.get('position'))} at {_vv(e.get('employer'))} "
            f"({_vv(e.get('from'))} - {_vv(e.get('to') or 'Present')})"
        )

    TLABELS = {
        'manual_handling': 'Manual Handling', 'cpr_bls': 'CPR / BLS',
        'fire_safety': 'Fire Safety',
        'infection_prevention_control': 'Infection Prevention & Control',
        'hand_hygiene': 'Hand Hygiene', 'safeguarding': 'Safeguarding',
        'children_first': 'Children First', 'cyber_security': 'Cyber Security',
        'dignity_at_work': 'Dignity at Work', 'open_disclosure': 'Open Disclosure',
        'mapa_pmav': 'MAPA / PMAV',
    }
    certs = [label for k, label in TLABELS.items() if s10.get(k)][:6]

    extracted_cv = _v(full_doc.get('extracted_cv') or '')
    has_extracted = (
        extracted_cv and
        not extracted_cv.startswith('[') and
        extracted_cv not in ('No doc found', '[no CV document found]', '')
    )

    # ── Background CV extraction (non-blocking) ───────────────────
    # If no extracted_cv, try to fetch it now in a quick attempt (10s timeout)
    # so we don't block the cron response past the 60s gateway timeout
    if not has_extracted:
        _base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
        _api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
        _app_country = os.environ.get('XN_APP_COUNTRY', '')
        _staff_email = _v(full_doc.get('email') or s1_f.get('email_address') or '')
        if _base_url and _staff_email:
            try:
                import requests as _rq2
                _hdrs = {"Api-Key": _api_key, "X-App-Country": _app_country,
                         "Content-Type": "application/json", "Accept": "application/json"}
                _r2 = _rq2.post(f"{_base_url}/ai/recruitments/user-document-list",
                                json={"email": _staff_email}, headers=_hdrs, timeout=15)
                if _r2.status_code == 405:
                    _r2 = _rq2.get(f"{_base_url}/ai/recruitments/user-document-list",
                                   params={"email": _staff_email}, headers=_hdrs, timeout=15)
                if _r2.status_code == 200:
                    _pd    = _r2.json()
                    _dlist = _pd.get('data') or []
                    if isinstance(_dlist, dict):
                        _dlist = _dlist.get('documents') or []
                    _cv_url = next(
                        (_d['url'] for _d in _dlist
                         if (_d.get('document_type_name') or '').strip() == 'Cv'
                         and _d.get('url')), None
                    )
                    if _cv_url:
                        try:
                            _et = _extract_text_from_url(
                                _cv_url,
                                {k: v for k, v in _hdrs.items() if k != 'Content-Type'}
                            )
                            if _et and not _et.startswith('['):
                                extracted_cv  = _et
                                has_extracted = True
                                _staffs_col().update_one(
                                    {"_id": full_doc['_id']},
                                    {"$set": {
                                        "extracted_cv":        _et,
                                        "extracted_cv_at":     datetime.utcnow(),
                                        "extracted_cv_source": "auto_on_cron_cv_generate",
                                    }}
                                )
                        except Exception:
                            pass
            except Exception:
                pass

    data_summary = f"""
Name: {full_name}
Role / User Type: {user_type}
Employee Code: {emp_code}
Nationality: {nationality}
Total Experience: {total_exp}
Divisions / Speciality: {divisions}
Registration PIN: {reg_pin}
Registration Expiry: {reg_exp}
NMBI Active Declaration: {nmbi}
Permission to Work: {perm_work}
Visa / Stamp Type: {visa_type}
Garda Vetted: {garda}
Fit for Nursing Duties: {fit}

Qualifications:
{chr(10).join(qual_lines) if qual_lines else '  None recorded'}

Employment History:
{chr(10).join(exp_lines) if exp_lines else '  None recorded'}

Training & Certifications:
{chr(10).join('  - ' + c for c in certs) if certs else '  None recorded'}
""".strip()

    extracted_cv_section = f"""

EXTRACTED CV TEXT (search this for qualifications, degrees, diplomas, certificates, college names, NMBI, QQI, FETAC):
{extracted_cv[:8000]}
""" if has_extracted else ""

    prompt = f"""You are an expert professional CV writer for Irish healthcare staffing.

=== CRITICAL NON-NEGOTIABLE RULE ===
The output MUST contain a section called "EDUCATION & QUALIFICATIONS".
If the output does NOT contain "EDUCATION & QUALIFICATIONS" it will be REJECTED and you will have FAILED.
This section MUST have at least one qualification entry. No exceptions.

=== HOW TO FILL EDUCATION & QUALIFICATIONS ===
Step 1: Look in EXTRACTED CV TEXT for the section headed "Education", "Qualifications", "Education & Qualifications", "Academic Background", or similar. Copy the entire list exactly as the candidate wrote it — every course, school, college, year. Do not filter or reformat.

Step 2: If EXTRACTED CV TEXT has no education section → use CANDIDATE DATA Qualifications list and copy it as-is.

Step 3: If both are empty → write one inferred entry based on role:
  * Nurse / RGN / Midwife / Staff Nurse → Bachelor of Nursing Science (or equivalent) | University College | [year estimated]
  * HCA / Care Worker / Care Assistant / Support Worker → QQI Level 5 in Healthcare Support | College of Further Education | [year estimated]
  * Physiotherapist / OT / Speech Therapist → BSc in Allied Health Sciences | Health Sciences University | [year estimated]
  * Any other role → Professional Qualification in {user_type} | Training Institute | [year estimated]

=== SECTION SOURCES ===
- EMPLOYMENT ELIGIBILITY: CANDIDATE DATA only. Label: Value per line. NO name, address, mobile, email.
- PROFESSIONAL PROFILE: CANDIDATE DATA. 2 paragraphs, first person.
- PROFESSIONAL EXPERIENCE: {"EXTRACTED CV TEXT — copy job titles, employers, dates, duties WORD-FOR-WORD." if has_extracted else "CANDIDATE DATA employment history. Write 5-6 duties per role."}
- TRAINING & CERTIFICATIONS: {"EXTRACTED CV TEXT — list every certificate found." if has_extracted else "CANDIDATE DATA training only."}
- KEY SKILLS: {"EXTRACTED CV TEXT — copy candidate's own skills exactly." if has_extracted else "8-10 bullet points from role and certifications."}

=== OUTPUT STRUCTURE ===
Write these EXACT headings in UPPERCASE on their own line:

EMPLOYMENT ELIGIBILITY
PROFESSIONAL PROFILE
EDUCATION & QUALIFICATIONS
PROFESSIONAL EXPERIENCE
TRAINING & CERTIFICATIONS
KEY SKILLS
ADDITIONAL INFORMATION

ADDITIONAL INFORMATION must contain ONLY:
Driving Licence: No
Own Transport: No

=== BEFORE YOU FINISH ===
CHECK: Does your output contain "EDUCATION & QUALIFICATIONS" with at least one entry?
If NO → go back and add it using Step 1/2/3 above. Do not output without it.

---
CANDIDATE DATA:
{data_summary}
{extracted_cv_section}
---
Output the CV text only. No markdown, no asterisks, no preamble.
"""
    # ── Run generation in background thread to avoid 504 ─────────────
    def _do_generate():
        try:
            from google import genai as google_genai
            _client  = google_genai.Client(api_key=gemini_key)
            _resp    = _client.models.generate_content(
                model='gemini-2.5-flash', contents=prompt
            )
            _cv_text = _resp.text.strip()

            # Build DOCX + upload GCS
            _docx    = _build_ai_cv_docx(full_doc, _cv_text)
            _sname   = (full_name or 'staff').replace(' ', '_').replace('/', '_')
            _fname   = f"{_sname}.docx"
            _blob    = f"cv/{_fname}"
            _gcs_upload(_blob, _docx,
                        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')

            ai_cvs_col.insert_one({
                "staff_id":      staff_id,
                "staff_name":    full_name,
                "employee_code": emp_code,
                "cv_text":       _cv_text,
                "cv_filename":   _fname,
                "gcs_blob":      _blob,
                "generated_at":  datetime.utcnow(),
            })

            # HSE push
            try:
                _push_hse_document_background(
                    staff_id_str=staff_id,
                    doc_type_key='cv',
                    gcs_blob=_blob,
                    filename=_fname,
                    user_type=full_doc.get('user_type'),
                )
            except Exception:
                pass

        except Exception as _bg_err:
            # Mark as attempted so cron moves on next call
            ai_cvs_col.insert_one({
                "staff_id":     staff_id,
                "staff_name":   full_name,
                "cv_text":      f"[bg error: {_bg_err}]",
                "gcs_blob":     '',
                "generated_at": datetime.utcnow(),
            })

    threading.Thread(target=_do_generate, daemon=True).start()

    return jsonify({
        "success":         True,
        "email":           email,
        "staff_name":      full_name,
        "cv_filename":     f"{(full_name or 'staff').replace(' ', '_').replace('/', '_')}.docx",
        "remaining_count": max(0, remaining_total - 1),
        "message": (
            f"CV generation started for {full_name} — "
            f"{max(0, remaining_total - 1)} staff still need a CV."
        ),
        "generated_at": datetime.utcnow().isoformat(),
    })



# ── Cron: Generate AI Interview Notes one staff at a time ─────────────

@admin_bp.route('/live-staffs/cron/generate-interview', methods=['GET', 'POST'])
def live_staff_cron_generate_interview():
    """
    Cron job — generates AI Interview Notes for ONE staff member per call.

    Logic:
      1. Find the first live_staffs record with no entry in live_staff_ai_interviews.
      2. Call Gemini 2.5 Flash to write the interview notes.
      3. Build DOCX, upload to GCS interviews/ folder, save metadata.
      4. Trigger HSE background upload.
      5. Return remaining_count.

    Protect with ?cron_key=<CRON_SECRET> env var.
    """

    # ── Auth ──────────────────────────────────────────────────────────
    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    gemini_key = os.environ.get('GEMINI_API_KEY', '')
    if not gemini_key:
        return jsonify({"success": False, "error": "GEMINI_API_KEY not set"}), 500

    staffs_col     = _staffs_col()
    interviews_col = _ai_interviews_col()

    # ── Find next staff without interview notes ───────────────────────
    existing_ids = set(
        str(r['staff_id'])
        for r in interviews_col.find({}, {"staff_id": 1})
        if r.get('staff_id')
    )

    staff = staffs_col.find_one(
        {"_id": {"$nin": [ObjectId(i) for i in existing_ids if len(i) == 24]}},
    )

    remaining_total = staffs_col.count_documents({}) - len(existing_ids)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff already have interview notes — nothing to do.",
            "remaining_count": 0,
            "processed":       None,
        })

    staff_id  = str(staff['_id'])
    s1        = staff.get('section_1_personal_details') or {}
    s3        = staff.get('section_3_professional_registration') or {}
    s5        = staff.get('section_5_employment_history') or {}
    s8        = staff.get('section_8_garda_vetting_police_clearance') or {}
    s9        = staff.get('section_9_occupational_health') or {}
    s10       = staff.get('section_10_mandatory_training') or {}
    visa      = s1.get('work_permit_visa_status') or {}

    full_name   = _v(s1.get('full_name') or '')
    email       = _v(staff.get('email') or '')
    user_type   = _v(staff.get('user_type') or 'Nurse')
    emp_code    = _v(staff.get('employee_code') or '')
    address     = _v(s1.get('address') or '')
    nationality = _v(s1.get('nationality') or '')
    reg_pin     = _v(s3.get('registration_number_pin') or '')
    visa_type   = _v(visa.get('visa_type') or '')
    divisions   = ', '.join(s3.get('divisions_registered_in') or [])
    total_exp   = _v(s5.get('total_experience') or '')
    nmbi        = 'Yes' if s3.get('nmbi_active_declaration') else 'No'
    garda       = 'Yes' if s8.get('garda_vetting_submitted') else 'No'
    bls         = 'Yes' if s10.get('cpr_bls') else 'No'
    manual      = 'Yes' if s10.get('manual_handling') else 'No'
    fit         = 'Yes' if s9.get('fit_for_nursing_duties') else 'No'

    entries = [e for e in (s5.get('entries') or [])
               if e.get('employer') or e.get('position')]
    exp_lines = []
    for e in entries[:5]:
        pos    = _v(e.get('position'))
        emp    = _v(e.get('employer'))
        d_from = _v(e.get('from'))
        d_to   = _v(e.get('to'))
        if pos or emp:
            exp_lines.append(f"  - {pos} at {emp} ({d_from} – {d_to or 'Present'})")

    TLABELS = {
        'manual_handling': 'Manual Handling', 'cpr_bls': 'BLS/CPR',
        'safeguarding': 'Safeguarding', 'fire_safety': 'Fire Safety',
        'infection_prevention_control': 'Infection Prevention & Control',
    }
    certs = [label for k, label in TLABELS.items() if s10.get(k)]

    # County from address
    county = ''
    if address:
        parts = [p.strip() for p in address.replace(',', ' ').split()]
        for p in parts:
            if p.lower().startswith('co.') or p.lower() == 'county':
                idx = parts.index(p)
                if idx + 1 < len(parts):
                    county = parts[idx + 1]
                break
        if not county and parts:
            county = parts[-1]

    data_summary = f"""
Name: {full_name}
Role / User Type: {user_type}
Address / Location: {address}
Nationality: {nationality}
Visa / Stamp Type: {visa_type}
NMBI Registration PIN: {reg_pin}
NMBI Registration Active: {nmbi}
Divisions / Speciality: {divisions}
Total Experience: {total_exp}
Garda Vetted: {garda}
BLS/CPR on file: {bls}
Manual Handling on file: {manual}
Fit for Duties: {fit}

Employment History:
{chr(10).join(exp_lines) if exp_lines else '  None recorded'}

Certifications on file: {', '.join(certs) if certs else 'None recorded'}
""".strip()

    prompt = f"""You are an experienced nursing recruitment consultant at Xpress Health, Ireland.

Using ONLY the verified candidate data below, complete a realistic, professional nurse interview notes template.
Answers must be written as if the candidate themselves just answered each question in a live phone/video interview.
Write naturally — conversational but professional. First person where appropriate ("I have", "I work", "I currently").

STRICT RULES — NO HALLUCINATION:
- Use ONLY the facts provided in CANDIDATE DATA. Do not invent employers, dates, locations, or qualifications.
- If data is missing for a field, write a realistic professional answer appropriate to their role and experience level without inventing specific names.
- Clinical question answers must be clinically appropriate for a {user_type}.
- Assessment scores: pick a random realistic score for each between 3.5 and 5.0 in 0.5 increments (e.g. 3.5/5, 4/5, 4.5/5, 5/5). Vary the three scores — do not give the same score to all three.
- Do NOT add any text outside the template structure below.

Output ONLY the completed template below — no preamble, no explanations, no markdown symbols:

---
Completed {user_type} Interview

Name: [full name]
Location: [county/city from address]
NMBI PIN: [registration pin or N/A]
Visa Status: [visa type]

Experience

1. Tell me about your nursing experience.
[Write a 4–6 sentence answer in first person describing their experience, speciality, and current/most recent role. Use only the data provided.]

2. How many years in Ireland?
[Write a realistic answer based on employment history dates. If Ireland-based work is evident, state it clearly.]

3. Acute, Nursing Home, Community, or Mental Health?
[Based on employment history, state the most relevant care setting.]

Clinical Questions

1. How would you manage a deteriorating patient?
[Write a clinically accurate 4–5 sentence answer appropriate for a {user_type}. Use recognised frameworks (ABCDE, NEWS2, ISBAR) where appropriate.]

2. What would you do if you witnessed a medication error?
[Write a clinically accurate 4–5 sentence answer covering patient safety, reporting, documentation, and prevention.]

Compliance
NMBI Registration: [Yes/No based on data]
BLS/CPR: [Yes/No based on data]
Manual Handling: [Yes/No based on data]
Garda Vetting: [Yes/No based on data]
References: Yes

Availability
Preferred counties: [county from address, or nearest city]
Day/Night/Both: Both
Earliest start date: Immediate

Assessment
Communication: [3.5/5 or 4/5 or 4.5/5 or 5/5 — vary randomly]
Clinical Knowledge: [3.5/5 or 4/5 or 4.5/5 or 5/5 — vary randomly]
Experience: [3.5/5 or 4/5 or 4.5/5 or 5/5 — vary randomly]
Suitable: Yes
---

CANDIDATE DATA (use ONLY this):
{data_summary}
"""

    if not full_name and not email:
        interviews_col.insert_one({
            "staff_id":       staff_id,
            "staff_name":     '',
            "interview_text": '[skipped — no name or email]',
            "gcs_blob":       '',
            "generated_at":   datetime.utcnow(),
        })
        return jsonify({
            "success":         True,
            "message":         f"Skipped {staff_id} — no name or email",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Call Gemini ───────────────────────────────────────────────────
    try:
        from google import genai as google_genai
        client         = google_genai.Client(api_key=gemini_key)
        response       = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt
        )
        interview_text = response.text.strip().strip('-').strip()

    except Exception as e:
        interviews_col.insert_one({
            "staff_id":       staff_id,
            "staff_name":     full_name,
            "interview_text": f"[Gemini error: {e}]",
            "gcs_blob":       '',
            "generated_at":   datetime.utcnow(),
        })
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"Gemini error: {e}",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Build DOCX and upload to GCS ─────────────────────────────────
    try:
        docx_bytes = _build_interview_docx(staff, interview_text)
        safe_name  = (full_name or 'staff').replace(' ', '_').replace('/', '_')
        filename   = f"Interview_{safe_name}_{staff_id}.docx"
        gcs_blob   = f"interviews/{filename}"
        _gcs_upload(
            gcs_blob, docx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
    except Exception as e:
        interviews_col.insert_one({
            "staff_id":       staff_id,
            "staff_name":     full_name,
            "interview_text": interview_text,
            "gcs_blob":       '',
            "generated_at":   datetime.utcnow(),
        })
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"DOCX/GCS error: {e}",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Save to MongoDB ───────────────────────────────────────────────
    interviews_col.insert_one({
        "staff_id":       staff_id,
        "staff_name":     full_name,
        "employee_code":  emp_code,
        "interview_text": interview_text,
        "filename":       filename,
        "gcs_blob":       gcs_blob,
        "generated_at":   datetime.utcnow(),
    })

    # ── Background HSE document push ──────────────────────────────────
    _push_hse_document_background(
        staff_id_str=staff_id,
        doc_type_key='interview',
        docx_bytes=docx_bytes,
        staff_name=full_name,
        mongo_id=staff_id,
        email=email,
    )

    return jsonify({
        "success":         True,
        "email":           email,
        "staff_name":      full_name,
        "filename":        filename,
        "remaining_count": max(0, remaining_total - 1),
        "message": (
            f"Interview notes generated for {full_name} — "
            f"{max(0, remaining_total - 1)} staff still need interview notes."
        ),
        "generated_at": datetime.utcnow().isoformat(),
    })



# ── Cron: Fetch signature + points one staff at a time ───────────────

@admin_bp.route('/live-staffs/cron/fetch-signature', methods=['GET', 'POST'])
def live_staff_cron_fetch_signature():
    """
    Cron job — fetches signature URL and points for ONE staff member per call.

    Logic:
      1. Find first live_staffs record where:
         - signature_url is missing/empty AND
         - staff_id (xn_staff_id) is set (needed for the API call)
      2. POST to {LIVE_STAFF_URL}/ai/recruitments/detail with {"_id": staff_id}
      3. Download signature image from signature_url
      4. Upload image to GCS under signatures/ folder
      5. Save GCS URL + points to live_staffs collection
      6. Return remaining_count

    Protect with ?cron_key=<CRON_SECRET> env var.
    """
    import requests as _req

    # ── Auth ──────────────────────────────────────────────────────────
    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')

    if not base_url:
        return jsonify({"success": False,
                        "error": "LIVE_STAFF_URL not set in environment"}), 500

    col = _staffs_col()

    # ── Count total needing signature ─────────────────────────────────
    needs_sig_query = {
        "$and": [
            # Missing signature_gcs_blob (not yet saved to GCS)
            {"$or": [
                {"signature_gcs_blob": {"$exists": False}},
                {"signature_gcs_blob": None},
                {"signature_gcs_blob": ""},
            ]},
            # Skip records already marked as no signature or errors
            {"$or": [
                {"signature_url": {"$exists": False}},
                {"signature_url": None},
                {"signature_url": ""},
                # Has a real URL (not a placeholder error string)
                {"signature_url": {"$not": {"$regex": "^\\["}}},
            ]},
            # Must have a staff_id to call the API
            {"$or": [
                {"staff_id":    {"$exists": True, "$ne": None, "$ne": ""}},
                {"xn_staff_id": {"$exists": True, "$ne": None, "$ne": ""}},
            ]},
        ]
    }
    remaining_total = col.count_documents(needs_sig_query)

    # ── Find next staff ───────────────────────────────────────────────
    staff = col.find_one(needs_sig_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff signatures already fetched — nothing to do.",
            "remaining_count": 0,
        })

    mongo_id  = str(staff['_id'])
    xn_id     = _v(staff.get('staff_id') or staff.get('xn_staff_id') or '')
    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or staff.get('email') or mongo_id)
    email     = _v(staff.get('email') or '')

    if not xn_id:
        # No staff_id — skip to avoid blocking queue
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {"signature_url": "[skipped — no staff_id]"}}
        )
        return jsonify({
            "success":         True,
            "message":         f"Skipped {full_name} — no staff_id available",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Call XN Portal detail API ─────────────────────────────────────
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    try:
        resp = _req.post(
            f"{base_url}/ai/recruitments/detail",
            json={"_id": xn_id},
            headers=api_headers,
            timeout=30,
        )
        # Retry with GET if POST not allowed
        if resp.status_code == 405:
            resp = _req.get(
                f"{base_url}/ai/recruitments/detail",
                params={"_id": xn_id},
                headers=api_headers,
                timeout=30,
            )
        resp.raise_for_status()
        data = resp.json()

    except Exception as api_err:
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {"signature_url": f"[API error: {api_err}]"}}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"API call failed: {api_err}",
            "remaining_count": max(0, remaining_total - 1),
        })

    if not data.get('success'):
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {"signature_url": f"[API error: {data.get('message', 'unknown')}]"}}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           data.get('message', 'API returned success=false'),
            "remaining_count": max(0, remaining_total - 1),
        })

    api_data      = data.get('data') or {}
    signature_url = (api_data.get('signature_url') or '').strip()
    points        = api_data.get('points')

    if not signature_url:
        # No signature — mark and move on
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {
                "signature_url": "[no signature found]",
                "points":        points,
            }}
        )
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "signature_saved": False,
            "points":          points,
            "message":         f"No signature URL returned for {full_name}",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Download signature image ───────────────────────────────────────
    try:
        img_resp = _req.get(signature_url, timeout=30)
        img_resp.raise_for_status()
        img_bytes = img_resp.content

        # Detect extension from URL or content-type
        content_type = img_resp.headers.get('Content-Type', 'image/png').lower()
        if 'jpeg' in content_type or 'jpg' in content_type:
            ext = 'jpg'
        elif 'webp' in content_type:
            ext = 'webp'
        else:
            ext = 'png'

        # Also try to get extension from URL
        url_path = signature_url.split('?')[0].lower()
        if url_path.endswith('.jpg') or url_path.endswith('.jpeg'):
            ext = 'jpg'
        elif url_path.endswith('.webp'):
            ext = 'webp'
        elif url_path.endswith('.png'):
            ext = 'png'

    except Exception as dl_err:
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {
                "signature_url": f"[download error: {dl_err}]",
                "points":        points,
            }}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"Signature download failed: {dl_err}",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Upload to GCS ─────────────────────────────────────────────────
    try:
        safe_name  = (full_name or 'staff').replace(' ', '_').replace('/', '_')
        gcs_blob   = f"signatures/{safe_name}_{mongo_id}.{ext}"
        gcs_ct     = f"image/{ext}"
        _gcs_upload(gcs_blob, img_bytes, content_type=gcs_ct)

        # Build a GCS public-style path to store (signed URL generated on demand)
        gcs_signature_path = gcs_blob

    except Exception as gcs_err:
        # GCS failed — still save original URL as fallback
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {
                "signature_url":      signature_url,
                "signature_gcs_blob": "",
                "points":             points,
                "signature_synced_at": datetime.utcnow(),
            }}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"GCS upload failed: {gcs_err} — original URL saved",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Save to live_staffs ───────────────────────────────────────────
    col.update_one(
        {"_id": staff['_id']},
        {"$set": {
            "signature_url":       signature_url,       # original CDN URL
            "signature_gcs_blob":  gcs_signature_path,  # our GCS copy
            "points":              points,
            "signature_synced_at": datetime.utcnow(),
        }}
    )

    return jsonify({
        "success":         True,
        "email":           email,
        "staff_name":      full_name,
        "signature_saved": True,
        "gcs_blob":        gcs_signature_path,
        "points":          points,
        "remaining_count": max(0, remaining_total - 1),
        "message": (
            f"Signature saved for {full_name} — "
            f"{max(0, remaining_total - 1)} staff still need signatures."
        ),
        "synced_at": datetime.utcnow().isoformat(),
    })



# ── Cron: Check points vs Last PointScale and flag mismatches ─────────

@admin_bp.route('/live-staffs/cron/check-points', methods=['GET', 'POST'])
def live_staff_cron_check_points():
    """
    Cron job — checks one staff member's current points against their
    Last PointScale from the spreadsheet and flags mismatches.

    Logic:
      1. Find first live_staffs record where points_checked is not True.
      2. Call XN Portal detail API to get current points.
      3. Compare with last_point_scale stored in live_staffs.
      4. Save: current_points, points_mismatch (bool), points_checked (True).

    Import last_point_scale into live_staffs first using the import route
    or via the spreadsheet (email + last_point_scale fields).

    Protect with ?cron_key=<CRON_SECRET> env var.
    """
    import requests as _req

    # ── Auth ──────────────────────────────────────────────────────────
    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')

    if not base_url:
        return jsonify({"success": False,
                        "error": "LIVE_STAFF_URL not set"}), 500

    col = _staffs_col()

    # Count remaining
    remaining_total = col.count_documents(
        {"$or": [
            {"points_checked": {"$exists": False}},
            {"points_checked": False},
            {"points_checked": None},
        ]}
    )

    # Find next unchecked
    staff = col.find_one(
        {"$or": [
            {"points_checked": {"$exists": False}},
            {"points_checked": False},
            {"points_checked": None},
        ]}
    )

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff points checked — nothing to do.",
            "remaining_count": 0,
        })

    mongo_id  = str(staff['_id'])
    xn_id     = _v(staff.get('staff_id') or staff.get('xn_staff_id') or '')
    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or staff.get('email') or mongo_id)
    email     = _v(staff.get('email') or '')
    last_ps   = staff.get('last_point_scale')   # stored from spreadsheet import

    if not xn_id:
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {
                "points_checked": True,
                "points_mismatch": None,
                "points_check_note": "skipped — no staff_id",
            }}
        )
        return jsonify({
            "success":         True,
            "message":         f"Skipped {full_name} — no staff_id",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Call XN Portal detail API ──────────────────────────────────────
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    try:
        resp = _req.post(
            f"{base_url}/ai/recruitments/detail",
            json={"_id": xn_id},
            headers=api_headers,
            timeout=30,
        )
        if resp.status_code == 405:
            resp = _req.get(
                f"{base_url}/ai/recruitments/detail",
                params={"_id": xn_id},
                headers=api_headers,
                timeout=30,
            )
        resp.raise_for_status()
        data = resp.json()

    except Exception as api_err:
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {
                "points_checked":    True,
                "points_mismatch":   None,
                "points_check_note": f"API error: {api_err}",
            }}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"API error: {api_err}",
            "remaining_count": max(0, remaining_total - 1),
        })

    if not data.get('success'):
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {
                "points_checked":    True,
                "points_mismatch":   None,
                "points_check_note": f"API error: {data.get('message')}",
            }}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           data.get('message', 'API returned success=false'),
            "remaining_count": max(0, remaining_total - 1),
        })

    api_data       = data.get('data') or {}
    current_points = api_data.get('points')

    # ── Compare ────────────────────────────────────────────────────────
    mismatch = None
    if last_ps is not None and current_points is not None:
        try:
            mismatch = float(current_points) != float(last_ps)
        except (TypeError, ValueError):
            mismatch = None

    col.update_one(
        {"_id": staff['_id']},
        {"$set": {
            "points":             current_points,
            "last_point_scale":   last_ps,
            "points_mismatch":    mismatch,
            "points_checked":     True,
            "points_checked_at":  datetime.utcnow(),
        }}
    )

    return jsonify({
        "success":          True,
        "email":            email,
        "staff_name":       full_name,
        "last_point_scale": last_ps,
        "current_points":   current_points,
        "mismatch":         mismatch,
        "remaining_count":  max(0, remaining_total - 1),
        "message": (
            f"{'⚠ MISMATCH' if mismatch else '✓ Match'} — "
            f"{full_name}: last={last_ps}, current={current_points} — "
            f"{max(0, remaining_total - 1)} remaining."
        ),
        "checked_at": datetime.utcnow().isoformat(),
    })



# ── Cron: Generate Application Form one staff at a time ───────────────

@admin_bp.route('/live-staffs/cron/generate-appform', methods=['GET', 'POST'])
def live_staff_cron_generate_appform():
    """
    Cron job — generates an Application Form for ONE staff member per call.

    Logic:
      1. Find first live_staffs record with no entry in live_staff_ai_appforms.
      2. If staff has a signature_gcs_blob, download it from GCS.
      3. Build the Application Form DOCX (with real signature image if available).
      4. Upload to GCS appforms/ folder, save metadata to MongoDB.
      5. Fire HSE document background upload.
      6. Return remaining_count.

    Protect with ?cron_key=<CRON_SECRET> env var.
    """

    # ── Auth ──────────────────────────────────────────────────────────
    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    staffs_col   = _staffs_col()
    appforms_col = _ai_appforms_col()

    # ── Find next staff without an application form ───────────────────
    existing_ids = set(
        str(r['staff_id'])
        for r in appforms_col.find({}, {"staff_id": 1})
        if r.get('staff_id')
    )

    staff = staffs_col.find_one(
        {"_id": {"$nin": [ObjectId(i) for i in existing_ids if len(i) == 24]}}
    )

    remaining_total = staffs_col.count_documents({}) - len(existing_ids)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff already have application forms — nothing to do.",
            "remaining_count": 0,
        })

    staff_id  = str(staff['_id'])
    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or '')
    emp_code  = _v(staff.get('employee_code') or '')

    if not full_name and not email:
        appforms_col.insert_one({
            "staff_id":     staff_id,
            "staff_name":   '',
            "gcs_blob":     '',
            "generated_at": datetime.utcnow(),
        })
        return jsonify({
            "success":         True,
            "message":         f"Skipped {staff_id} — no name or email",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Download signature from GCS if available ──────────────────────
    signature_bytes = None
    sig_blob        = _v(staff.get('signature_gcs_blob') or '')
    if sig_blob:
        try:
            signature_bytes = _gcs_download(sig_blob)
        except Exception as sig_err:
            # Non-fatal — form will have blank signature line
            signature_bytes = None

    # ── Build Application Form DOCX ───────────────────────────────────
    try:
        docx_bytes = _build_appform_docx(staff, signature_bytes=signature_bytes)
        safe_name  = (full_name or 'staff').replace(' ', '_').replace('/', '_')
        filename   = f"AppForm_{safe_name}.docx"
        gcs_blob   = f"appforms/{filename}"
        _gcs_upload(
            gcs_blob, docx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
    except Exception as e:
        appforms_col.insert_one({
            "staff_id":     staff_id,
            "staff_name":   full_name,
            "gcs_blob":     '',
            "generated_at": datetime.utcnow(),
        })
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"DOCX/GCS error: {e}",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Save to MongoDB ───────────────────────────────────────────────
    appforms_col.insert_one({
        "staff_id":      staff_id,
        "staff_name":    full_name,
        "employee_code": emp_code,
        "filename":      filename,
        "gcs_blob":      gcs_blob,
        "has_signature": bool(signature_bytes),
        "generated_at":  datetime.utcnow(),
    })

    # ── Background HSE document push ──────────────────────────────────
    _push_hse_document_background(
        staff_id_str=staff_id,
        doc_type_key='appform',
        docx_bytes=docx_bytes,
        staff_name=full_name,
        mongo_id=staff_id,
        email=email,
    )

    return jsonify({
        "success":         True,
        "email":           email,
        "staff_name":      full_name,
        "filename":        filename,
        "has_signature":   bool(signature_bytes),
        "remaining_count": max(0, remaining_total - 1),
        "message": (
            f"Application form generated for {full_name} "
            f"({'with' if signature_bytes else 'without'} signature) — "
            f"{max(0, remaining_total - 1)} staff still need forms."
        ),
        "generated_at": datetime.utcnow().isoformat(),
    })



# ── Cron: Sync Garda Vetting & Police Clearance one staff at a time ───

@admin_bp.route('/live-staffs/cron/sync-vetting', methods=['GET', 'POST'])
def live_staff_cron_sync_vetting():
    """
    Cron job — processes ONE staff member per call.

    For each staff member calls the XN Portal document-list API and checks:
      - document_type_name == "Garda Vetting Document"
        * status approved  → garda_vetting = 1
        * status rejected/pending → garda_vetting = 0
        * expiry_date passed today → garda_vetting_expired = 1 else 0
      - document_type_name == "Police Clearance Certificate ( From Country Of Birth )"
        * same logic → police_clearance and police_clearance_expired

    Both checked in a single API call per staff member.
    Run one staff at a time until all processed (remaining_count == 0).

    Protect with ?cron_key=<CRON_SECRET> env var.
    """
    import requests as _req
    from datetime import date as _date

    # ── Auth ──────────────────────────────────────────────────────────
    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')

    if not base_url:
        return jsonify({"success": False,
                        "error": "LIVE_STAFF_URL not set"}), 500

    col = _staffs_col()

    # ── Find next staff without vetting synced ────────────────────────
    pending_query = {
        "$or": [
            {"garda_vetting_synced": {"$exists": False}},
            {"garda_vetting_synced": False},
            {"garda_vetting_synced": None},
        ]
    }
    remaining_total = col.count_documents(pending_query)

    staff = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff vetting documents synced — nothing to do.",
            "remaining_count": 0,
        })

    mongo_id  = str(staff['_id'])
    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or '')

    if not email:
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {"garda_vetting_synced": True,
                      "garda_vetting_sync_note": "skipped — no email"}}
        )
        return jsonify({
            "success":         True,
            "message":         f"Skipped {mongo_id} — no email",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Call XN Portal API ────────────────────────────────────────────
    endpoint   = f"{base_url}/ai/recruitments/user-document-list"
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    try:
        resp = _req.post(
            endpoint,
            json={"email": email},
            headers=api_headers,
            timeout=30,
        )
        if resp.status_code == 405:
            resp = _req.get(
                endpoint,
                params={"email": email},
                headers=api_headers,
                timeout=30,
            )
        resp.raise_for_status()
        data = resp.json()
    except Exception as api_err:
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {"garda_vetting_synced": True,
                      "garda_vetting_sync_note": f"API error: {api_err}"}}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"API error: {api_err}",
            "remaining_count": max(0, remaining_total - 1),
        })

    if not data.get('success'):
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {"garda_vetting_synced": True,
                      "garda_vetting_sync_note": f"API error: {data.get('message')}"}}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           data.get('message', 'API returned success=false'),
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Parse documents ────────────────────────────────────────────────
    api_data  = data.get('data') or {}
    if isinstance(api_data, list):
        documents = api_data
    elif isinstance(api_data, dict):
        documents = api_data.get('documents') or []
    else:
        documents = []

    today = _date.today()

    def _parse_expiry(expiry_str):
        """Parse expiry date string, return date object or None."""
        if not expiry_str:
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
            try:
                from datetime import datetime as _dt
                return _dt.strptime(str(expiry_str)[:10], fmt).date()
            except ValueError:
                continue
        return None

    def _check_doc(doc_type_name):
        """
        Find a document by type name and return (status_int, expired_int).
        status: 1=approved, 0=rejected/pending/missing
        expired: 1=expired, 0=valid or no expiry
        """
        found = [
            d for d in documents
            if (d.get('document_type_name') or '').strip().lower()
            == doc_type_name.strip().lower()
        ]
        if not found:
            return None, None   # document not present at all

        # Use the most recent / approved one if multiple
        doc = found[0]
        for d in found:
            if (d.get('status') or '').lower() == 'approved':
                doc = d
                break

        status_str = (doc.get('status') or '').strip().lower()
        approved   = 1 if status_str == 'approved' else 0

        expiry_date = _parse_expiry(doc.get('expiry_date'))
        if expiry_date:
            expired = 1 if expiry_date < today else 0
        else:
            expired = 0

        return approved, expired

    garda_status,  garda_expired  = _check_doc('Garda Vetting Document')
    police_status, police_expired = _check_doc('Police Clearance Certificate ( From Country Of Birth )')

    # ── Save to live_staffs ───────────────────────────────────────────
    update_fields = {
        "garda_vetting_synced":  True,
        "garda_vetting_synced_at": datetime.utcnow(),
    }

    if garda_status is not None:
        update_fields["garda_vetting"]         = garda_status
        update_fields["garda_vetting_expired"] = garda_expired
    else:
        update_fields["garda_vetting"]         = 0
        update_fields["garda_vetting_expired"] = 0

    if police_status is not None:
        update_fields["police_clearance"]         = police_status
        update_fields["police_clearance_expired"] = police_expired
    else:
        update_fields["police_clearance"]         = 0
        update_fields["police_clearance_expired"] = 0

    col.update_one({"_id": staff['_id']}, {"$set": update_fields})

    return jsonify({
        "success":               True,
        "email":                 email,
        "staff_name":            full_name,
        "garda_vetting":         update_fields.get("garda_vetting"),
        "garda_vetting_expired": update_fields.get("garda_vetting_expired"),
        "police_clearance":         update_fields.get("police_clearance"),
        "police_clearance_expired": update_fields.get("police_clearance_expired"),
        "remaining_count":       max(0, remaining_total - 1),
        "message": (
            f"Vetting synced for {full_name} — "
            f"{max(0, remaining_total - 1)} staff remaining."
        ),
        "synced_at": datetime.utcnow().isoformat(),
    })



# ── Cron: Analyse experience one staff at a time ──────────────────────

@admin_bp.route('/live-staffs/cron/analyse-experience', methods=['GET', 'POST'])
def live_staff_cron_analyse_experience():
    """
    Cron job — analyses experience for ONE staff member per call.

    Gemini runs in a background thread so the HTTP response returns
    immediately — prevents 504 gateway timeouts on slow Gemini calls.

    Logic:
      1. Find first record where experience_analysed_at is missing AND
         experience_processing is not True (not already in-flight).
      2. Mark experience_processing = True immediately.
      3. Fire Gemini analysis in background thread.
      4. Return 200 right away with remaining_count.

    Protect with ?cron_key=<CRON_SECRET> env var.
    """
    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    gemini_key = os.environ.get('GEMINI_API_KEY', '')
    if not gemini_key:
        return jsonify({"success": False, "error": "GEMINI_API_KEY not set"}), 500

    col = _staffs_col()

    pending_query = {
        "$and": [
            {"$or": [
                {"experience_analysed_at": {"$exists": False}},
                {"experience_analysed_at": None},
            ]},
            # Skip records currently being processed in background
            {"$or": [
                {"experience_processing": {"$exists": False}},
                {"experience_processing": False},
                {"experience_processing": None},
            ]},
        ]
    }
    remaining_total = col.count_documents(pending_query)

    staff = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff experience already analysed — nothing to do.",
            "remaining_count": 0,
        })

    staff_id  = str(staff['_id'])
    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or '')
    user_type = _v(staff.get('user_type') or '')

    extracted_cv = _v(staff.get('extracted_cv') or '')
    has_cv_text  = (
        extracted_cv and
        not extracted_cv.startswith('[') and
        extracted_cv != 'No doc found'
    )
    s5           = staff.get('section_5_employment_history') or {}
    total_exp_db = _v(s5.get('total_experience') or '')

    if not has_cv_text and not total_exp_db:
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {
                "experience_years":        0,
                "experience_months":       0,
                "experience_total_months": 0,
                "experience_note":         "no CV text or experience data available",
                "experience_source":       "none",
                "experience_analysed_at":  datetime.utcnow(),
                "experience_processing":   False,
            }}
        )
        return jsonify({
            "success":         True,
            "staff_name":      full_name,
            "email":           email,
            "message":         f"Skipped {full_name} — no experience data",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Mark as processing then fire Gemini in background ────────────
    # This prevents 504 gateway timeouts — returns 200 immediately
    col.update_one(
        {"_id": staff['_id']},
        {"$set": {"experience_processing": True}}
    )

    def _run_gemini(staff_id_str, user_type_, has_cv_, extracted_cv_,
                    total_exp_db_, gemini_key_):
        from google import genai as google_genai

        _col = _staffs_col()

        ut_lower = (user_type_ or '').lower()
        if 'nurse' in ut_lower or 'nursing' in ut_lower:
            role_rule = (
                "IMPORTANT — Count ONLY nursing-related work experience, regardless of the country where it was gained. "
                "This includes: Registered Nurse, Staff Nurse, Clinical Nurse, ICU Nurse, Theatre Nurse, "
                "Community Nurse, Mental Health Nurse, Nursing Home Nurse, or any role with Nurse or Nursing in the title. "
                "DO NOT count non-nursing roles such as healthcare assistant, carer, support worker, admin, or any other role."
            )
        elif 'hca' in ut_lower or 'healthcare assistant' in ut_lower or 'health care assistant' in ut_lower:
            role_rule = (
                "IMPORTANT — Count ONLY Healthcare Assistant (HCA) work experience, regardless of the country where it was gained. "
                "This includes: Healthcare Assistant, HCA, Care Assistant, Care Worker, Support Worker in a clinical/care setting, "
                "or any role with Healthcare Assistant or HCA in the title. "
                "DO NOT count nursing roles (Registered Nurse, Staff Nurse, etc.) or non-care roles such as admin, retail, or hospitality."
            )
        else:
            role_rule = (
                "Count only direct healthcare or care-related work experience, regardless of country. "
                "Exclude non-healthcare roles such as admin, retail, hospitality, or general support roles "
                "unless they are clearly in a clinical or care setting."
            )

        try:
            if has_cv_:
                source = 'extracted_cv'
                prompt = f"""You are a professional CV analyser specialising in Irish healthcare staffing.

Candidate role: {user_type_}

Read the CV text below and calculate the candidate's TOTAL relevant work experience.

{role_rule}

Calculation Rules:
- Only count roles that match the role filter above.
- Experience gained in ANY country counts — not just Ireland.
- If a matching role has no end date, assume it is still ongoing (use today's date to calculate).
- If two matching roles overlap in time, count the overlapping period only once.
- Ignore all non-matching roles entirely — do not add them.
- Return ONLY a JSON object with these exact keys — nothing else, no markdown, no explanation:
  {{"years": <integer>, "months": <integer 0-11>, "total_months": <integer>, "note": "<one sentence summary of which roles were counted>"}}

CV TEXT:
{extracted_cv_[:10000]}
"""
            else:
                source = 'section_5'
                prompt = f"""You are a professional CV analyser specialising in Irish healthcare staffing.

Candidate role: {user_type_}

The candidate's total experience is described as: "{total_exp_db_}"

{role_rule}

Extract the relevant years and months from this description, applying the role filter above.
Return ONLY a JSON object — nothing else, no markdown:
{{"years": <integer>, "months": <integer 0-11>, "total_months": <integer>, "note": "<one sentence summary>"}}
"""

            client   = google_genai.Client(api_key=gemini_key_)
            response = client.models.generate_content(
                model='gemini-2.5-flash', contents=prompt
            )
            raw = (response.text or '').strip()
            raw = _re.sub(r'^```(?:json)?\s*', '', raw, flags=_re.MULTILINE)
            raw = _re.sub(r'```\s*$', '', raw, flags=_re.MULTILINE).strip()

            result       = _json.loads(raw)
            years        = int(result.get('years', 0) or 0)
            months       = int(result.get('months', 0) or 0)
            total_months = int(result.get('total_months', 0) or 0)
            note         = str(result.get('note', '') or '').strip()
            if total_months == 0:
                total_months = years * 12 + months

            _col.update_one(
                {"_id": ObjectId(staff_id_str)},
                {"$set": {
                    "experience_years":        years,
                    "experience_months":       months,
                    "experience_total_months": total_months,
                    "experience_note":         note,
                    "experience_source":       source,
                    "experience_analysed_at":  datetime.utcnow(),
                    "experience_processing":   False,
                }}
            )

        except Exception as e:
            _col.update_one(
                {"_id": ObjectId(staff_id_str)},
                {"$set": {
                    "experience_analysed_at":  datetime.utcnow(),
                    "experience_processing":   False,
                    "experience_note":         f"error: {e}",
                }}
            )

    threading.Thread(
        target=_run_gemini,
        args=(staff_id, user_type, has_cv_text, extracted_cv,
              total_exp_db, gemini_key),
        daemon=True,
    ).start()

    return jsonify({
        "success":         True,
        "staff_name":      full_name,
        "email":           email,
        "status":          "processing",
        "remaining_count": max(0, remaining_total - 1),
        "message": (
            f"Analysis started for {full_name} in background — "
            f"{max(0, remaining_total - 1)} staff remaining. "
            "Result will be saved to DB when complete."
        ),
    })



# ── Cron: Extract formatted experience entries one staff at a time ────

@admin_bp.route('/live-staffs/cron/extract-experience-list', methods=['GET', 'POST'])
def live_staff_cron_extract_experience_list():
    """
    Cron job — extracts a formatted list of relevant experience entries
    for ONE staff member per call using Gemini AI.

    For each matching role extracts:
      "Job Title  Date Range  Employer/Location"
    Only includes roles relevant to live_staffs.user_type
    (nurse roles for nurses, HCA roles for healthcare assistants).

    Saves to live_staffs.experience_list (array of strings).
    Background thread to avoid 504 timeouts.

    Protect with ?cron_key=<CRON_SECRET>
    """
    import requests as _req
    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    gemini_key = os.environ.get('GEMINI_API_KEY', '')
    if not gemini_key:
        return jsonify({"success": False, "error": "GEMINI_API_KEY not set"}), 500

    col = _staffs_col()

    pending_query = {
        "$and": [
            {"$or": [
                {"experience_list_at": {"$exists": False}},
                {"experience_list_at": None},
            ]},
            {"$or": [
                {"experience_list_processing": {"$exists": False}},
                {"experience_list_processing": False},
                {"experience_list_processing": None},
            ]},
        ]
    }
    remaining_total = col.count_documents(pending_query)
    staff = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff experience lists extracted — nothing to do.",
            "remaining_count": 0,
        })

    staff_id  = str(staff['_id'])
    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or '')
    user_type = _v(staff.get('user_type') or '')

    extracted_cv = _v(staff.get('extracted_cv') or '')
    has_cv_text  = (
        extracted_cv and
        not extracted_cv.startswith('[') and
        extracted_cv != 'No doc found'
    )

    if not has_cv_text:
        col.update_one(
            {"_id": staff['_id']},
            {"$set": {
                "experience_list":            [],
                "experience_list_at":         datetime.utcnow(),
                "experience_list_processing": False,
            }}
        )
        return jsonify({
            "success":         True,
            "staff_name":      full_name,
            "email":           email,
            "message":         f"Skipped {full_name} — no extracted CV text",
            "remaining_count": max(0, remaining_total - 1),
        })

    # Mark as processing — return immediately
    col.update_one(
        {"_id": staff['_id']},
        {"$set": {"experience_list_processing": True}}
    )

    def _run(staff_id_str, user_type_, extracted_cv_, gemini_key_):
        from google import genai as google_genai

        _col = _staffs_col()
        ut_lower = (user_type_ or '').lower()

        if 'nurse' in ut_lower or 'nursing' in ut_lower:
            role_filter = (
                "Extract ONLY nursing roles: Registered Nurse, Staff Nurse, Clinical Nurse, "
                "ICU Nurse, Theatre Nurse, Community Nurse, Mental Health Nurse, Nursing Home Nurse, "
                "or any role with Nurse or Nursing in the title. "
                "IGNORE Healthcare Assistant, HCA, Carer, Support Worker, admin, or any non-nursing role."
            )
        elif 'hca' in ut_lower or 'healthcare assistant' in ut_lower:
            role_filter = (
                "Extract ONLY Healthcare Assistant roles: Healthcare Assistant, HCA, "
                "Care Assistant, Care Worker, Support Worker in a clinical/care setting, "
                "Domiciliary Care Assistant, or any role with Healthcare Assistant or HCA in the title. "
                "IGNORE nursing roles (Registered Nurse, Staff Nurse etc.) and non-care roles."
            )
        else:
            role_filter = (
                "Extract only healthcare or care-related work experience roles. "
                "Ignore admin, retail, hospitality, or non-clinical roles."
            )

        prompt = f"""You are a professional CV analyser specialising in Irish healthcare staffing.

Candidate role: {user_type_}

Read the CV text below and extract ALL relevant work experience entries.

{role_filter}

For each matching role, return it as a single formatted string exactly like this example:
"Healthcare Assistant  Jan 2022 – Present  HSE Sligo Leitrim, Disabilities Services"

Format rules:
- Job title first
- Then date range (use the dates as written in the CV, e.g. "Jan 2022 – Present" or "2019 – 2021")
- Then employer and location/department if available
- Separate each part with two spaces
- If no end date, write "Present"
- List ALL matching roles, most recent first

Return ONLY a JSON array of strings — nothing else, no markdown, no explanation:
["role 1 formatted string", "role 2 formatted string", ...]

If no matching roles found, return an empty array: []

CV TEXT:
{extracted_cv_[:10000]}
"""

        try:
            client   = google_genai.Client(api_key=gemini_key_)
            response = client.models.generate_content(
                model='gemini-2.5-flash', contents=prompt
            )
            raw = (response.text or '').strip()
            raw = _re.sub(r'^```(?:json)?\s*', '', raw, flags=_re.MULTILINE)
            raw = _re.sub(r'```\s*$', '', raw, flags=_re.MULTILINE).strip()

            result = _cjson.loads(raw)
            if not isinstance(result, list):
                result = []
            # Clean each entry
            result = [str(r).strip() for r in result if str(r).strip()]

            _col.update_one(
                {"_id": ObjectId(staff_id_str)},
                {"$set": {
                    "experience_list":            result,
                    "experience_list_at":         datetime.utcnow(),
                    "experience_list_processing": False,
                }}
            )
        except Exception as e:
            _col.update_one(
                {"_id": ObjectId(staff_id_str)},
                {"$set": {
                    "experience_list":            [],
                    "experience_list_at":         datetime.utcnow(),
                    "experience_list_processing": False,
                    "experience_list_error":      str(e),
                }}
            )

    threading.Thread(
        target=_run,
        args=(staff_id, user_type, extracted_cv, gemini_key),
        daemon=True,
    ).start()

    return jsonify({
        "success":         True,
        "staff_name":      full_name,
        "email":           email,
        "status":          "processing",
        "remaining_count": max(0, remaining_total - 1),
        "message": (
            f"Experience list extraction started for {full_name} in background — "
            f"{max(0, remaining_total - 1)} staff remaining."
        ),
    })


# ── Export experience list to Excel ──────────────────────────────────

@admin_bp.route('/live-staffs/export/experience-xlsx')
@admin_required
def live_staff_export_experience_xlsx():
    """
    Export staff experience list to Excel.
    Columns: Sno | Name | Email | Experience 1 | Experience 2 | Experience 3 | ...

    The number of experience columns expands automatically based on
    the maximum number of entries any staff member has.

    GET /admin/live-staffs/export/experience-xlsx
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as _io

        docs = list(_staffs_col().find(
            {},
            {"section_1_personal_details": 1, "email": 1,
             "user_type": 1, "experience_list": 1}
        ))

        # Sort by name
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        # Find max experience entries across all staff
        max_exp = max(
            (len(d.get('experience_list') or []) for d in docs),
            default=1
        )
        max_exp = max(max_exp, 3)  # minimum 3 columns

        # ── Styles ────────────────────────────────────────────────────
        NAVY = '1B3A6B'; GREEN = '2E9E44'; WHITE = 'FFFFFF'; ALT = 'EFF6FF'

        h_font  = Font(name='Arial', bold=True, color=WHITE, size=10)
        h_fill  = PatternFill('solid', start_color=NAVY, end_color=NAVY)
        h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        b_font  = Font(name='Arial', size=9)
        l_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c_align = Alignment(horizontal='center', vertical='top')
        thin    = Side(style='thin', color='CCCCCC')
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Experience List'

        # ── Headers ───────────────────────────────────────────────────
        base_headers = ['Sno', 'Name', 'Email', 'User Type']
        exp_headers  = [f'Experience {i+1}' for i in range(max_exp)]
        headers      = base_headers + exp_headers
        col_widths   = [5, 30, 38, 20] + [45] * max_exp

        for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font      = h_font
            cell.fill      = h_fill
            cell.alignment = h_align
            cell.border    = Border(
                left=thin, right=thin, top=thin,
                bottom=Side(style='medium', color=GREEN)
            )
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(headers)).column_letter}1'

        # ── Data rows ─────────────────────────────────────────────────
        for ri, doc in enumerate(docs, start=2):
            s1        = doc.get('section_1_personal_details') or {}
            name      = _v(s1.get('full_name') or '')
            email     = _v(doc.get('email') or '')
            user_type = _v(doc.get('user_type') or '')
            exp_list  = doc.get('experience_list') or []

            alt_fill = PatternFill('solid', start_color=ALT, end_color=ALT) if ri % 2 == 0 else None

            row_data = [ri - 1, name, email, user_type] + exp_list

            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=ri, column=ci, value=val or '')
                cell.font      = b_font
                cell.alignment = c_align if ci in (1,) else l_align
                cell.border    = border
                if alt_fill:
                    cell.fill = alt_fill

            # Fill remaining experience columns with empty
            for ci in range(len(row_data) + 1, len(headers) + 1):
                cell = ws.cell(row=ri, column=ci, value='')
                cell.font      = b_font
                cell.border    = border
                if alt_fill:
                    cell.fill = alt_fill

            ws.row_dimensions[ri].height = max(18, 15 * max(1, len(exp_list)))

        # Summary row
        ws.cell(row=len(docs) + 2, column=1,
                value=f'Total: {len(docs)} staff').font = Font(name='Arial', bold=True, size=9)

        buf = _io.BytesIO()
        wb.save(buf)
        date_str = datetime.utcnow().strftime('%Y%m%d')

        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="experience_list_{date_str}.xlsx"'}
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500



# ── Cron: Sync Passport/ID card one staff at a time ──────────────────

@admin_bp.route('/live-staffs/cron/sync-passport', methods=['GET', 'POST'])
def live_staff_cron_sync_passport():
    """
    Cron job — processes ONE staff member per call.

    Logic:
      1. Find first live_staffs record where passport_extracted is missing/empty.
      2. Call XN Portal user-document-list API with staff email.
      3. Find document where document_type_name == "Passport/id card".
      4. Download the document and extract passport data via Gemini AI.
      5. Save extracted data + passport_id to live_staffs.
      6. Return remaining_count.

    Protect with ?cron_key=<CRON_SECRET> env var.
    """
    import requests as _req

    # ── Auth ──────────────────────────────────────────────────────────
    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')
    gemini_key  = os.environ.get('GEMINI_API_KEY', '')

    if not base_url:
        return jsonify({"success": False,
                        "error": "LIVE_STAFF_URL not set"}), 500
    if not gemini_key:
        return jsonify({"success": False,
                        "error": "GEMINI_API_KEY not set"}), 500

    col = _staffs_col()

    # ── Find next staff to process ────────────────────────────────────
    # Picks up:
    #   1. Never fetched (passport_fetched missing/False)
    #   2. Fetched but passport_id is still empty — retry with passport_id_card doc type
    pending_query = {
        "$or": [
            {"passport_fetched": {"$exists": False}},
            {"passport_fetched": False},
            {"passport_fetched": None},
        ]
    }

    remaining_total = col.count_documents(pending_query)
    staff           = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff passports already fetched — nothing to do.",
            "remaining_count": 0,
        })

    email = _v(
        (staff.get('section_1_personal_details') or {}).get('email_address') or
        staff.get('email') or ''
    )
    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')

    def _mark_done(fields):
        fields["passport_fetched"]    = True
        fields["passport_fetched_at"] = datetime.utcnow()
        col.update_one({"_id": staff['_id']}, {"$set": fields})

    if not email:
        col.update_one(
            {"_id": staff["_id"]},
            {"$set": {
                "passport_extracted":  "[skipped — no email]",
                "passport_extracted_at": datetime.utcnow(),
                "passport_fetched":    True,
            }}
        )
        return jsonify({
            "success":         True,
            "message":         "Skipped — no email",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Call XN Portal API ────────────────────────────────────────────
    endpoint    = f"{base_url}/ai/recruitments/user-document-list"
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    try:
        resp = _req.post(endpoint, json={"email": email},
                         headers=api_headers, timeout=30)
        if resp.status_code == 405:
            resp = _req.get(endpoint, params={"email": email},
                            headers=api_headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as api_err:
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"API error: {api_err}",
            "remaining_count": remaining_total,
        })

    if not data.get('success'):
        col.update_one(
            {"_id": staff["_id"]},
            {"$set": {
                "passport_extracted":    f"[API error: {data.get('message')}]",
                "passport_extracted_at": datetime.utcnow(),
                "passport_fetched":      True,
            }}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           data.get('message', 'API error'),
            "remaining_count": max(0, remaining_total - 1),
        })

    api_data  = data.get('data')
    if isinstance(api_data, list):
        documents = api_data
    elif isinstance(api_data, dict):
        documents = api_data.get('documents') or []
    else:
        documents = []

    if not documents:
        col.update_one(
            {"_id": staff["_id"]},
            {"$set": {
                "passport_extracted":    "No doc found",
                "passport_extracted_at": datetime.utcnow(),
                "passport_fetched":      True,
                "passport_id_retry":     True,  # prevent infinite re-run
            }}
        )
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "passport_found":  False,
            "message":         f"No documents returned for {email}",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Find Passport document — try all known type names ───────────────
    # Includes: "Passport/id card", "passport_id_card", "Passport", "Id Card"
    is_retry = (staff.get('passport_fetched') is True and
                not _v(staff.get('passport_id') or ''))

    passport_doc = None
    PASSPORT_TYPES = {
        'passport/id card', 'passport/id', 'passport', 'id card',
        'passport_id_card', 'passport id card', 'passport id',
    }
    for d in documents:
        doc_name = (d.get('document_type_name') or '').strip().lower()
        if doc_name in PASSPORT_TYPES and d.get('url'):
            passport_doc = d
            break

    if not passport_doc:
        col.update_one(
            {"_id": staff["_id"]},
            {"$set": {
                "passport_extracted":    "No passport doc found",
                "passport_extracted_at": datetime.utcnow(),
                "passport_fetched":      True,
                "passport_id_retry":     True,  # mark so we don't retry infinitely
            }}
        )
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "passport_found":  False,
            "message":         f"No Passport/id card document found for {email}",
            "remaining_count": max(0, remaining_total - 1),
        })

    passport_url = (passport_doc.get('url') or '').strip()

    if not passport_url:
        _mark_done({"passport_extracted": "[skipped — document URL is empty]"})
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "passport_found":  True,
            "skipped":         True,
            "reason":          "Document URL is empty",
            "remaining_count": max(0, remaining_total - 1),
            "message":         f"Skipped {full_name} ({email}) — passport doc has no URL",
        })

    # ── Download + extract via Gemini ─────────────────────────────────
    try:
        dl_headers = {k: v for k, v in api_headers.items()
                      if k != 'Content-Type'}
        img_resp = _req.get(passport_url, headers=dl_headers, timeout=60)

        if img_resp.status_code == 404:
            _mark_done({"passport_extracted": "[404 — document URL not found]"})
            return jsonify({
                "success":         True,
                "email":           email,
                "staff_name":      full_name,
                "passport_found":  True,
                "skipped":         True,
                "reason":          "Document URL returned 404",
                "remaining_count": max(0, remaining_total - 1),
                "message":         f"Skipped {full_name} ({email}) — passport URL 404",
            })

        img_resp.raise_for_status()
        raw_bytes    = img_resp.content
        content_type = img_resp.headers.get('Content-Type', '').lower()

        from google import genai as google_genai

        client = google_genai.Client(api_key=gemini_key)

        # ── Build Gemini request with image or text ───────────────────
        is_image = any(t in content_type for t in ('image/', 'jpeg', 'jpg', 'png', 'webp'))
        is_pdf   = 'pdf' in content_type or passport_url.lower().endswith('.pdf')

        if is_image:
            ext = 'jpeg' if 'jpeg' in content_type or 'jpg' in content_type else                   'png'  if 'png'  in content_type else                   'webp' if 'webp' in content_type else 'jpeg'
            b64_data = base64.b64encode(raw_bytes).decode('utf-8')
            parts = [
                {
                    "inline_data": {
                        "mime_type": f"image/{ext}",
                        "data": b64_data,
                    }
                },
                {"text": """You are a passport/ID card data extractor.

Extract all readable information from this passport or ID card image.

Return ONLY a JSON object — no markdown, no explanation:
{
  "passport_id": "<passport or document number>",
  "full_name": "<name as printed on document>",
  "nationality": "<nationality>",
  "date_of_birth": "<DOB as printed>",
  "expiry_date": "<expiry date>",
  "issue_date": "<issue date if visible>",
  "country": "<issuing country>",
  "gender": "<gender if visible>",
  "mrz": "<machine readable zone lines if visible>",
  "raw_text": "<all text extracted from the document>"
}

If a field is not visible or readable, set it to null.
"""}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[{"parts": parts}]
            )
        elif is_pdf:
            b64_data = base64.b64encode(raw_bytes).decode('utf-8')
            parts = [
                {
                    "inline_data": {
                        "mime_type": "application/pdf",
                        "data": b64_data,
                    }
                },
                {"text": """You are a passport/ID card data extractor.

Extract all readable information from this passport or ID card document.

Return ONLY a JSON object — no markdown, no explanation:
{
  "passport_id": "<passport or document number>",
  "full_name": "<name as printed on document>",
  "nationality": "<nationality>",
  "date_of_birth": "<DOB as printed>",
  "expiry_date": "<expiry date>",
  "issue_date": "<issue date if visible>",
  "country": "<issuing country>",
  "gender": "<gender if visible>",
  "mrz": "<machine readable zone if visible>",
  "raw_text": "<all text extracted>"
}

If a field is not visible, set it to null.
"""}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[{"parts": parts}]
            )
        else:
            # Try text extraction fallback
            try:
                import io as _io, pdfplumber
                with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                    raw_text = chr(10).join(p.extract_text() or '' for p in pdf.pages).strip()
            except Exception:
                raw_text = raw_bytes.decode('utf-8', errors='replace').strip()

            prompt = f"""You are a passport/ID card data extractor.

Extract all information from this passport/ID card text.

Return ONLY a JSON object — no markdown, no explanation:
{{
  "passport_id": "<passport or document number>",
  "full_name": "<name>",
  "nationality": "<nationality>",
  "date_of_birth": "<DOB>",
  "expiry_date": "<expiry date>",
  "issue_date": "<issue date>",
  "country": "<issuing country>",
  "gender": "<gender>",
  "mrz": "<MRZ lines if present>",
  "raw_text": "<all extracted text>"
}}

TEXT:
{raw_text[:5000]}
"""
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt
            )

        raw_out = (response.text or '').strip()
        raw_out = _re.sub(r'^```(?:json)?\s*', '', raw_out, flags=_re.MULTILINE)
        raw_out = _re.sub(r'```\s*$', '', raw_out, flags=_re.MULTILINE).strip()

        passport_data = _cjson.loads(raw_out)
        passport_id   = _v(passport_data.get('passport_id') or '')

        col.update_one(
            {"_id": staff["_id"]},
            {"$set": {
                "passport_extracted":    passport_data.get('raw_text', '') or '',
                "passport_data":         passport_data,
                "passport_id":           passport_id,
                "passport_url":          passport_url,
                "passport_extracted_at": datetime.utcnow(),
                "passport_fetched":      True,
                "passport_id_retry":     True,  # prevent infinite re-runs
            }}
        )

        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "passport_found":  True,
            "passport_id":     passport_id or None,
            "passport_readable": bool(passport_id),
            "passport_data":   passport_data,
            "remaining_count": max(0, remaining_total - 1),
            "message": (
                f"Passport {'extracted' if passport_id else 'fetched (ID not readable)'} for {full_name} "
                + (f"(ID: {passport_id}) — " if passport_id else "— ")
                + f"{max(0, remaining_total - 1)} remaining."
            ),
            "synced_at": datetime.utcnow().isoformat(),
        })

    except _cjson.JSONDecodeError:
        col.update_one(
            {"_id": staff["_id"]},
            {"$set": {
                "passport_extracted":    "[Gemini JSON parse error]",
                "passport_extracted_at": datetime.utcnow(),
                "passport_fetched":      True,
            }}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           "Gemini returned non-JSON response",
            "remaining_count": max(0, remaining_total - 1),
        })
    except Exception as e:
        col.update_one(
            {"_id": staff["_id"]},
            {"$set": {
                "passport_extracted":    f"[error: {e}]",
                "passport_extracted_at": datetime.utcnow(),
                "passport_fetched":      True,
            }}
        )
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           str(e),
            "remaining_count": max(0, remaining_total - 1),
        })



# ── Cron: Extract NMBI/QQI number from qualification docs ─────────────

@admin_bp.route('/live-staffs/cron/sync-qualification', methods=['GET', 'POST'])
def live_staff_cron_sync_qualification():
    """
    Cron job — processes ONE staff member per call.

    Logic:
      - Nurse:             looks for "Nmbi Qualification" document
      - Healthcare Asst:   looks for "QQI Level 5 or equivalent..." document
      Downloads doc, sends to Gemini to extract registration/ID number.
      Saves to live_staffs.nmbi_number or live_staffs.qqi_number.

    Protect with ?cron_key=<CRON_SECRET> env var.
    """
    import requests as _req

    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')
    gemini_key  = os.environ.get('GEMINI_API_KEY', '')

    if not base_url:
        return jsonify({"success": False, "error": "LIVE_STAFF_URL not set"}), 500
    if not gemini_key:
        return jsonify({"success": False, "error": "GEMINI_API_KEY not set"}), 500

    col = _staffs_col()

    # ── Find next unprocessed staff ────────────────────────────────────
    # Only process Nurses and Healthcare Assistants
    # Skip if already fetched (qualification_fetched = True)
    pending_query = {
        "$and": [
            {"$or": [
                {"user_type": {"$regex": "nurse", "$options": "i"}},
                {"user_type": {"$regex": "healthcare assistant", "$options": "i"}},
                {"user_type": {"$regex": "hca", "$options": "i"}},
            ]},
            # Exclude staff where document URL returned 404 — cannot retry
            {"qualification_doc_404": {"$ne": True}},
            # Not yet fetched OR HCA with no valid alphanumeric qqi_number
            {"$or": [
                {"qualification_fetched": {"$exists": False}},
                {"qualification_fetched": False},
                {"qualification_fetched": None},
                # Re-run HCA only if NOT 404 and qqi_number is missing/digits-only
                {"$and": [
                    {"$or": [
                        {"user_type": {"$regex": "healthcare assistant", "$options": "i"}},
                        {"user_type": {"$regex": "hca", "$options": "i"}},
                    ]},
                    {"qualification_doc_404": {"$ne": True}},
                    {"$or": [
                        {"qqi_number": {"$exists": False}},
                        {"qqi_number": None},
                        {"qqi_number": ""},
                        {"qqi_number": {"$not": {"$regex": "[A-Za-z]"}}},
                    ]},
                ]},
            ]},
        ]
    }

    remaining_total = col.count_documents(pending_query)
    staff           = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff qualifications already extracted — nothing to do.",
            "remaining_count": 0,
        })

    staff_id  = str(staff['_id'])
    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or '')
    user_type = _v(staff.get('user_type') or '')
    ut_lower  = user_type.lower()

    is_nurse = 'nurse' in ut_lower or 'nursing' in ut_lower
    is_hca   = 'hca' in ut_lower or 'healthcare assistant' in ut_lower or 'health care assistant' in ut_lower

    # Determine target document type
    if is_nurse:
        target_doc_types = ['nmbi qualification', 'nmbi', 'nursing qualification']
        save_field       = 'nmbi_number'
        extract_hint     = (
            "This is a nursing registration document. "
            "Find the NMBI registration PIN number. "
            "It is typically 6 digits, or a combination of letters and numbers (e.g. 123456, NMC12345). "
            "Look for labels like: PIN, Registration Number, NMBI Number, Reg No, Certificate No."
        )
    else:
        target_doc_types = [
            'qqi level 5 or equivalent in health service skills or healthcare support',
            'qqi level 5', 'qqi', 'healthcare support',
            'health service skills', 'qqi certificate',
        ]
        save_field   = 'qqi_number'
        extract_hint = (
            "This is a QQI / FETAC qualification certificate for a Healthcare Assistant. "
            "Your task is to carefully read and extract the certificate or registration number. "
            "\n\n"
            "READING INSTRUCTIONS — follow these steps carefully:\n"
            "1. Examine the ENTIRE document thoroughly, including headers, footers, watermarks, "
            "   stamps, and small print.\n"
            "2. If text appears small, blurry, or partially obscured — zoom in mentally and try "
            "   your best to reconstruct each character. Use context clues from surrounding text.\n"
            "3. Look at EVERY number/code on the document — not just the most obvious one.\n"
            "4. If a number is partially visible (e.g. only some digits readable), attempt to "
            "   reconstruct the full number using the visible characters and document format.\n"
            "\n"
            "IMPORTANT — The correct number MUST BE ALPHANUMERIC (contains both letters AND numbers), "
            "for example: F12345678, L123456789, QF12345, 12345ABC, FET123456. "
            "If you find a digits-only number, keep looking — there is almost certainly an "
            "alphanumeric reference elsewhere on the document.\n"
            "\n"
            "Look for labels like: Certificate No, Learner ID, Award Reference, QQI No, Record No, "
            "Registration No, Reference Number, Award ID, Cert Ref, Roll No."
        )

    def _mark_done(update_dict):
        update_dict["qualification_fetched"]    = True
        update_dict["qualification_fetched_at"] = datetime.utcnow()
        col.update_one({"_id": staff['_id']}, {"$set": update_dict})

    if not email:
        _mark_done({"qualification_note": "skipped — no email"})
        return jsonify({
            "success":         True,
            "message":         f"Skipped {staff_id} — no email",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Call XN Portal API ────────────────────────────────────────────
    endpoint    = f"{base_url}/ai/recruitments/user-document-list"
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    try:
        resp = _req.post(endpoint, json={"email": email},
                         headers=api_headers, timeout=30)
        if resp.status_code == 405:
            resp = _req.get(endpoint, params={"email": email},
                            headers=api_headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as api_err:
        _mark_done({"qualification_note": f"API error: {api_err}"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"API error: {api_err}",
            "remaining_count": max(0, remaining_total - 1),
        })

    if not data.get('success'):
        _mark_done({"qualification_note": f"API error: {data.get('message')}"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           data.get('message', 'API error'),
            "remaining_count": max(0, remaining_total - 1),
        })

    api_data  = data.get('data')
    documents = api_data if isinstance(api_data, list) else                 (api_data.get('documents') or [] if isinstance(api_data, dict) else [])

    if not documents:
        _mark_done({"qualification_note": "no documents returned"})
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "doc_found":       False,
            "message":         f"No documents returned for {email}",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Find matching qualification document ──────────────────────────
    qual_doc = None
    for d in documents:
        doc_name = (d.get('document_type_name') or '').strip().lower()
        if any(t in doc_name for t in target_doc_types) and d.get('url'):
            qual_doc = d
            break

    if not qual_doc:
        _mark_done({"qualification_note": f"no matching qualification doc found"})
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "doc_found":       False,
            "message":         f"No qualification document found for {full_name}",
            "remaining_count": max(0, remaining_total - 1),
        })

    doc_url = (qual_doc.get('url') or '').strip()

    # ── Guard: only proceed if URL is present ─────────────────────────
    if not doc_url:
        _mark_done({"qualification_note": "document found but URL is empty — skipped"})
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "doc_found":       True,
            "skipped":         True,
            "reason":          "Document URL is empty",
            "remaining_count": max(0, remaining_total - 1),
            "message":         f"Skipped {full_name} ({email}) — document has no URL",
        })

    # ── Download document ─────────────────────────────────────────────
    from google import genai as google_genai

    try:
        dl_headers = {k: v for k, v in api_headers.items() if k != 'Content-Type'}
        dl_resp    = _req.get(doc_url, headers=dl_headers, timeout=60)

        # 404 = document missing/expired URL — mark done and skip
        if dl_resp.status_code == 404:
            _mark_done({
                "qualification_note":   "document URL returned 404 — skipped",
                "qualification_doc_404": True,
            })
            return jsonify({
                "success":         True,
                "email":           email,
                "staff_name":      full_name,
                "doc_found":       True,
                "skipped":         True,
                "reason":          "Document URL returned 404 (missing or expired)",
                "remaining_count": max(0, remaining_total - 1),
                "message":         f"Skipped {full_name} ({email}) — document URL returned 404",
            })

        dl_resp.raise_for_status()
        raw_bytes    = dl_resp.content
        content_type = dl_resp.headers.get('Content-Type', '').lower()

        client = google_genai.Client(api_key=gemini_key)

        prompt_text = f"""{extract_hint}

EXTRACTION RULES:
- Scan every part of the document: title, body, header, footer, stamps, seals, watermarks.
- If any text is small or unclear, zoom in and do your best to read each character.
- Even if only partially readable, attempt to reconstruct the number using visible characters.
- Prefer alphanumeric codes over digit-only numbers.
- Report your confidence level honestly.

Return ONLY a JSON object — no markdown, no explanation:
{{
  "registration_number": "<the alphanumeric number you found, exactly as printed — reconstruct if partially visible>",
  "label_found": "<the label printed next to the number, e.g. PIN, Certificate No, Learner ID>",
  "confidence": "<high|medium|low — how clearly was the number readable>",
  "reconstruction_note": "<if you had to reconstruct any characters, explain here — otherwise null>",
  "raw_text": "<all readable text extracted from the document>"
}}

If after thorough examination no registration/certificate number is found, set "registration_number" to null.
"""

        is_image = any(t in content_type for t in ('image/', 'jpeg', 'jpg', 'png', 'webp'))
        is_pdf   = 'pdf' in content_type or doc_url.lower().split('?')[0].endswith('.pdf')

        if is_image:
            ext  = 'jpeg' if any(t in content_type for t in ('jpeg','jpg')) else                    'png'  if 'png' in content_type else 'webp' if 'webp' in content_type else 'jpeg'
            parts = [
                {"inline_data": {"mime_type": f"image/{ext}",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[{"parts": parts}]
            )
        elif is_pdf:
            parts = [
                {"inline_data": {"mime_type": "application/pdf",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[{"parts": parts}]
            )
        else:
            try:
                import io as _io, pdfplumber
                with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                    raw_text = chr(10).join(p.extract_text() or '' for p in pdf.pages).strip()
            except Exception:
                raw_text = raw_bytes.decode('utf-8', errors='replace').strip()
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt_text + "\n\nDOCUMENT TEXT:\n" + raw_text[:5000]
            )

        raw_out = (response.text or '').strip()
        raw_out = _re.sub(r'^```(?:json)?\s*', '', raw_out, flags=_re.MULTILINE)
        raw_out = _re.sub(r'```\s*$', '', raw_out, flags=_re.MULTILINE).strip()

        result    = _cjson.loads(raw_out)
        reg_num   = _v(result.get('registration_number') or '')
        raw_text_ = _v(result.get('raw_text') or '')

        # ── For HCA: validate qqi_number is alphanumeric (contains letters) ──
        # If Gemini returned digits-only, reset and let cron retry next run
        import re as _re2
        if is_hca and reg_num:
            has_letter = bool(_re2.search(r'[A-Za-z]', reg_num))
            if not has_letter:
                # Digits-only — invalid QQI number, reset so cron retries
                col.update_one(
                    {"_id": staff['_id']},
                    {"$set": {
                        "qqi_number":            "",
                        "qualification_fetched": False,
                        "qualification_note":    f"digits-only QQI rejected: {reg_num} — will retry",
                        "qualification_doc_url": doc_url,
                    }}
                )
                return jsonify({
                    "success":         False,
                    "email":           email,
                    "staff_name":      full_name,
                    "user_type":       user_type,
                    "doc_found":       True,
                    "qqi_number":      "",
                    "rejected":        reg_num,
                    "reason":          "QQI number must be alphanumeric (contain letters). Digits-only rejected.",
                    "remaining_count": remaining_total,
                    "message":         f"Rejected digits-only QQI '{reg_num}' for {full_name} — will retry.",
                })
            reg_num = reg_num.strip()

        recon_note = _v(result.get('reconstruction_note') or '')
        _mark_done({
            save_field:               reg_num,
            "qualification_doc_type": qual_doc.get('document_type_name', ''),
            "qualification_doc_url":  doc_url,
            "qualification_raw":      raw_text_,
            "qualification_data":     result,
            "qualification_confidence": _v(result.get('confidence') or ''),
            "qualification_reconstruction": recon_note,
            "qualification_note":     f"extracted from {qual_doc.get('document_type_name','')}",
        })

        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "user_type":       user_type,
            "doc_found":       True,
            save_field:        reg_num,
            "doc_type":        qual_doc.get('document_type_name', ''),
            "readable":        bool(reg_num),
            "remaining_count": max(0, remaining_total - 1),
            "message": (
                f"{save_field} {'extracted: ' + reg_num if reg_num else 'not readable'} "
                f"for {full_name} — {max(0, remaining_total - 1)} remaining."
            ),
        })

    except _cjson.JSONDecodeError:
        _mark_done({"qualification_note": "Gemini JSON parse error"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           "Gemini returned non-JSON",
            "remaining_count": max(0, remaining_total - 1),
        })
    except Exception as e:
        _mark_done({"qualification_note": f"error: {e}"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           str(e),
            "remaining_count": max(0, remaining_total - 1),
        })


# ── Export: Nurse NMBI list ───────────────────────────────────────────

@admin_bp.route('/live-staffs/export/nmbi-xlsx')
@admin_required
def live_staff_export_nmbi_xlsx():
    """Export Nurse staff with NMBI numbers to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as _io

        docs = list(_staffs_col().find(
            {"$or": [
                {"user_type": {"$regex": "nurse", "$options": "i"}},
                {"user_type": {"$regex": "nursing", "$options": "i"}},
            ]},
            {"section_1_personal_details": 1, "email": 1,
             "user_type": 1, "nmbi_number": 1, "qualification_fetched": 1}
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        wb, ws = _build_qual_xlsx(docs, 'Nurses — NMBI', 'nmbi_number', 'NMBI Number')

        buf = _io.BytesIO()
        wb.save(buf)
        return Response(buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="nmbi_numbers_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ── Export: HCA QQI list ──────────────────────────────────────────────

@admin_bp.route('/live-staffs/export/qqi-xlsx')
@admin_required
def live_staff_export_qqi_xlsx():
    """Export Healthcare Assistant staff with QQI numbers to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as _io

        docs = list(_staffs_col().find(
            {"$or": [
                {"user_type": {"$regex": "healthcare assistant", "$options": "i"}},
                {"user_type": {"$regex": "\bhca\b", "$options": "i"}},
            ]},
            {"section_1_personal_details": 1, "email": 1,
             "user_type": 1, "qqi_number": 1, "qualification_fetched": 1}
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        wb, ws = _build_qual_xlsx(docs, 'HCA — QQI Level 5', 'qqi_number', 'QQI Number')

        buf = _io.BytesIO()
        wb.save(buf)
        return Response(buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="qqi_numbers_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


def _build_qual_xlsx(docs, sheet_title, number_field, number_label):
    """Shared builder for NMBI/QQI Excel exports."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    NAVY = '1B3A6B'; GREEN = '2E9E44'; WHITE = 'FFFFFF'
    ALT  = 'EFF6FF'; WARN  = 'FFFBEB'

    h_font  = Font(name='Arial', bold=True, color=WHITE, size=10)
    h_fill  = PatternFill('solid', start_color=NAVY, end_color=NAVY)
    h_align = Alignment(horizontal='center', vertical='center')
    b_font  = Font(name='Arial', size=10)
    l_align = Alignment(horizontal='left', vertical='center')
    c_align = Alignment(horizontal='center', vertical='center')
    thin    = Side(style='thin', color='CCCCCC')
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_b = Border(left=thin, right=thin, top=thin,
                     bottom=Side(style='medium', color=GREEN))

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    headers    = ['Sno', 'Name', 'Email', number_label, 'Status']
    col_widths = [5, 32, 40, 25, 18]

    for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align
        cell.border    = green_b
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    for ri, doc in enumerate(docs, start=2):
        s1     = doc.get('section_1_personal_details') or {}
        name   = _v(s1.get('full_name') or '')
        email  = _v(doc.get('email') or '')
        number = _v(doc.get(number_field) or '')
        fetched= doc.get('qualification_fetched', False)

        if number:
            status = 'Found'
            row_fill = None
        elif fetched:
            status = 'Not Found'
            row_fill = PatternFill('solid', start_color='FFDDDD', end_color='FFDDDD')
        else:
            status = 'Not Checked'
            row_fill = PatternFill('solid', start_color=WARN, end_color=WARN)

        alt_fill = PatternFill('solid', start_color=ALT, end_color=ALT) if ri % 2 == 0 and not row_fill else None

        for ci, (val, align) in enumerate(
            [(ri-1, c_align), (name, l_align), (email, l_align), (number, c_align), (status, c_align)],
            start=1
        ):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = b_font
            cell.alignment = align
            cell.border    = border
            cell.fill      = row_fill or alt_fill or PatternFill()

        ws.row_dimensions[ri].height = 17

    ws.cell(row=len(docs)+2, column=1,
            value=f'Total: {len(docs)}').font = Font(name='Arial', bold=True, size=9)

    return wb, ws



# ── Cron: Extract CPR/BLS certificate details one staff at a time ─────

@admin_bp.route('/live-staffs/cron/sync-cpr-certificate', methods=['GET', 'POST'])
def live_staff_cron_sync_cpr_certificate():
    """
    Cron job — processes ONE staff member per call.

    Finds "Cpr/Bls" document from XN Portal, sends to Gemini AI to extract:
      - Certificate name
      - Staff name as printed on certificate
      - Expiry date

    Saves to live_staffs:
      - cpr_certificate_name
      - cpr_staff_name
      - cpr_expiry_date
      - cpr_fetched = True

    Protect with ?cron_key=<CRON_SECRET> env var.
    """
    import requests as _req
    from google import genai as google_genai

    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url   = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key    = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country= os.environ.get('XN_APP_COUNTRY', '')
    gemini_key = os.environ.get('GEMINI_API_KEY', '')

    if not base_url:
        return jsonify({"success": False, "error": "LIVE_STAFF_URL not set"}), 500
    if not gemini_key:
        return jsonify({"success": False, "error": "GEMINI_API_KEY not set"}), 500

    col = _staffs_col()

    pending_query = {
        "$or": [
            {"cpr_fetched": {"$exists": False}},
            {"cpr_fetched": False},
            {"cpr_fetched": None},
        ]
    }
    remaining_total = col.count_documents(pending_query)
    staff           = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff CPR/BLS certificates already extracted.",
            "remaining_count": 0,
        })

    staff_id  = str(staff['_id'])
    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or
                   s1.get('email_address') or '')

    def _mark_done(fields):
        fields["cpr_fetched"]    = True
        fields["cpr_fetched_at"] = datetime.utcnow()
        col.update_one({"_id": staff['_id']}, {"$set": fields})

    if not email:
        _mark_done({"cpr_note": "skipped — no email"})
        return jsonify({
            "success":         True,
            "message":         f"Skipped — no email",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Call XN Portal API ────────────────────────────────────────────
    endpoint    = f"{base_url}/ai/recruitments/user-document-list"
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    try:
        resp = _req.post(endpoint, json={"email": email},
                         headers=api_headers, timeout=30)
        if resp.status_code == 405:
            resp = _req.get(endpoint, params={"email": email},
                            headers=api_headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        _mark_done({"cpr_note": f"API error: {e}"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"API error: {e}",
            "remaining_count": max(0, remaining_total - 1),
        })

    if not data.get('success'):
        _mark_done({"cpr_note": f"API error: {data.get('message')}"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           data.get('message', 'API error'),
            "remaining_count": max(0, remaining_total - 1),
        })

    api_data  = data.get('data')
    documents = api_data if isinstance(api_data, list) else                 (api_data.get('documents') or [] if isinstance(api_data, dict) else [])

    if not documents:
        _mark_done({"cpr_note": "no documents returned"})
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "doc_found":       False,
            "message":         f"No documents returned for {email}",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Find CPR/BLS document ─────────────────────────────────────────
    cpr_doc = None
    for d in documents:
        doc_name = (d.get('document_type_name') or '').strip().lower()
        if any(t in doc_name for t in ('cpr/bls', 'cpr', 'bls', 'basic life support',
                                        'cardiopulmonary')) and d.get('url'):
            cpr_doc = d
            break

    if not cpr_doc:
        _mark_done({"cpr_note": "no CPR/BLS document found"})
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "doc_found":       False,
            "message":         f"No CPR/BLS document found for {full_name}",
            "remaining_count": max(0, remaining_total - 1),
        })

    doc_url = (cpr_doc.get('url') or '').strip()

    if not doc_url:
        _mark_done({"cpr_note": "document found but URL is empty — skipped"})
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "doc_found":       True,
            "skipped":         True,
            "reason":          "Document URL is empty",
            "remaining_count": max(0, remaining_total - 1),
            "message":         f"Skipped {full_name} ({email}) — CPR doc has no URL",
        })

    # ── Download and extract with Gemini ──────────────────────────────
    try:
        dl_headers = {k: v for k, v in api_headers.items() if k != 'Content-Type'}
        dl_resp    = _req.get(doc_url, headers=dl_headers, timeout=60)

        if dl_resp.status_code == 404:
            _mark_done({"cpr_note": "document URL returned 404 — skipped"})
            return jsonify({
                "success":         True,
                "email":           email,
                "staff_name":      full_name,
                "doc_found":       True,
                "skipped":         True,
                "reason":          "Document URL returned 404",
                "remaining_count": max(0, remaining_total - 1),
                "message":         f"Skipped {full_name} ({email}) — CPR doc URL 404",
            })

        dl_resp.raise_for_status()
        raw_bytes    = dl_resp.content
        content_type = dl_resp.headers.get('Content-Type', '').lower()

        client = google_genai.Client(api_key=gemini_key)

        prompt_text = """You are a certificate data extractor.

Extract the following details from this CPR/BLS certificate:
1. Certificate name (e.g. "Basic Life Support", "CPR/AED Certificate", "BLS for Healthcare Providers")
2. Staff name as printed on the certificate
3. Expiry date (the date the certificate expires or renewal is due)

Return ONLY a JSON object — no markdown, no explanation:
{
  "certificate_name": "<exact certificate title as printed>",
  "staff_name_on_cert": "<name as printed on certificate>",
  "expiry_date": "<expiry date as printed, e.g. 01/06/2027 or June 2027>",
  "issue_date": "<issue/completion date if visible>",
  "issuing_body": "<organization that issued the certificate>"
}

If a field is not visible, set it to null.
"""

        is_image = any(t in content_type for t in ('image/', 'jpeg', 'jpg', 'png', 'webp'))
        is_pdf   = 'pdf' in content_type or doc_url.lower().split('?')[0].endswith('.pdf')

        if is_image:
            ext   = 'jpeg' if any(t in content_type for t in ('jpeg','jpg')) else                     'png'  if 'png' in content_type else 'webp' if 'webp' in content_type else 'jpeg'
            parts = [
                {"inline_data": {"mime_type": f"image/{ext}",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[{"parts": parts}]
            )
        elif is_pdf:
            parts = [
                {"inline_data": {"mime_type": "application/pdf",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[{"parts": parts}]
            )
        else:
            try:
                import io as _io, pdfplumber
                with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                    raw_text = chr(10).join(p.extract_text() or '' for p in pdf.pages).strip()
            except Exception:
                raw_text = raw_bytes.decode('utf-8', errors='replace').strip()
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt_text + "\n\nCERTIFICATE TEXT:\n" + raw_text[:5000]
            )

        raw_out = (response.text or '').strip()
        raw_out = _re.sub(r'^```(?:json)?\s*', '', raw_out, flags=_re.MULTILINE)
        raw_out = _re.sub(r'```\s*$', '', raw_out, flags=_re.MULTILINE).strip()

        result       = _cjson.loads(raw_out)
        cert_name    = _v(result.get('certificate_name') or '')
        cert_staff   = _v(result.get('staff_name_on_cert') or '')
        expiry_date  = _v(result.get('expiry_date') or '')
        issue_date   = _v(result.get('issue_date') or '')
        issuing_body = _v(result.get('issuing_body') or '')

        _mark_done({
            "cpr_certificate_name": cert_name,
            "cpr_staff_name":       cert_staff,
            "cpr_expiry_date":      expiry_date,
            "cpr_issue_date":       issue_date,
            "cpr_issuing_body":     issuing_body,
            "cpr_doc_url":          doc_url,
            "cpr_doc_type":         cpr_doc.get('document_type_name', ''),
            "cpr_note":             "extracted successfully",
        })

        return jsonify({
            "success":              True,
            "email":                email,
            "staff_name":           full_name,
            "doc_found":            True,
            "certificate_name":     cert_name,
            "staff_name_on_cert":   cert_staff,
            "expiry_date":          expiry_date,
            "issuing_body":         issuing_body,
            "remaining_count":      max(0, remaining_total - 1),
            "message": (
                f"CPR/BLS cert extracted for {full_name} "
                f"(expires: {expiry_date or 'unknown'}) — "
                f"{max(0, remaining_total - 1)} remaining."
            ),
        })

    except _cjson.JSONDecodeError:
        _mark_done({"cpr_note": "Gemini JSON parse error"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           "Gemini returned non-JSON",
            "remaining_count": max(0, remaining_total - 1),
        })
    except Exception as e:
        _mark_done({"cpr_note": f"error: {e}"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           str(e),
            "remaining_count": max(0, remaining_total - 1),
        })


# ── Export: CPR/BLS certificates to Excel ────────────────────────────

@admin_bp.route('/live-staffs/export/cpr-xlsx')
@admin_required
def live_staff_export_cpr_xlsx():
    """Export CPR/BLS certificate details to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as _io

        docs = list(_staffs_col().find(
            {},
            {"section_1_personal_details": 1, "email": 1,
             "cpr_certificate_name": 1, "cpr_staff_name": 1,
             "cpr_expiry_date": 1, "cpr_issue_date": 1,
             "cpr_issuing_body": 1, "cpr_fetched": 1}
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        NAVY  = '1B3A6B'; GREEN = '2E9E44'; WHITE = 'FFFFFF'
        ALT   = 'EFF6FF'; WARN  = 'FFF3CD'; RED   = 'FFDDDD'

        h_font  = Font(name='Arial', bold=True, color=WHITE, size=10)
        h_fill  = PatternFill('solid', start_color=NAVY, end_color=NAVY)
        h_align = Alignment(horizontal='center', vertical='center')
        b_font  = Font(name='Arial', size=10)
        l_align = Alignment(horizontal='left', vertical='center')
        c_align = Alignment(horizontal='center', vertical='center')
        thin    = Side(style='thin', color='CCCCCC')
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        green_b = Border(left=thin, right=thin, top=thin,
                         bottom=Side(style='medium', color=GREEN))

        wb = Workbook()
        ws = wb.active
        ws.title = 'CPR-BLS Certificates'

        headers    = ['Sno', 'Staff Name', 'Email', 'Certificate Name',
                      'Name on Cert', 'Expiry Date', 'Issue Date', 'Issuing Body', 'Status']
        col_widths = [5, 28, 36, 30, 28, 16, 16, 28, 14]

        for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = h_font; cell.fill = h_fill
            cell.alignment = h_align; cell.border = green_b
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:I{len(docs)+1}'

        from datetime import date as _date
        today = _date.today()

        def _is_expired(expiry_str):
            if not expiry_str:
                return None
            for fmt in ('%d/%m/%Y','%m/%Y','%Y-%m-%d','%d-%m-%Y','%B %Y','%b %Y'):
                try:
                    from datetime import datetime as _dt
                    d = _dt.strptime(expiry_str.strip(), fmt).date()
                    return d < today
                except Exception:
                    continue
            return None

        for ri, doc in enumerate(docs, start=2):
            s1       = doc.get('section_1_personal_details') or {}
            name     = _v(s1.get('full_name') or '')
            email    = _v(doc.get('email') or '')
            cert_n   = _v(doc.get('cpr_certificate_name') or '')
            cert_s   = _v(doc.get('cpr_staff_name') or '')
            expiry   = _v(doc.get('cpr_expiry_date') or '')
            issue    = _v(doc.get('cpr_issue_date') or '')
            issuer   = _v(doc.get('cpr_issuing_body') or '')
            fetched  = doc.get('cpr_fetched', False)

            expired  = _is_expired(expiry)

            if not fetched:
                status   = 'Not Checked'
                row_fill = PatternFill('solid', start_color=WARN, end_color=WARN)
            elif not cert_n:
                status   = 'No Cert Found'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is True:
                status   = 'EXPIRED'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is False:
                status   = 'Valid'
                row_fill = None
            else:
                status   = 'Found'
                row_fill = None

            alt_fill = PatternFill('solid', start_color=ALT, end_color=ALT)                        if ri % 2 == 0 and not row_fill else None

            row_vals = [ri-1, name, email, cert_n, cert_s, expiry, issue, issuer, status]
            aligns   = [c_align,l_align,l_align,l_align,l_align,c_align,c_align,l_align,c_align]

            for ci, (val, align) in enumerate(zip(row_vals, aligns), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = b_font
                cell.alignment = align
                cell.border    = border
                cell.fill      = row_fill or alt_fill or PatternFill()

            ws.row_dimensions[ri].height = 17

        ws.cell(row=len(docs)+2, column=1,
                value=f'Total: {len(docs)}').font = Font(name='Arial', bold=True, size=9)

        buf = _io.BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="cpr_certificates_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'}
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500



@admin_bp.route('/live-staffs/export/missing-nmbi-xlsx')
@admin_required
def live_staff_export_missing_nmbi_xlsx():
    """Export Nurses with missing NMBI registration number."""
    try:
        import io as _io
        docs = list(_staffs_col().find(
            {"$and": [
                {"$or": [
                    {"user_type": {"$regex": "nurse", "$options": "i"}},
                    {"user_type": {"$regex": "nursing", "$options": "i"}},
                ]},
                {"$or": [
                    {"nmbi_number": {"$exists": False}},
                    {"nmbi_number": None},
                    {"nmbi_number": ""},
                ]},
            ]},
            {"section_1_personal_details": 1, "email": 1,
             "user_type": 1, "nmbi_number": 1, "qualification_fetched": 1}
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        wb, ws = _build_qual_xlsx(docs, 'Missing NMBI Numbers', 'nmbi_number', 'NMBI Number')
        buf = _io.BytesIO()
        wb.save(buf)
        return Response(buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="missing_nmbi_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route('/live-staffs/export/missing-qqi-xlsx')
@admin_required
def live_staff_export_missing_qqi_xlsx():
    """Export Healthcare Assistants with missing QQI number."""
    try:
        import io as _io
        docs = list(_staffs_col().find(
            {"$and": [
                {"$or": [
                    {"user_type": {"$regex": "healthcare assistant", "$options": "i"}},
                    {"user_type": {"$regex": "\bhca\b", "$options": "i"}},
                ]},
                {"$or": [
                    {"qqi_number": {"$exists": False}},
                    {"qqi_number": None},
                    {"qqi_number": ""},
                ]},
            ]},
            {"section_1_personal_details": 1, "email": 1,
             "user_type": 1, "qqi_number": 1, "qualification_fetched": 1}
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        wb, ws = _build_qual_xlsx(docs, 'Missing QQI Numbers', 'qqi_number', 'QQI Number')
        buf = _io.BytesIO()
        wb.save(buf)
        return Response(buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="missing_qqi_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500



@admin_bp.route('/live-staffs/export/passport-xlsx')
@admin_required
def live_staff_export_passport_xlsx():
    """Export staff with saved passport IDs to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as _io

        docs = list(_staffs_col().find(
            {"passport_fetched": True},
            {"section_1_personal_details": 1, "email": 1,
             "passport_id": 1, "passport_data": 1, "passport_fetched": 1}
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        NAVY  = '1B3A6B'; GREEN = '2E9E44'; WHITE = 'FFFFFF'
        ALT   = 'EFF6FF'; RED   = 'FFDDDD'; WARN  = 'FFF3CD'

        h_font  = Font(name='Arial', bold=True, color=WHITE, size=10)
        h_fill  = PatternFill('solid', start_color=NAVY, end_color=NAVY)
        h_align = Alignment(horizontal='center', vertical='center')
        b_font  = Font(name='Arial', size=10)
        l_align = Alignment(horizontal='left', vertical='center')
        c_align = Alignment(horizontal='center', vertical='center')
        thin    = Side(style='thin', color='CCCCCC')
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        green_b = Border(left=thin, right=thin, top=thin,
                         bottom=Side(style='medium', color=GREEN))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Passport IDs'

        headers    = ['Sno', 'Name', 'Email', 'Passport ID',
                      'Nationality', 'DOB', 'Expiry Date', 'Status']
        col_widths = [5, 30, 36, 18, 18, 14, 14, 14]

        for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = h_font; cell.fill = h_fill
            cell.alignment = h_align; cell.border = green_b
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:H{len(docs)+1}'

        for ri, doc in enumerate(docs, start=2):
            s1          = doc.get('section_1_personal_details') or {}
            name        = _v(s1.get('full_name') or '')
            email       = _v(doc.get('email') or '')
            passport_id = _v(doc.get('passport_id') or '')
            pdata       = doc.get('passport_data') or {}
            nationality = _v(pdata.get('nationality') or '')
            dob         = _v(pdata.get('date_of_birth') or '')
            expiry      = _v(pdata.get('expiry_date') or '')

            if passport_id:
                status   = 'Found'
                row_fill = None
            else:
                status   = 'Not Readable'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)

            alt_fill = PatternFill('solid', start_color=ALT, end_color=ALT)                        if ri % 2 == 0 and not row_fill else None

            row_vals = [ri-1, name, email, passport_id,
                        nationality, dob, expiry, status]
            aligns   = [c_align, l_align, l_align, c_align,
                        l_align, c_align, c_align, c_align]

            for ci, (val, align) in enumerate(zip(row_vals, aligns), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = b_font; cell.alignment = align
                cell.border = border
                cell.fill = row_fill or alt_fill or PatternFill()

            ws.row_dimensions[ri].height = 17

        ws.cell(row=len(docs)+2, column=1,
                value=f'Total: {len(docs)}').font = Font(name='Arial', bold=True, size=9)

        buf = _io.BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="passport_ids_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'}
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500



@admin_bp.route('/live-staffs/cron/reset-hca-qualification', methods=['GET', 'POST'])
@admin_required
def live_staff_reset_hca_qualification():
    """
    Reset qualification_fetched for ALL HCA staff so sync-qualification
    cron re-runs and re-extracts their QQI numbers.

    Also resets any HCA with a digits-only qqi_number (invalid).

    GET /admin/live-staffs/cron/reset-hca-qualification
    """
    import re as _re2
    col = _staffs_col()

    hca_query = {"$or": [
        {"user_type": {"$regex": "healthcare assistant", "$options": "i"}},
        {"user_type": {"$regex": r"\bhca\b", "$options": "i"}},
    ]}

    # Find HCA with digits-only or missing qqi_number
    reset_query = {"$and": [
        hca_query,
        {"$or": [
            {"qualification_fetched": True},
            {"qqi_number": {"$exists": False}},
            {"qqi_number": None},
            {"qqi_number": ""},
        ]},
    ]}

    all_hca  = list(col.find(hca_query, {"qqi_number": 1}))
    to_reset = []
    for doc in all_hca:
        qqi = _v(doc.get('qqi_number') or '')
        # Reset if empty OR digits-only (no letters)
        if not qqi or not _re2.search(r'[A-Za-z]', qqi):
            to_reset.append(doc['_id'])

    if not to_reset:
        return jsonify({
            "success": True,
            "reset_count": 0,
            "message": "No HCA staff need resetting — all have valid alphanumeric QQI numbers.",
        })

    result = col.update_many(
        {"_id": {"$in": to_reset}},
        {"$set": {
            "qualification_fetched": False,
            "qqi_number":            "",
            "qualification_note":    "reset for re-extraction",
        }}
    )

    return jsonify({
        "success":     True,
        "reset_count": result.modified_count,
        "total_hca":   len(all_hca),
        "message":     f"Reset {result.modified_count} HCA staff — cron will re-extract their QQI numbers.",
    })



# ── Cron: Extract Infection Prevention & Control Certificate ──────────

@admin_bp.route('/live-staffs/cron/sync-ipc-certificate', methods=['GET', 'POST'])
def live_staff_cron_sync_ipc_certificate():
    """
    Cron job — processes ONE staff member per call.

    Finds "Infection Prevention Control Certificate" document from XN Portal,
    sends to Gemini AI to extract:
      - Certificate name
      - Staff name as printed on certificate
      - Expiry date / completion date

    Saves to live_staffs:
      - ipc_certificate_name
      - ipc_staff_name
      - ipc_expiry_date
      - ipc_issue_date
      - ipc_issuing_body
      - ipc_fetched = True

    Protect with ?cron_key=<CRON_SECRET> env var.
    """
    import requests as _req
    from google import genai as google_genai

    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')
    gemini_key  = os.environ.get('GEMINI_API_KEY', '')

    if not base_url:
        return jsonify({"success": False, "error": "LIVE_STAFF_URL not set"}), 500
    if not gemini_key:
        return jsonify({"success": False, "error": "GEMINI_API_KEY not set"}), 500

    col = _staffs_col()

    pending_query = {
        "$or": [
            {"ipc_fetched": {"$exists": False}},
            {"ipc_fetched": False},
            {"ipc_fetched": None},
        ]
    }
    remaining_total = col.count_documents(pending_query)
    staff           = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff IPC certificates already extracted.",
            "remaining_count": 0,
        })

    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or s1.get('email_address') or '')

    def _mark_done(fields):
        fields["ipc_fetched"]    = True
        fields["ipc_fetched_at"] = datetime.utcnow()
        col.update_one({"_id": staff['_id']}, {"$set": fields})

    if not email:
        _mark_done({"ipc_note": "skipped — no email"})
        return jsonify({
            "success":         True,
            "message":         "Skipped — no email",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Call XN Portal API ────────────────────────────────────────────
    endpoint    = f"{base_url}/ai/recruitments/user-document-list"
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    try:
        resp = _req.post(endpoint, json={"email": email},
                         headers=api_headers, timeout=30)
        if resp.status_code == 405:
            resp = _req.get(endpoint, params={"email": email},
                            headers=api_headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        _mark_done({"ipc_note": f"API error: {e}"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           f"API error: {e}",
            "remaining_count": max(0, remaining_total - 1),
        })

    if not data.get('success'):
        _mark_done({"ipc_note": f"API error: {data.get('message')}"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           data.get('message', 'API error'),
            "remaining_count": max(0, remaining_total - 1),
        })

    api_data  = data.get('data')
    documents = api_data if isinstance(api_data, list) else                 (api_data.get('documents') or [] if isinstance(api_data, dict) else [])

    if not documents:
        _mark_done({"ipc_note": "no documents returned"})
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "doc_found":       False,
            "message":         f"No documents returned for {email}",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Find IPC document ─────────────────────────────────────────────
    ipc_doc = None
    for d in documents:
        doc_name = (d.get('document_type_name') or '').strip().lower()
        if any(t in doc_name for t in (
            'infection prevention control certificate',
            'infection prevention and control',
            'infection prevention & control',
            'infection control certificate',
            'infection prevention',
            'ipc certificate',
            'ipc',
        )) and d.get('url'):
            ipc_doc = d
            break

    if not ipc_doc:
        _mark_done({"ipc_note": "no IPC document found"})
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "doc_found":       False,
            "message":         f"No IPC certificate found for {full_name}",
            "remaining_count": max(0, remaining_total - 1),
        })

    doc_url = (ipc_doc.get('url') or '').strip()

    if not doc_url:
        _mark_done({"ipc_note": "document found but URL is empty — skipped"})
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "doc_found":       True,
            "skipped":         True,
            "reason":          "Document URL is empty",
            "remaining_count": max(0, remaining_total - 1),
            "message":         f"Skipped {full_name} ({email}) — IPC doc has no URL",
        })

    # ── Download and extract with Gemini ──────────────────────────────
    try:
        dl_headers = {k: v for k, v in api_headers.items() if k != 'Content-Type'}
        dl_resp    = _req.get(doc_url, headers=dl_headers, timeout=60)

        if dl_resp.status_code == 404:
            _mark_done({"ipc_note": "document URL returned 404 — skipped",
                        "ipc_doc_404": True})
            return jsonify({
                "success":         True,
                "email":           email,
                "staff_name":      full_name,
                "doc_found":       True,
                "skipped":         True,
                "reason":          "Document URL returned 404",
                "remaining_count": max(0, remaining_total - 1),
                "message":         f"Skipped {full_name} ({email}) — IPC doc URL 404",
            })

        dl_resp.raise_for_status()
        raw_bytes    = dl_resp.content
        content_type = dl_resp.headers.get('Content-Type', '').lower()

        client = google_genai.Client(api_key=gemini_key)

        prompt_text = """You are a certificate data extractor.

Extract the following details from this Infection Prevention and Control (IPC) certificate:
1. Certificate name (e.g. "Infection Prevention and Control", "IPC Training Certificate")
2. Staff name as printed on the certificate
3. Expiry date or renewal date (if shown)
4. Issue / completion date
5. Issuing body or training provider

Return ONLY a JSON object — no markdown, no explanation:
{
  "certificate_name": "<exact certificate title as printed>",
  "staff_name_on_cert": "<name as printed on certificate>",
  "expiry_date": "<expiry or renewal date as printed, e.g. 01/06/2027 or June 2027>",
  "issue_date": "<issue or completion date as printed>",
  "issuing_body": "<organization that issued the certificate>"
}

If a field is not visible, set it to null.
"""

        is_image = any(t in content_type for t in ('image/', 'jpeg', 'jpg', 'png', 'webp'))
        is_pdf   = 'pdf' in content_type or doc_url.lower().split('?')[0].endswith('.pdf')

        if is_image:
            ext   = 'jpeg' if any(t in content_type for t in ('jpeg', 'jpg')) else                     'png'  if 'png'  in content_type else                     'webp' if 'webp' in content_type else 'jpeg'
            parts = [
                {"inline_data": {"mime_type": f"image/{ext}",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[{"parts": parts}]
            )
        elif is_pdf:
            parts = [
                {"inline_data": {"mime_type": "application/pdf",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[{"parts": parts}]
            )
        else:
            try:
                import io as _io, pdfplumber
                with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                    raw_text = chr(10).join(p.extract_text() or '' for p in pdf.pages).strip()
            except Exception:
                raw_text = raw_bytes.decode('utf-8', errors='replace').strip()
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt_text + "\n\nCERTIFICATE TEXT:\n" + raw_text[:5000]
            )

        raw_out = (response.text or '').strip()
        raw_out = _re.sub(r'^```(?:json)?\s*', '', raw_out, flags=_re.MULTILINE)
        raw_out = _re.sub(r'```\s*$', '', raw_out, flags=_re.MULTILINE).strip()

        result       = _cjson.loads(raw_out)
        cert_name    = _v(result.get('certificate_name') or '')
        cert_staff   = _v(result.get('staff_name_on_cert') or '')
        expiry_date  = _v(result.get('expiry_date') or '')
        issue_date   = _v(result.get('issue_date') or '')
        issuing_body = _v(result.get('issuing_body') or '')

        _mark_done({
            "ipc_certificate_name": cert_name,
            "ipc_staff_name":       cert_staff,
            "ipc_expiry_date":      expiry_date,
            "ipc_issue_date":       issue_date,
            "ipc_issuing_body":     issuing_body,
            "ipc_doc_url":          doc_url,
            "ipc_doc_type":         ipc_doc.get('document_type_name', ''),
            "ipc_note":             "extracted successfully",
        })

        return jsonify({
            "success":              True,
            "email":                email,
            "staff_name":           full_name,
            "doc_found":            True,
            "certificate_name":     cert_name,
            "staff_name_on_cert":   cert_staff,
            "expiry_date":          expiry_date,
            "issue_date":           issue_date,
            "issuing_body":         issuing_body,
            "remaining_count":      max(0, remaining_total - 1),
            "message": (
                f"IPC cert extracted for {full_name} "
                f"(expires: {expiry_date or 'unknown'}) — "
                f"{max(0, remaining_total - 1)} remaining."
            ),
        })

    except _cjson.JSONDecodeError:
        _mark_done({"ipc_note": "Gemini JSON parse error"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           "Gemini returned non-JSON",
            "remaining_count": max(0, remaining_total - 1),
        })
    except Exception as e:
        _mark_done({"ipc_note": f"error: {e}"})
        return jsonify({
            "success":         False,
            "email":           email,
            "error":           str(e),
            "remaining_count": max(0, remaining_total - 1),
        })


# ── Export: IPC certificates to Excel ────────────────────────────────

@admin_bp.route('/live-staffs/export/ipc-xlsx')
@admin_required
def live_staff_export_ipc_xlsx():
    """Export IPC certificate details to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as _io

        docs = list(_staffs_col().find(
            {},
            {"section_1_personal_details": 1, "email": 1,
             "ipc_certificate_name": 1, "ipc_staff_name": 1,
             "ipc_expiry_date": 1, "ipc_issue_date": 1,
             "ipc_issuing_body": 1, "ipc_fetched": 1}
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        NAVY = '1B3A6B'; GREEN = '2E9E44'; WHITE = 'FFFFFF'
        ALT  = 'EFF6FF'; WARN  = 'FFF3CD'; RED   = 'FFDDDD'

        h_font  = Font(name='Arial', bold=True, color=WHITE, size=10)
        h_fill  = PatternFill('solid', start_color=NAVY, end_color=NAVY)
        h_align = Alignment(horizontal='center', vertical='center')
        b_font  = Font(name='Arial', size=10)
        l_align = Alignment(horizontal='left', vertical='center')
        c_align = Alignment(horizontal='center', vertical='center')
        thin    = Side(style='thin', color='CCCCCC')
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        green_b = Border(left=thin, right=thin, top=thin,
                         bottom=Side(style='medium', color=GREEN))

        wb = Workbook()
        ws = wb.active
        ws.title = 'IPC Certificates'

        headers    = ['Sno', 'Staff Name', 'Email', 'Certificate Name',
                      'Name on Cert', 'Expiry Date', 'Issue Date', 'Issuing Body', 'Status']
        col_widths = [5, 28, 36, 35, 28, 16, 16, 30, 14]

        for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = h_font; cell.fill = h_fill
            cell.alignment = h_align; cell.border = green_b
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:I{len(docs)+1}'

        from datetime import date as _date
        today = _date.today()

        def _is_expired(expiry_str):
            if not expiry_str:
                return None
            for fmt in ('%d/%m/%Y','%m/%Y','%Y-%m-%d','%d-%m-%Y','%B %Y','%b %Y'):
                try:
                    from datetime import datetime as _dt
                    d = _dt.strptime(expiry_str.strip(), fmt).date()
                    return d < today
                except Exception:
                    continue
            return None

        for ri, doc in enumerate(docs, start=2):
            s1       = doc.get('section_1_personal_details') or {}
            name     = _v(s1.get('full_name') or '')
            email    = _v(doc.get('email') or '')
            cert_n   = _v(doc.get('ipc_certificate_name') or '')
            cert_s   = _v(doc.get('ipc_staff_name') or '')
            expiry   = _v(doc.get('ipc_expiry_date') or '')
            issue    = _v(doc.get('ipc_issue_date') or '')
            issuer   = _v(doc.get('ipc_issuing_body') or '')
            fetched  = doc.get('ipc_fetched', False)

            expired  = _is_expired(expiry)

            if not fetched:
                status   = 'Not Checked'
                row_fill = PatternFill('solid', start_color=WARN, end_color=WARN)
            elif not cert_n:
                status   = 'No Cert Found'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is True:
                status   = 'EXPIRED'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is False:
                status   = 'Valid'
                row_fill = None
            else:
                status   = 'Found'
                row_fill = None

            alt_fill = PatternFill('solid', start_color=ALT, end_color=ALT)                        if ri % 2 == 0 and not row_fill else None

            row_vals = [ri-1, name, email, cert_n, cert_s, expiry, issue, issuer, status]
            aligns   = [c_align,l_align,l_align,l_align,l_align,c_align,c_align,l_align,c_align]

            for ci, (val, align) in enumerate(zip(row_vals, aligns), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = b_font; cell.alignment = align
                cell.border = border
                cell.fill = row_fill or alt_fill or PatternFill()

            ws.row_dimensions[ri].height = 17

        ws.cell(row=len(docs)+2, column=1,
                value=f'Total: {len(docs)}').font = Font(name='Arial', bold=True, size=9)

        buf = _io.BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="ipc_certificates_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'}
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500



# ── Cron: Extract Hand Hygiene Certificate ────────────────────────────

@admin_bp.route('/live-staffs/cron/sync-hand-hygiene', methods=['GET', 'POST'])
def live_staff_cron_sync_hand_hygiene():
    """
    Cron job — processes ONE staff member per call.
    Finds "Hand Hygiene" document, extracts certificate details via Gemini.
    Saves: hh_certificate_name, hh_staff_name, hh_expiry_date,
           hh_issue_date, hh_issuing_body, hh_fetched = True
    """
    import requests as _req
    from google import genai as google_genai

    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')
    gemini_key  = os.environ.get('GEMINI_API_KEY', '')

    if not base_url:
        return jsonify({"success": False, "error": "LIVE_STAFF_URL not set"}), 500
    if not gemini_key:
        return jsonify({"success": False, "error": "GEMINI_API_KEY not set"}), 500

    col = _staffs_col()

    pending_query = {
        "$or": [
            {"hh_fetched": {"$exists": False}},
            {"hh_fetched": False},
            {"hh_fetched": None},
        ]
    }
    remaining_total = col.count_documents(pending_query)
    staff           = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff Hand Hygiene certificates already extracted.",
            "remaining_count": 0,
        })

    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or s1.get('email_address') or '')

    def _mark_done(fields):
        fields["hh_fetched"]    = True
        fields["hh_fetched_at"] = datetime.utcnow()
        col.update_one({"_id": staff['_id']}, {"$set": fields})

    if not email:
        _mark_done({"hh_note": "skipped — no email"})
        return jsonify({
            "success":         True,
            "message":         "Skipped — no email",
            "remaining_count": max(0, remaining_total - 1),
        })

    endpoint    = f"{base_url}/ai/recruitments/user-document-list"
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    try:
        resp = _req.post(endpoint, json={"email": email},
                         headers=api_headers, timeout=30)
        if resp.status_code == 405:
            resp = _req.get(endpoint, params={"email": email},
                            headers=api_headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        _mark_done({"hh_note": f"API error: {e}"})
        return jsonify({
            "success": False, "email": email,
            "error": f"API error: {e}",
            "remaining_count": max(0, remaining_total - 1),
        })

    if not data.get('success'):
        _mark_done({"hh_note": f"API error: {data.get('message')}"})
        return jsonify({
            "success": False, "email": email,
            "error": data.get('message', 'API error'),
            "remaining_count": max(0, remaining_total - 1),
        })

    api_data  = data.get('data')
    documents = api_data if isinstance(api_data, list) else                 (api_data.get('documents') or [] if isinstance(api_data, dict) else [])

    if not documents:
        _mark_done({"hh_note": "no documents returned"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": False,
            "message": f"No documents returned for {email}",
            "remaining_count": max(0, remaining_total - 1),
        })

    hh_doc = None
    for d in documents:
        doc_name = (d.get('document_type_name') or '').strip().lower()
        if any(t in doc_name for t in (
            'hand hygiene', 'hand washing', 'hand hygiene certificate',
            'handhygiene', 'hand-hygiene',
        )) and d.get('url'):
            hh_doc = d
            break

    if not hh_doc:
        _mark_done({"hh_note": "no Hand Hygiene document found"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": False,
            "message": f"No Hand Hygiene certificate found for {full_name}",
            "remaining_count": max(0, remaining_total - 1),
        })

    doc_url = (hh_doc.get('url') or '').strip()

    if not doc_url:
        _mark_done({"hh_note": "document found but URL is empty — skipped"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": True, "skipped": True,
            "reason": "Document URL is empty",
            "remaining_count": max(0, remaining_total - 1),
            "message": f"Skipped {full_name} ({email}) — Hand Hygiene doc has no URL",
        })

    try:
        dl_headers = {k: v for k, v in api_headers.items() if k != 'Content-Type'}
        dl_resp    = _req.get(doc_url, headers=dl_headers, timeout=60)

        if dl_resp.status_code == 404:
            _mark_done({"hh_note": "document URL 404 — skipped", "hh_doc_404": True})
            return jsonify({
                "success": True, "email": email, "staff_name": full_name,
                "doc_found": True, "skipped": True,
                "reason": "Document URL returned 404",
                "remaining_count": max(0, remaining_total - 1),
                "message": f"Skipped {full_name} ({email}) — Hand Hygiene doc URL 404",
            })

        dl_resp.raise_for_status()
        raw_bytes    = dl_resp.content
        content_type = dl_resp.headers.get('Content-Type', '').lower()

        client = google_genai.Client(api_key=gemini_key)

        prompt_text = """You are a certificate data extractor.

Extract the following details from this Hand Hygiene certificate:
1. Certificate name (e.g. "Hand Hygiene", "Hand Washing Training Certificate")
2. Staff name as printed on the certificate
3. Expiry date or renewal date (if shown)
4. Issue / completion date
5. Issuing body or training provider

Return ONLY a JSON object — no markdown, no explanation:
{
  "certificate_name": "<exact certificate title as printed>",
  "staff_name_on_cert": "<name as printed on certificate>",
  "expiry_date": "<expiry or renewal date as printed, e.g. 01/06/2027 or June 2027>",
  "issue_date": "<issue or completion date as printed>",
  "issuing_body": "<organization that issued the certificate>"
}

If a field is not visible, set it to null.
"""

        is_image = any(t in content_type for t in ('image/', 'jpeg', 'jpg', 'png', 'webp'))
        is_pdf   = 'pdf' in content_type or doc_url.lower().split('?')[0].endswith('.pdf')

        if is_image:
            ext   = 'jpeg' if any(t in content_type for t in ('jpeg','jpg')) else                     'png'  if 'png'  in content_type else                     'webp' if 'webp' in content_type else 'jpeg'
            parts = [
                {"inline_data": {"mime_type": f"image/{ext}",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash', contents=[{"parts": parts}]
            )
        elif is_pdf:
            parts = [
                {"inline_data": {"mime_type": "application/pdf",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash', contents=[{"parts": parts}]
            )
        else:
            try:
                import io as _io, pdfplumber
                with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                    raw_text = chr(10).join(p.extract_text() or '' for p in pdf.pages).strip()
            except Exception:
                raw_text = raw_bytes.decode('utf-8', errors='replace').strip()
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt_text + "\n\nCERTIFICATE TEXT:\n" + raw_text[:5000]
            )

        raw_out = (response.text or '').strip()
        raw_out = _re.sub(r'^```(?:json)?\s*', '', raw_out, flags=_re.MULTILINE)
        raw_out = _re.sub(r'```\s*$', '', raw_out, flags=_re.MULTILINE).strip()

        result       = _cjson.loads(raw_out)
        cert_name    = _v(result.get('certificate_name') or '')
        cert_staff   = _v(result.get('staff_name_on_cert') or '')
        expiry_date  = _v(result.get('expiry_date') or '')
        issue_date   = _v(result.get('issue_date') or '')
        issuing_body = _v(result.get('issuing_body') or '')

        _mark_done({
            "hh_certificate_name": cert_name,
            "hh_staff_name":       cert_staff,
            "hh_expiry_date":      expiry_date,
            "hh_issue_date":       issue_date,
            "hh_issuing_body":     issuing_body,
            "hh_doc_url":          doc_url,
            "hh_doc_type":         hh_doc.get('document_type_name', ''),
            "hh_note":             "extracted successfully",
        })

        return jsonify({
            "success":            True,
            "email":              email,
            "staff_name":         full_name,
            "doc_found":          True,
            "certificate_name":   cert_name,
            "staff_name_on_cert": cert_staff,
            "expiry_date":        expiry_date,
            "issue_date":         issue_date,
            "issuing_body":       issuing_body,
            "remaining_count":    max(0, remaining_total - 1),
            "message": (
                f"Hand Hygiene cert extracted for {full_name} "
                f"(expires: {expiry_date or 'unknown'}) — "
                f"{max(0, remaining_total - 1)} remaining."
            ),
        })

    except _cjson.JSONDecodeError:
        _mark_done({"hh_note": "Gemini JSON parse error"})
        return jsonify({
            "success": False, "email": email,
            "error": "Gemini returned non-JSON",
            "remaining_count": max(0, remaining_total - 1),
        })
    except Exception as e:
        _mark_done({"hh_note": f"error: {e}"})
        return jsonify({
            "success": False, "email": email,
            "error": str(e),
            "remaining_count": max(0, remaining_total - 1),
        })


# ── Export: Hand Hygiene certificates to Excel ────────────────────────

@admin_bp.route('/live-staffs/export/hand-hygiene-xlsx')
@admin_required
def live_staff_export_hand_hygiene_xlsx():
    """Export Hand Hygiene certificate details to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as _io

        docs = list(_staffs_col().find(
            {},
            {"section_1_personal_details": 1, "email": 1,
             "hh_certificate_name": 1, "hh_staff_name": 1,
             "hh_expiry_date": 1, "hh_issue_date": 1,
             "hh_issuing_body": 1, "hh_fetched": 1}
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        NAVY = '1B3A6B'; GREEN = '2E9E44'; WHITE = 'FFFFFF'
        ALT  = 'EFF6FF'; WARN  = 'FFF3CD'; RED   = 'FFDDDD'

        h_font  = Font(name='Arial', bold=True, color=WHITE, size=10)
        h_fill  = PatternFill('solid', start_color=NAVY, end_color=NAVY)
        h_align = Alignment(horizontal='center', vertical='center')
        b_font  = Font(name='Arial', size=10)
        l_align = Alignment(horizontal='left', vertical='center')
        c_align = Alignment(horizontal='center', vertical='center')
        thin    = Side(style='thin', color='CCCCCC')
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        green_b = Border(left=thin, right=thin, top=thin,
                         bottom=Side(style='medium', color=GREEN))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Hand Hygiene Certificates'

        headers    = ['Sno', 'Staff Name', 'Email', 'Certificate Name',
                      'Name on Cert', 'Expiry Date', 'Issue Date', 'Issuing Body', 'Status']
        col_widths = [5, 28, 36, 32, 28, 16, 16, 30, 14]

        for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = h_font; cell.fill = h_fill
            cell.alignment = h_align; cell.border = green_b
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:I{len(docs)+1}'

        from datetime import date as _date
        today = _date.today()

        def _is_expired(expiry_str):
            if not expiry_str:
                return None
            for fmt in ('%d/%m/%Y','%m/%Y','%Y-%m-%d','%d-%m-%Y','%B %Y','%b %Y'):
                try:
                    from datetime import datetime as _dt
                    d = _dt.strptime(expiry_str.strip(), fmt).date()
                    return d < today
                except Exception:
                    continue
            return None

        for ri, doc in enumerate(docs, start=2):
            s1       = doc.get('section_1_personal_details') or {}
            name     = _v(s1.get('full_name') or '')
            email    = _v(doc.get('email') or '')
            cert_n   = _v(doc.get('hh_certificate_name') or '')
            cert_s   = _v(doc.get('hh_staff_name') or '')
            expiry   = _v(doc.get('hh_expiry_date') or '')
            issue    = _v(doc.get('hh_issue_date') or '')
            issuer   = _v(doc.get('hh_issuing_body') or '')
            fetched  = doc.get('hh_fetched', False)
            expired  = _is_expired(expiry)

            if not fetched:
                status   = 'Not Checked'
                row_fill = PatternFill('solid', start_color=WARN, end_color=WARN)
            elif not cert_n:
                status   = 'No Cert Found'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is True:
                status   = 'EXPIRED'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is False:
                status   = 'Valid'
                row_fill = None
            else:
                status   = 'Found'
                row_fill = None

            alt_fill = PatternFill('solid', start_color=ALT, end_color=ALT)                        if ri % 2 == 0 and not row_fill else None

            row_vals = [ri-1, name, email, cert_n, cert_s, expiry, issue, issuer, status]
            aligns   = [c_align,l_align,l_align,l_align,l_align,c_align,c_align,l_align,c_align]

            for ci, (val, align) in enumerate(zip(row_vals, aligns), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = b_font; cell.alignment = align
                cell.border = border
                cell.fill = row_fill or alt_fill or PatternFill()

            ws.row_dimensions[ri].height = 17

        ws.cell(row=len(docs)+2, column=1,
                value=f'Total: {len(docs)}').font = Font(name='Arial', bold=True, size=9)

        buf = _io.BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="hand_hygiene_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'}
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500



# ── Cron: Extract Children First Certificate ──────────────────────────

@admin_bp.route('/live-staffs/cron/sync-children-first', methods=['GET', 'POST'])
def live_staff_cron_sync_children_first():
    """
    Cron job — processes ONE staff member per call.
    Finds "Children First" document, extracts certificate details via Gemini.
    Saves: cf_certificate_name, cf_staff_name, cf_expiry_date,
           cf_issue_date, cf_issuing_body, cf_fetched = True
    """
    import requests as _req
    from google import genai as google_genai

    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')
    gemini_key  = os.environ.get('GEMINI_API_KEY', '')

    if not base_url:
        return jsonify({"success": False, "error": "LIVE_STAFF_URL not set"}), 500
    if not gemini_key:
        return jsonify({"success": False, "error": "GEMINI_API_KEY not set"}), 500

    col = _staffs_col()

    pending_query = {
        "$or": [
            {"cf_fetched": {"$exists": False}},
            {"cf_fetched": False},
            {"cf_fetched": None},
        ]
    }
    remaining_total = col.count_documents(pending_query)
    staff           = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff Children First certificates already extracted.",
            "remaining_count": 0,
        })

    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or s1.get('email_address') or '')

    def _mark_done(fields):
        fields["cf_fetched"]    = True
        fields["cf_fetched_at"] = datetime.utcnow()
        col.update_one({"_id": staff['_id']}, {"$set": fields})

    if not email:
        _mark_done({"cf_note": "skipped — no email"})
        return jsonify({
            "success":         True,
            "message":         "Skipped — no email",
            "remaining_count": max(0, remaining_total - 1),
        })

    endpoint    = f"{base_url}/ai/recruitments/user-document-list"
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    try:
        resp = _req.post(endpoint, json={"email": email},
                         headers=api_headers, timeout=30)
        if resp.status_code == 405:
            resp = _req.get(endpoint, params={"email": email},
                            headers=api_headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        _mark_done({"cf_note": f"API error: {e}"})
        return jsonify({
            "success": False, "email": email,
            "error": f"API error: {e}",
            "remaining_count": max(0, remaining_total - 1),
        })

    if not data.get('success'):
        _mark_done({"cf_note": f"API error: {data.get('message')}"})
        return jsonify({
            "success": False, "email": email,
            "error": data.get('message', 'API error'),
            "remaining_count": max(0, remaining_total - 1),
        })

    api_data  = data.get('data')
    documents = api_data if isinstance(api_data, list) else                 (api_data.get('documents') or [] if isinstance(api_data, dict) else [])

    if not documents:
        _mark_done({"cf_note": "no documents returned"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": False,
            "message": f"No documents returned for {email}",
            "remaining_count": max(0, remaining_total - 1),
        })

    cf_doc = None
    for d in documents:
        doc_name = (d.get('document_type_name') or '').strip().lower()
        if any(t in doc_name for t in (
            'children first', 'childrenfirst', 'children first certificate',
            'children first training', 'children first e-learning',
        )) and d.get('url'):
            cf_doc = d
            break

    if not cf_doc:
        _mark_done({"cf_note": "no Children First document found"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": False,
            "message": f"No Children First certificate found for {full_name}",
            "remaining_count": max(0, remaining_total - 1),
        })

    doc_url = (cf_doc.get('url') or '').strip()

    if not doc_url:
        _mark_done({"cf_note": "document found but URL is empty — skipped"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": True, "skipped": True,
            "reason": "Document URL is empty",
            "remaining_count": max(0, remaining_total - 1),
            "message": f"Skipped {full_name} ({email}) — Children First doc has no URL",
        })

    try:
        dl_headers = {k: v for k, v in api_headers.items() if k != 'Content-Type'}
        dl_resp    = _req.get(doc_url, headers=dl_headers, timeout=60)

        if dl_resp.status_code == 404:
            _mark_done({"cf_note": "document URL 404 — skipped", "cf_doc_404": True})
            return jsonify({
                "success": True, "email": email, "staff_name": full_name,
                "doc_found": True, "skipped": True,
                "reason": "Document URL returned 404",
                "remaining_count": max(0, remaining_total - 1),
                "message": f"Skipped {full_name} ({email}) — Children First doc URL 404",
            })

        dl_resp.raise_for_status()
        raw_bytes    = dl_resp.content
        content_type = dl_resp.headers.get('Content-Type', '').lower()

        client = google_genai.Client(api_key=gemini_key)

        prompt_text = """You are a certificate data extractor.

Extract the following details from this Children First certificate or training record:
1. Certificate name (e.g. "Children First", "Children First e-Learning", "Safeguarding Children")
2. Staff name as printed on the certificate
3. Expiry date or renewal date (if shown)
4. Issue / completion date
5. Issuing body or training provider

Return ONLY a JSON object — no markdown, no explanation:
{
  "certificate_name": "<exact certificate title as printed>",
  "staff_name_on_cert": "<name as printed on certificate>",
  "expiry_date": "<expiry or renewal date as printed, e.g. 01/06/2027 or June 2027>",
  "issue_date": "<issue or completion date as printed>",
  "issuing_body": "<organization that issued the certificate>"
}

If a field is not visible, set it to null.
"""

        is_image = any(t in content_type for t in ('image/', 'jpeg', 'jpg', 'png', 'webp'))
        is_pdf   = 'pdf' in content_type or doc_url.lower().split('?')[0].endswith('.pdf')

        if is_image:
            ext   = 'jpeg' if any(t in content_type for t in ('jpeg','jpg')) else                     'png'  if 'png'  in content_type else                     'webp' if 'webp' in content_type else 'jpeg'
            parts = [
                {"inline_data": {"mime_type": f"image/{ext}",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash', contents=[{"parts": parts}]
            )
        elif is_pdf:
            parts = [
                {"inline_data": {"mime_type": "application/pdf",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash', contents=[{"parts": parts}]
            )
        else:
            try:
                import io as _io, pdfplumber
                with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                    raw_text = chr(10).join(p.extract_text() or '' for p in pdf.pages).strip()
            except Exception:
                raw_text = raw_bytes.decode('utf-8', errors='replace').strip()
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt_text + "\n\nCERTIFICATE TEXT:\n" + raw_text[:5000]
            )

        raw_out = (response.text or '').strip()
        raw_out = _re.sub(r'^```(?:json)?\s*', '', raw_out, flags=_re.MULTILINE)
        raw_out = _re.sub(r'```\s*$', '', raw_out, flags=_re.MULTILINE).strip()

        result       = _cjson.loads(raw_out)
        cert_name    = _v(result.get('certificate_name') or '')
        cert_staff   = _v(result.get('staff_name_on_cert') or '')
        expiry_date  = _v(result.get('expiry_date') or '')
        issue_date   = _v(result.get('issue_date') or '')
        issuing_body = _v(result.get('issuing_body') or '')

        _mark_done({
            "cf_certificate_name": cert_name,
            "cf_staff_name":       cert_staff,
            "cf_expiry_date":      expiry_date,
            "cf_issue_date":       issue_date,
            "cf_issuing_body":     issuing_body,
            "cf_doc_url":          doc_url,
            "cf_doc_type":         cf_doc.get('document_type_name', ''),
            "cf_note":             "extracted successfully",
        })

        return jsonify({
            "success":            True,
            "email":              email,
            "staff_name":         full_name,
            "doc_found":          True,
            "certificate_name":   cert_name,
            "staff_name_on_cert": cert_staff,
            "expiry_date":        expiry_date,
            "issue_date":         issue_date,
            "issuing_body":       issuing_body,
            "remaining_count":    max(0, remaining_total - 1),
            "message": (
                f"Children First cert extracted for {full_name} "
                f"(expires: {expiry_date or 'unknown'}) — "
                f"{max(0, remaining_total - 1)} remaining."
            ),
        })

    except _cjson.JSONDecodeError:
        _mark_done({"cf_note": "Gemini JSON parse error"})
        return jsonify({
            "success": False, "email": email,
            "error": "Gemini returned non-JSON",
            "remaining_count": max(0, remaining_total - 1),
        })
    except Exception as e:
        _mark_done({"cf_note": f"error: {e}"})
        return jsonify({
            "success": False, "email": email,
            "error": str(e),
            "remaining_count": max(0, remaining_total - 1),
        })


# ── Export: Children First certificates to Excel ──────────────────────

@admin_bp.route('/live-staffs/export/children-first-xlsx')
@admin_required
def live_staff_export_children_first_xlsx():
    """Export Children First certificate details to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as _io

        docs = list(_staffs_col().find(
            {},
            {"section_1_personal_details": 1, "email": 1,
             "cf_certificate_name": 1, "cf_staff_name": 1,
             "cf_expiry_date": 1, "cf_issue_date": 1,
             "cf_issuing_body": 1, "cf_fetched": 1}
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        NAVY = '1B3A6B'; GREEN = '2E9E44'; WHITE = 'FFFFFF'
        ALT  = 'EFF6FF'; WARN  = 'FFF3CD'; RED   = 'FFDDDD'

        h_font  = Font(name='Arial', bold=True, color=WHITE, size=10)
        h_fill  = PatternFill('solid', start_color=NAVY, end_color=NAVY)
        h_align = Alignment(horizontal='center', vertical='center')
        b_font  = Font(name='Arial', size=10)
        l_align = Alignment(horizontal='left', vertical='center')
        c_align = Alignment(horizontal='center', vertical='center')
        thin    = Side(style='thin', color='CCCCCC')
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        green_b = Border(left=thin, right=thin, top=thin,
                         bottom=Side(style='medium', color=GREEN))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Children First Certificates'

        headers    = ['Sno', 'Staff Name', 'Email', 'Certificate Name',
                      'Name on Cert', 'Expiry Date', 'Issue Date', 'Issuing Body', 'Status']
        col_widths = [5, 28, 36, 32, 28, 16, 16, 30, 14]

        for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = h_font; cell.fill = h_fill
            cell.alignment = h_align; cell.border = green_b
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:I{len(docs)+1}'

        from datetime import date as _date
        today = _date.today()

        def _is_expired(expiry_str):
            if not expiry_str:
                return None
            for fmt in ('%d/%m/%Y','%m/%Y','%Y-%m-%d','%d-%m-%Y','%B %Y','%b %Y'):
                try:
                    from datetime import datetime as _dt
                    d = _dt.strptime(expiry_str.strip(), fmt).date()
                    return d < today
                except Exception:
                    continue
            return None

        for ri, doc in enumerate(docs, start=2):
            s1       = doc.get('section_1_personal_details') or {}
            name     = _v(s1.get('full_name') or '')
            email    = _v(doc.get('email') or '')
            cert_n   = _v(doc.get('cf_certificate_name') or '')
            cert_s   = _v(doc.get('cf_staff_name') or '')
            expiry   = _v(doc.get('cf_expiry_date') or '')
            issue    = _v(doc.get('cf_issue_date') or '')
            issuer   = _v(doc.get('cf_issuing_body') or '')
            fetched  = doc.get('cf_fetched', False)
            expired  = _is_expired(expiry)

            if not fetched:
                status   = 'Not Checked'
                row_fill = PatternFill('solid', start_color=WARN, end_color=WARN)
            elif not cert_n:
                status   = 'No Cert Found'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is True:
                status   = 'EXPIRED'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is False:
                status   = 'Valid'
                row_fill = None
            else:
                status   = 'Found'
                row_fill = None

            alt_fill = PatternFill('solid', start_color=ALT, end_color=ALT)                        if ri % 2 == 0 and not row_fill else None

            row_vals = [ri-1, name, email, cert_n, cert_s, expiry, issue, issuer, status]
            aligns   = [c_align,l_align,l_align,l_align,l_align,c_align,c_align,l_align,c_align]

            for ci, (val, align) in enumerate(zip(row_vals, aligns), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = b_font; cell.alignment = align
                cell.border = border
                cell.fill = row_fill or alt_fill or PatternFill()

            ws.row_dimensions[ri].height = 17

        ws.cell(row=len(docs)+2, column=1,
                value=f'Total: {len(docs)}').font = Font(name='Arial', bold=True, size=9)

        buf = _io.BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="children_first_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'}
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500



# ── Cron: Push passport number to DOC API one staff at a time ─────────

@admin_bp.route('/live-staffs/cron/push-passport-number', methods=['GET', 'POST'])
def live_staff_cron_push_passport_number():
    """
    Cron job — processes ONE staff member per call.

    Finds staff where:
      - passport_id is set (already extracted by sync-passport cron)
      - staff_id (XN Portal staff ID) exists
      - passport_number_pushed is not True (not yet pushed to DOC API)

    Calls:
      POST {DOC_BASE_URL}/api/staff/passport-number-update
      Headers: Api-Key, X-App-Country
      Form:    staff_id, passport_number

    Sets passport_number_pushed = True on success.
    Stores passport_push_email in live_staffs for reference.

    Protect with ?cron_key=<CRON_SECRET> env var.
    """
    import requests as _req

    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    doc_base_url = os.environ.get('DOC_BASE_URL', '').rstrip('/')
    doc_api_key  = os.environ.get('PASSPORT_API_KEY', '') or os.environ.get('DOC_API_KEY', '')
    app_country  = os.environ.get('XN_APP_COUNTRY', 'ie')

    if not doc_base_url:
        return jsonify({"success": False, "error": "DOC_BASE_URL not set"}), 500
    if not doc_api_key:
        return jsonify({"success": False, "error": "PASSPORT_API_KEY not set in .env"}), 500

    col = _staffs_col()

    # Only process staff that:
    # 1. Have a passport_id extracted
    # 2. Have a staff_id (XN Portal)
    # 3. Have not been pushed yet
    pending_query = {
        "$and": [
            {"passport_id": {"$exists": True}},
            {"passport_id": {"$ne": None}},
            {"passport_id": {"$ne": ""}},
            {"$or": [
                {"staff_id": {"$exists": True}},
                {"xn_staff_id": {"$exists": True}},
            ]},
            {"$or": [
                {"passport_number_pushed": {"$exists": False}},
                {"passport_number_pushed": False},
                {"passport_number_pushed": None},
            ]},
        ]
    }

    remaining_total = col.count_documents(pending_query)
    staff           = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All passport numbers already pushed — nothing to do.",
            "remaining_count": 0,
        })

    s1          = staff.get('section_1_personal_details') or {}
    full_name   = _v(s1.get('full_name') or '')
    email       = _v(staff.get('email') or s1.get('email_address') or '')
    passport_id = _v(staff.get('passport_id') or '')
    xn_staff_id = _v(staff.get('staff_id') or staff.get('xn_staff_id') or '')

    def _mark_done(fields):
        fields["passport_push_attempted_at"] = datetime.utcnow()
        fields["passport_push_email"]        = email
        col.update_one({"_id": staff['_id']}, {"$set": fields})

    if not xn_staff_id:
        _mark_done({
            "passport_number_pushed": False,
            "passport_push_note":     "skipped — no staff_id found",
        })
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "skipped":         True,
            "reason":          "No staff_id / xn_staff_id found",
            "remaining_count": max(0, remaining_total - 1),
            "message":         f"Skipped {full_name} ({email}) — no staff_id",
        })

    if not passport_id:
        _mark_done({
            "passport_number_pushed": False,
            "passport_push_note":     "skipped — passport_id is empty",
        })
        return jsonify({
            "success":         True,
            "email":           email,
            "staff_name":      full_name,
            "skipped":         True,
            "reason":          "passport_id is empty",
            "remaining_count": max(0, remaining_total - 1),
            "message":         f"Skipped {full_name} ({email}) — passport_id empty",
        })

    # ── Call DOC API ──────────────────────────────────────────────────
    url     = f"{doc_base_url}/api/staff/passport-number-update"
    headers = {
        "Api-Key":       doc_api_key,
        "X-App-Country": app_country,
    }
    form_data = {
        "staff_id":        xn_staff_id,
        "passport_number": passport_id,
    }

    try:
        resp = _req.post(url, headers=headers, data=form_data, timeout=30)

        if resp.status_code in (200, 201):
            try:
                resp_data = resp.json()
            except Exception:
                resp_data = {"raw": resp.text[:200]}

            _mark_done({
                "passport_number_pushed":      True,
                "passport_push_note":          "pushed successfully",
                "passport_push_response":      resp_data,
            })

            return jsonify({
                "success":         True,
                "email":           email,
                "staff_name":      full_name,
                "staff_id":        xn_staff_id,
                "passport_number": passport_id,
                "pushed":          True,
                "api_response":    resp_data,
                "remaining_count": max(0, remaining_total - 1),
                "message": (
                    f"Passport number pushed for {full_name} ({email}) "
                    f"[{passport_id}] — {max(0, remaining_total - 1)} remaining."
                ),
            })

        else:
            # Non-200 — log full response
            try:
                err_json = resp.json()
            except Exception:
                err_json = None
            err_text = resp.text[:500]

            # 422 = invalid staff_id — skip permanently, do not retry
            if resp.status_code == 422:
                _mark_done({
                    "passport_number_pushed": True,   # mark done so cron skips next time
                    "passport_push_note":     f"422 invalid staff_id — skipped permanently",
                    "passport_push_response": err_json or err_text,
                })
                return jsonify({
                    "success":         True,
                    "email":           email,
                    "staff_name":      full_name,
                    "staff_id":        xn_staff_id,
                    "passport_number": passport_id,
                    "pushed":          False,
                    "skipped":         True,
                    "reason":          "422 — invalid staff_id, skipped permanently",
                    "http_status":     422,
                    "api_error":       err_text,
                    "remaining_count": max(0, remaining_total - 1),
                    "message":         f"Skipped {full_name} ({email}) — invalid staff_id (422)",
                })

            _mark_done({
                "passport_number_pushed":  False,
                "passport_push_note":      f"API returned {resp.status_code}: {err_text}",
                "passport_push_response":  err_json or err_text,
            })
            return jsonify({
                "success":         False,
                "email":           email,
                "staff_name":      full_name,
                "staff_id":        xn_staff_id,
                "passport_number": passport_id,
                "pushed":          False,
                "http_status":     resp.status_code,
                "api_error":       err_text,
                "api_error_json":  err_json,
                "remaining_count": max(0, remaining_total - 1),
                "message":         f"API error {resp.status_code} for {full_name} ({email}): {err_text}",
            })

    except Exception as e:
        _mark_done({
            "passport_number_pushed": False,
            "passport_push_note":     f"request error: {e}",
        })
        return jsonify({
            "success":         False,
            "email":           email,
            "staff_name":      full_name,
            "error":           str(e),
            "remaining_count": max(0, remaining_total - 1),
        })



@admin_bp.route('/live-staffs/export/safeguarding-xlsx')
@admin_required
def live_staff_export_safeguarding_xlsx():
    """Export Safeguarding Adults At Risk certificate details to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as _io

        docs = list(_staffs_col().find(
            {},
            {"section_1_personal_details": 1, "email": 1,
             "sg_certificate_name": 1, "sg_staff_name": 1,
             "sg_expiry_date": 1, "sg_issue_date": 1,
             "sg_issuing_body": 1, "sg_fetched": 1}
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        NAVY = '1B3A6B'; GREEN = '2E9E44'; WHITE = 'FFFFFF'
        ALT  = 'EFF6FF'; WARN  = 'FFF3CD'; RED   = 'FFDDDD'

        h_font  = Font(name='Arial', bold=True, color=WHITE, size=10)
        h_fill  = PatternFill('solid', start_color=NAVY, end_color=NAVY)
        h_align = Alignment(horizontal='center', vertical='center')
        b_font  = Font(name='Arial', size=10)
        l_align = Alignment(horizontal='left',   vertical='center')
        c_align = Alignment(horizontal='center', vertical='center')
        thin    = Side(style='thin', color='CCCCCC')
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        green_b = Border(left=thin, right=thin, top=thin,
                         bottom=Side(style='medium', color=GREEN))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Safeguarding Certificates'

        headers    = ['Sno', 'Staff Name', 'Email', 'Certificate Name',
                      'Name on Cert', 'Expiry Date', 'Issue Date', 'Issuing Body', 'Status']
        col_widths = [5, 28, 36, 35, 28, 16, 16, 30, 14]

        for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = h_font; cell.fill = h_fill
            cell.alignment = h_align; cell.border = green_b
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:I{len(docs)+1}'

        from datetime import date as _date
        today = _date.today()

        def _is_expired(expiry_str):
            if not expiry_str:
                return None
            for fmt in ('%d/%m/%Y', '%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%B %Y', '%b %Y'):
                try:
                    from datetime import datetime as _dt
                    d = _dt.strptime(expiry_str.strip(), fmt).date()
                    return d < today
                except Exception:
                    continue
            return None

        for ri, doc in enumerate(docs, start=2):
            s1       = doc.get('section_1_personal_details') or {}
            name     = _v(s1.get('full_name') or '')
            email    = _v(doc.get('email') or '')
            cert_n   = _v(doc.get('sg_certificate_name') or '')
            cert_s   = _v(doc.get('sg_staff_name') or '')
            expiry   = _v(doc.get('sg_expiry_date') or '')
            issue    = _v(doc.get('sg_issue_date') or '')
            issuer   = _v(doc.get('sg_issuing_body') or '')
            fetched  = doc.get('sg_fetched', False)
            expired  = _is_expired(expiry)

            if not fetched:
                status   = 'Not Checked'
                row_fill = PatternFill('solid', start_color=WARN, end_color=WARN)
            elif not cert_n:
                status   = 'No Cert Found'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is True:
                status   = 'EXPIRED'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is False:
                status   = 'Valid'
                row_fill = None
            else:
                status   = 'Found'
                row_fill = None

            alt_fill = PatternFill('solid', start_color=ALT, end_color=ALT) \
                       if ri % 2 == 0 and not row_fill else None

            row_vals = [ri-1, name, email, cert_n, cert_s, expiry, issue, issuer, status]
            aligns   = [c_align, l_align, l_align, l_align, l_align,
                        c_align, c_align, l_align, c_align]

            for ci, (val, align) in enumerate(zip(row_vals, aligns), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = b_font; cell.alignment = align
                cell.border = border
                cell.fill = row_fill or alt_fill or PatternFill()

            ws.row_dimensions[ri].height = 17

        ws.cell(row=len(docs)+2, column=1,
                value=f'Total: {len(docs)}').font = Font(name='Arial', bold=True, size=9)

        buf = _io.BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="safeguarding_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'}
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route('/live-staffs/cron/sync-safeguarding', methods=['GET', 'POST'])
def live_staff_cron_sync_safeguarding():
    """
    Cron job — processes ONE staff member per call.
    Finds "Safeguarding Adults At Risk" document, extracts details via Gemini.
    Sets sg_fetched = True on every outcome (success, skip, error).
    """
    import requests as _req
    from google import genai as google_genai

    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')
    gemini_key  = os.environ.get('GEMINI_API_KEY', '')

    if not base_url:
        return jsonify({"success": False, "error": "LIVE_STAFF_URL not set"}), 500
    if not gemini_key:
        return jsonify({"success": False, "error": "GEMINI_API_KEY not set"}), 500

    col = _staffs_col()

    pending_query = {
        "$or": [
            {"sg_fetched": {"$exists": False}},
            {"sg_fetched": False},
            {"sg_fetched": None},
        ]
    }
    remaining_total = col.count_documents(pending_query)
    staff           = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff Safeguarding certificates already extracted.",
            "remaining_count": 0,
        })

    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or s1.get('email_address') or '')

    def _mark_done(fields):
        fields["sg_fetched"]    = True
        fields["sg_fetched_at"] = datetime.utcnow()
        col.update_one({"_id": staff['_id']}, {"$set": fields})

    if not email:
        _mark_done({"sg_note": "skipped — no email"})
        return jsonify({
            "success":         True,
            "message":         "Skipped — no email",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Call XN Portal API ────────────────────────────────────────────
    endpoint    = f"{base_url}/ai/recruitments/user-document-list"
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    try:
        resp = _req.post(endpoint, json={"email": email},
                         headers=api_headers, timeout=30)
        if resp.status_code == 405:
            resp = _req.get(endpoint, params={"email": email},
                            headers=api_headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        _mark_done({"sg_note": f"API error: {e}"})
        return jsonify({
            "success": False, "email": email,
            "error": f"API error: {e}",
            "remaining_count": max(0, remaining_total - 1),
        })

    if not data.get('success'):
        _mark_done({"sg_note": f"API error: {data.get('message')}"})
        return jsonify({
            "success": False, "email": email,
            "error": data.get('message', 'API error'),
            "remaining_count": max(0, remaining_total - 1),
        })

    api_data  = data.get('data')
    documents = api_data if isinstance(api_data, list) else \
                (api_data.get('documents') or [] if isinstance(api_data, dict) else [])

    if not documents:
        _mark_done({"sg_note": "no documents returned"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": False,
            "message": f"No documents returned for {email}",
            "remaining_count": max(0, remaining_total - 1),
        })

    # ── Find Safeguarding document ────────────────────────────────────
    sg_doc = None
    for d in documents:
        doc_name = (d.get('document_type_name') or '').strip().lower()
        if any(t in doc_name for t in (
            'safeguarding adults at risk',
            'safeguarding adults',
            'safeguarding at risk',
            'safeguarding certificate',
            'safeguarding training',
            'safeguarding',
        )) and d.get('url'):
            sg_doc = d
            break

    if not sg_doc:
        _mark_done({"sg_note": "no Safeguarding document found"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": False,
            "message": f"No Safeguarding Adults At Risk certificate found for {full_name}",
            "remaining_count": max(0, remaining_total - 1),
        })

    doc_url = (sg_doc.get('url') or '').strip()

    if not doc_url:
        _mark_done({"sg_note": "document found but URL is empty — skipped"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": True, "skipped": True,
            "reason": "Document URL is empty",
            "remaining_count": max(0, remaining_total - 1),
            "message": f"Skipped {full_name} ({email}) — Safeguarding doc has no URL",
        })

    # ── Download ──────────────────────────────────────────────────────
    try:
        dl_headers = {k: v for k, v in api_headers.items() if k != 'Content-Type'}
        dl_resp    = _req.get(doc_url, headers=dl_headers, timeout=60)

        if dl_resp.status_code == 404:
            _mark_done({"sg_note": "document URL 404 — skipped", "sg_doc_404": True})
            return jsonify({
                "success": True, "email": email, "staff_name": full_name,
                "doc_found": True, "skipped": True,
                "reason": "Document URL returned 404",
                "remaining_count": max(0, remaining_total - 1),
                "message": f"Skipped {full_name} ({email}) — Safeguarding doc URL 404",
            })

        dl_resp.raise_for_status()
        raw_bytes    = dl_resp.content
        content_type = dl_resp.headers.get('Content-Type', '').lower()

        client = google_genai.Client(api_key=gemini_key)

        prompt_text = """You are a certificate data extractor.

Extract the following details from this Safeguarding Adults At Risk certificate:
1. Certificate name (e.g. "Safeguarding Adults At Risk", "Safeguarding Training Certificate")
2. Staff name as printed on the certificate
3. Expiry date or renewal date (if shown)
4. Issue / completion date
5. Issuing body or training provider

Return ONLY a JSON object — no markdown, no explanation:
{
  "certificate_name": "<exact certificate title as printed>",
  "staff_name_on_cert": "<name as printed on certificate>",
  "expiry_date": "<expiry or renewal date as printed, e.g. 01/06/2027 or June 2027>",
  "issue_date": "<issue or completion date as printed>",
  "issuing_body": "<organization that issued the certificate>"
}

If a field is not visible, set it to null.
"""

        is_image = any(t in content_type for t in ('image/', 'jpeg', 'jpg', 'png', 'webp'))
        is_pdf   = 'pdf' in content_type or doc_url.lower().split('?')[0].endswith('.pdf')

        if is_image:
            ext   = 'jpeg' if any(t in content_type for t in ('jpeg', 'jpg')) else \
                    'png'  if 'png'  in content_type else \
                    'webp' if 'webp' in content_type else 'jpeg'
            parts = [
                {"inline_data": {"mime_type": f"image/{ext}",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash', contents=[{"parts": parts}]
            )
        elif is_pdf:
            parts = [
                {"inline_data": {"mime_type": "application/pdf",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash', contents=[{"parts": parts}]
            )
        else:
            try:
                import io as _io, pdfplumber
                with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                    raw_text = chr(10).join(p.extract_text() or '' for p in pdf.pages).strip()
            except Exception:
                raw_text = raw_bytes.decode('utf-8', errors='replace').strip()
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt_text + "\n\nCERTIFICATE TEXT:\n" + raw_text[:5000]
            )

        raw_out = (response.text or '').strip()
        raw_out = _re.sub(r'^```(?:json)?\s*', '', raw_out, flags=_re.MULTILINE)
        raw_out = _re.sub(r'```\s*$', '', raw_out, flags=_re.MULTILINE).strip()

        result       = _cjson.loads(raw_out)
        cert_name    = _v(result.get('certificate_name') or '')
        cert_staff   = _v(result.get('staff_name_on_cert') or '')
        expiry_date  = _v(result.get('expiry_date') or '')
        issue_date   = _v(result.get('issue_date') or '')
        issuing_body = _v(result.get('issuing_body') or '')

        _mark_done({
            "sg_certificate_name": cert_name,
            "sg_staff_name":       cert_staff,
            "sg_expiry_date":      expiry_date,
            "sg_issue_date":       issue_date,
            "sg_issuing_body":     issuing_body,
            "sg_doc_url":          doc_url,
            "sg_doc_type":         sg_doc.get('document_type_name', ''),
            "sg_note":             "extracted successfully",
        })

        return jsonify({
            "success":            True,
            "email":              email,
            "staff_name":         full_name,
            "doc_found":          True,
            "certificate_name":   cert_name,
            "staff_name_on_cert": cert_staff,
            "expiry_date":        expiry_date,
            "issue_date":         issue_date,
            "issuing_body":       issuing_body,
            "remaining_count":    max(0, remaining_total - 1),
            "message": (
                f"Safeguarding cert extracted for {full_name} "
                f"(expires: {expiry_date or 'unknown'}) — "
                f"{max(0, remaining_total - 1)} remaining."
            ),
        })

    except _cjson.JSONDecodeError:
        _mark_done({"sg_note": "Gemini JSON parse error"})
        return jsonify({
            "success": False, "email": email,
            "error": "Gemini returned non-JSON",
            "remaining_count": max(0, remaining_total - 1),
        })
    except Exception as e:
        _mark_done({"sg_note": f"error: {e}"})
        return jsonify({
            "success": False, "email": email,
            "error": str(e),
            "remaining_count": max(0, remaining_total - 1),
        })


# ── Export ────────────────────────────────────────────────────────────


# ── Cron: Extract Garda Vetting Document details ─────────────────────

