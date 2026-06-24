from datetime import datetime, timezone
import re
import csv
import io

from flask import render_template, request, jsonify
from bson import ObjectId

from database import db
from . import admin_bp
from admin.views import admin_required


# ── Config ───────────────────────────────────────────────────────────
# The browser triggers the actual reminder call by hitting this endpoint
# on the same origin (mirrors the existing trigger_ai_call flow). This is
# a compliance-document reminder asking the user to submit their PCC /
# Garda Vetting certificate. Change this string if a dedicated endpoint
# is created for certificate reminders.
REMINDER_CALL_ENDPOINT = "compliance_document_call"

# Certificate types a user can be reminded about.
VALID_CERTS = ["PCC", "Garda Vetting"]

# Allowed call-status values stored on each reminder record.
VALID_STATUSES = {"pending", "triggered", "completed", "failed"}


# ── Collections ──────────────────────────────────────────────────────
def _users_col():
    return db.users


def _reminders_col():
    return db.certificate_reminder_calls


# ── Shared helpers ───────────────────────────────────────────────────
def _upsert_reminder(user_ref_id, name, phone, xn_user_id, certs,
                     now=None, source="manual"):
    """
    Insert or update a reminder record keyed on user_ref_id, so re-adding
    a user updates their required certificates instead of duplicating.
    Existing call_status is preserved on update.
    """
    now = now or datetime.now(timezone.utc)
    _reminders_col().update_one(
        {"user_ref_id": str(user_ref_id)},
        {
            "$set": {
                "name":                (name or "—"),
                "phone":               (phone or ""),
                "xn_user_id":          xn_user_id,
                "certificates_needed": certs,
                "updated_at":          now,
            },
            "$setOnInsert": {
                "user_ref_id": str(user_ref_id),
                "call_status": "pending",
                "created_at":  now,
                "source":      source,
            },
        },
        upsert=True,
    )


# Cell values that mean "not flagged" in the import sheet.
_NEGATIVE_TOKENS = {"", "0", "no", "n", "false", "f", "none", "nan", "na", "n/a", "-"}


def _is_flagged(value):
    """A cell counts as flagged if it holds any affirmative marker (x, yes, 1, ✓, …)."""
    if value is None:
        return False
    return str(value).strip().lower() not in _NEGATIVE_TOKENS


def _find_col(col_idx, *names):
    """Return the index of the first header name found in col_idx, else None."""
    for n in names:
        if n in col_idx:
            return col_idx[n]
    return None


def _parse_table(file_storage):
    """
    Read an uploaded .xlsx or .csv into (header_list, data_rows).
    Raises ValueError with a clear message on unreadable input.
    """
    filename = (file_storage.filename or "").lower()
    raw = file_storage.read()

    if filename.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        if not rows:
            return [], []
        return rows[0], rows[1:]

    # default: treat as xlsx
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("openpyxl is not installed — run: pip install openpyxl")

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    return list(rows[0]), [list(r) for r in rows[1:]]


# ── Pages ────────────────────────────────────────────────────────────
@admin_bp.route('/certificate_reminders')
@admin_required
def certificate_reminders():
    return render_template(
        'admin/cert_reminders.html',
        call_endpoint=REMINDER_CALL_ENDPOINT,
        cert_types=VALID_CERTS,
    )


# ── Search users ─────────────────────────────────────────────────────
@admin_bp.route('/certificate_reminders/search')
@admin_required
def certificate_reminders_search():
    q = request.args.get('q', '').strip()

    if not q:
        return jsonify({"success": True, "users": []})

    pattern = re.compile(re.escape(q), re.IGNORECASE)

    or_conditions = [
        {"first_name": pattern},
        {"last_name":  pattern},
        {"email":      pattern},
        {"phone":      pattern},
    ]

    # If query contains a space, also try matching first + last name split
    parts = q.split(None, 1)  # split on whitespace, max 2 parts
    if len(parts) == 2:
        first_pat = re.compile(re.escape(parts[0]), re.IGNORECASE)
        last_pat  = re.compile(re.escape(parts[1]), re.IGNORECASE)
        or_conditions.append({
            "first_name": first_pat,
            "last_name":  last_pat,
        })

    query = {"$or": or_conditions}

    try:
        items = list(
            _users_col()
            .find(query, {
                "_id":         1,
                "xn_user_id":  1,
                "first_name":  1,
                "last_name":   1,
                "email":       1,
                "phone":       1,
                "designation": 1,
                "country":     1,
                "created_at":  1,
            })
            .sort([("created_at", -1)])
            .limit(50)
        )

        # Figure out which of these users are already in the reminder list,
        # so the UI can show their existing certificates / call status.
        ids = [str(u["_id"]) for u in items]
        queued = {}
        if ids:
            for r in _reminders_col().find(
                {"user_ref_id": {"$in": ids}},
                {"user_ref_id": 1, "certificates_needed": 1, "call_status": 1},
            ):
                queued[r["user_ref_id"]] = {
                    "certificates_needed": r.get("certificates_needed", []),
                    "call_status":         r.get("call_status", "pending"),
                }

        # Serialize ObjectId and dates + attach queue info
        for u in items:
            u["_id"] = str(u["_id"])
            if "created_at" in u and hasattr(u["created_at"], "isoformat"):
                u["created_at"] = u["created_at"].isoformat()
            u["queued"] = queued.get(u["_id"])

        return jsonify({"success": True, "users": items})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ── Save a selected group to the reminder collection ─────────────────
@admin_bp.route('/certificate_reminders/save', methods=['POST'])
@admin_required
def certificate_reminders_save():
    data = request.get_json(silent=True) or {}
    selections = data.get('users', [])

    if not selections:
        return jsonify({"success": False, "error": "No users selected"}), 400

    now = datetime.now(timezone.utc)
    saved = 0
    skipped = []

    try:
        for s in selections:
            ref_id = str(s.get('user_id', '')).strip()
            certs = [c for c in (s.get('certificates') or []) if c in VALID_CERTS]

            if not ref_id or not certs:
                skipped.append({
                    "user_id": ref_id,
                    "name": s.get('name'),
                    "reason": "missing user id or no certificate selected",
                })
                continue

            name = (s.get('name') or '').strip() or '—'
            phone = (s.get('phone') or '').strip()
            xn_user_id = s.get('xn_user_id')

            _upsert_reminder(ref_id, name, phone, xn_user_id, certs, now=now)
            saved += 1

        return jsonify({"success": True, "saved": saved, "skipped": skipped})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ── Import from Excel / CSV ──────────────────────────────────────────
@admin_bp.route('/certificate_reminders/import', methods=['POST'])
@admin_required
def certificate_reminders_import():
    """
    Accept an .xlsx/.csv with columns: email, PCC, Garda Vetting, Both.
    For each row, look up the user by email and upsert a reminder record.
    """
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    f = request.files['file']
    if not f or not f.filename:
        return jsonify({"success": False, "error": "No file selected"}), 400

    try:
        header, data_rows = _parse_table(f)
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": f"Could not read file: {e}"}), 400

    if not header:
        return jsonify({"success": False, "error": "The file is empty"}), 400

    # Map header names (case-insensitive) to column indexes
    col_idx = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        col_idx[str(h).strip().lower()] = i

    email_i = _find_col(col_idx, "email", "e-mail", "email address")
    pcc_i   = _find_col(col_idx, "pcc", "police clearance", "police clearance certificate")
    garda_i = _find_col(col_idx, "garda vetting", "garda", "garda_vetting",
                        "gardavetting", "garda vetting certificate")
    both_i  = _find_col(col_idx, "both")

    if email_i is None:
        return jsonify({"success": False,
                        "error": "No 'email' column found in the file"}), 400
    if pcc_i is None and garda_i is None and both_i is None:
        return jsonify({"success": False,
                        "error": "No 'PCC', 'Garda Vetting', or 'Both' column found"}), 400

    def cell(row, i):
        return row[i] if (i is not None and i < len(row)) else None

    now = datetime.now(timezone.utc)
    imported = 0
    not_found = []   # email not in users collection
    no_cert = []     # row has no certificate flagged

    try:
        for n, row in enumerate(data_rows, start=2):   # row 2 = first data row under the header
            email = cell(row, email_i)
            if email is None or str(email).strip() == "":
                continue   # blank line

            email = str(email).strip()

            if _is_flagged(cell(row, both_i)):
                certs = ["PCC", "Garda Vetting"]
            else:
                certs = []
                if _is_flagged(cell(row, pcc_i)):
                    certs.append("PCC")
                if _is_flagged(cell(row, garda_i)):
                    certs.append("Garda Vetting")

            if not certs:
                no_cert.append({"row": n, "email": email})
                continue

            user = _users_col().find_one(
                {"email": {"$regex": f"^{re.escape(email)}$", "$options": "i"}},
                {"_id": 1, "xn_user_id": 1, "first_name": 1, "last_name": 1, "phone": 1},
            )
            if not user:
                not_found.append({"row": n, "email": email})
                continue

            name = " ".join(
                p for p in [user.get("first_name"), user.get("last_name")] if p
            ).strip() or "—"

            _upsert_reminder(
                user["_id"], name, user.get("phone"),
                user.get("xn_user_id"), certs, now=now, source="import",
            )
            imported += 1

        return jsonify({
            "success":     True,
            "imported":    imported,
            "not_found":   not_found,
            "no_cert":     no_cert,
            "total_rows":  len(data_rows),
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ── List queued reminders ────────────────────────────────────────────
@admin_bp.route('/certificate_reminders/list')
@admin_required
def certificate_reminders_list():
    status = request.args.get('status', '').strip()
    query = {}
    if status in VALID_STATUSES:
        query["call_status"] = status

    try:
        items = list(
            _reminders_col()
            .find(query)
            .sort([("created_at", -1)])
            .limit(500)
        )

        for r in items:
            r["_id"] = str(r["_id"])
            for f in ("created_at", "updated_at", "last_triggered_at"):
                if f in r and hasattr(r[f], "isoformat"):
                    r[f] = r[f].isoformat()

        return jsonify({"success": True, "reminders": items})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ── Update call status after a call is triggered ─────────────────────
@admin_bp.route('/certificate_reminders/mark', methods=['POST'])
@admin_required
def certificate_reminders_mark():
    data = request.get_json(silent=True) or {}
    reminder_id = str(data.get('id', '')).strip()
    status = str(data.get('status', '')).strip()

    if status not in VALID_STATUSES:
        return jsonify({"success": False, "error": "Invalid status"}), 400

    try:
        oid = ObjectId(reminder_id)
    except Exception:
        return jsonify({"success": False, "error": "Invalid reminder id"}), 400

    now = datetime.now(timezone.utc)
    update = {"call_status": status, "updated_at": now}
    if status == "triggered":
        update["last_triggered_at"] = now
        update["trigger_error"] = None
    if "error" in data:
        update["trigger_error"] = data.get("error")

    try:
        res = _reminders_col().update_one({"_id": oid}, {"$set": update})
        if res.matched_count == 0:
            return jsonify({"success": False, "error": "Reminder not found"}), 404
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ── Remove a queued reminder ─────────────────────────────────────────
@admin_bp.route('/certificate_reminders/delete', methods=['POST'])
@admin_required
def certificate_reminders_delete():
    data = request.get_json(silent=True) or {}
    reminder_id = str(data.get('id', '')).strip()

    try:
        oid = ObjectId(reminder_id)
    except Exception:
        return jsonify({"success": False, "error": "Invalid reminder id"}), 400

    try:
        _reminders_col().delete_one({"_id": oid})
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
