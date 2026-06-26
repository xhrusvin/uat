from flask import request, jsonify, Response
from datetime import datetime
import os

from database import db
from . import admin_bp
from admin.views import admin_required


# ── Helpers ───────────────────────────────────────────────────────────

def _staffs_col():
    return db.live_staffs


def _v(val):
    """Return value as string, or empty string if None/empty."""
    if val is None:
        return ''
    return str(val).strip()


# ── Cron: Sync Open Disclosure certificates ───────────────────────────

@admin_bp.route('/live-staffs/cron/sync-open-disclosure', methods=['GET', 'POST'])
def live_staff_cron_sync_open_disclosure():
    """
    Cron job — processes ONE staff member per call.
    Finds "Open Disclosure" document, extracts certificate details via Gemini.
    Saves: od_certificate_name, od_staff_name, od_expiry_date,
           od_issue_date, od_issuing_body, od_fetched = True
    """
    import requests as _req
    import json as _json, re as _re, base64
    from google import genai as google_genai

    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')
    gemini_key  = os.environ.get('GEMINI_API_KEY', '')

    if not base_url:
        return jsonify({"success": False, "error": "LIVE_STAFF_URL not set"}), 500
    if not gemini_key:
        return jsonify({"success": False, "error": "GEMINI_API_KEY not set"}), 500

    col = _staffs_col()

    pending_query = {
        "$or": [
            {"od_fetched": {"$exists": False}},
            {"od_fetched": False},
            {"od_fetched": None},
        ]
    }
    remaining_total = col.count_documents(pending_query)
    staff           = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff Open Disclosure certificates already extracted.",
            "remaining_count": 0,
        })

    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or s1.get('email_address') or '')

    def _mark_done(fields):
        fields["od_fetched"]    = True
        fields["od_fetched_at"] = datetime.utcnow()
        col.update_one({"_id": staff['_id']}, {"$set": fields})

    if not email:
        _mark_done({"od_note": "skipped — no email"})
        return jsonify({
            "success":         True,
            "message":         "Skipped — no email",
            "remaining_count": max(0, remaining_total - 1),
        })

    endpoint    = f"{base_url}/ai/recruitments/user-document-list"
    api_headers = {
        "Api-Key":       api_key,
        "X-App-Country": app_country,
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }

    try:
        resp = _req.post(endpoint, json={"email": email},
                         headers=api_headers, timeout=30)
        if resp.status_code == 405:
            resp = _req.get(endpoint, params={"email": email},
                            headers=api_headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        _mark_done({"od_note": f"API error: {e}"})
        return jsonify({
            "success": False, "email": email,
            "error": f"API error: {e}",
            "remaining_count": max(0, remaining_total - 1),
        })

    if not data.get('success'):
        _mark_done({"od_note": f"API error: {data.get('message')}"})
        return jsonify({
            "success": False, "email": email,
            "error": data.get('message', 'API error'),
            "remaining_count": max(0, remaining_total - 1),
        })

    api_data  = data.get('data')
    documents = api_data if isinstance(api_data, list) else \
                (api_data.get('documents') or [] if isinstance(api_data, dict) else [])

    if not documents:
        _mark_done({"od_note": "no documents returned"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": False,
            "message": f"No documents returned for {email}",
            "remaining_count": max(0, remaining_total - 1),
        })

    od_doc = None
    for d in documents:
        doc_name = (d.get('document_type_name') or '').strip().lower()
        if any(t in doc_name for t in (
            'open disclosure',
            'opendisclosure',
            'open disclosure training',
            'open disclosure certificate',
            'open disclosure e-learning',
        )) and d.get('url'):
            od_doc = d
            break

    if not od_doc:
        _mark_done({"od_note": "no Open Disclosure document found"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": False,
            "message": f"No Open Disclosure certificate found for {full_name}",
            "remaining_count": max(0, remaining_total - 1),
        })

    doc_url = (od_doc.get('url') or '').strip()

    if not doc_url:
        _mark_done({"od_note": "document found but URL is empty — skipped"})
        return jsonify({
            "success": True, "email": email, "staff_name": full_name,
            "doc_found": True, "skipped": True,
            "reason": "Document URL is empty",
            "remaining_count": max(0, remaining_total - 1),
            "message": f"Skipped {full_name} ({email}) — Open Disclosure doc has no URL",
        })

    try:
        dl_headers = {k: v for k, v in api_headers.items() if k != 'Content-Type'}
        dl_resp    = _req.get(doc_url, headers=dl_headers, timeout=60)

        if dl_resp.status_code == 404:
            _mark_done({"od_note": "document URL 404 — skipped", "od_doc_404": True})
            return jsonify({
                "success": True, "email": email, "staff_name": full_name,
                "doc_found": True, "skipped": True,
                "reason": "Document URL returned 404",
                "remaining_count": max(0, remaining_total - 1),
                "message": f"Skipped {full_name} ({email}) — Open Disclosure doc URL 404",
            })

        dl_resp.raise_for_status()
        raw_bytes    = dl_resp.content
        content_type = dl_resp.headers.get('Content-Type', '').lower()

        client = google_genai.Client(api_key=gemini_key)

        prompt_text = """You are a certificate data extractor.

Extract the following details from this Open Disclosure certificate or training record:
1. Certificate name (e.g. "Open Disclosure", "Open Disclosure Training", "Open Disclosure e-Learning")
2. Staff name as printed on the certificate
3. Expiry date or renewal date (if shown)
4. Issue / completion date
5. Issuing body or training provider

Return ONLY a JSON object — no markdown, no explanation:
{
  "certificate_name": "<exact certificate title as printed>",
  "staff_name_on_cert": "<name as printed on certificate>",
  "expiry_date": "<expiry or renewal date as printed, e.g. 01/06/2027 or June 2027>",
  "issue_date": "<issue or completion date as printed>",
  "issuing_body": "<organization that issued the certificate>"
}

If a field is not visible, set it to null.
"""

        is_image = any(t in content_type for t in ('image/', 'jpeg', 'jpg', 'png', 'webp'))
        is_pdf   = 'pdf' in content_type or doc_url.lower().split('?')[0].endswith('.pdf')

        if is_image:
            ext   = 'jpeg' if any(t in content_type for t in ('jpeg', 'jpg')) else \
                    'png'  if 'png'  in content_type else \
                    'webp' if 'webp' in content_type else 'jpeg'
            parts = [
                {"inline_data": {"mime_type": f"image/{ext}",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash', contents=[{"parts": parts}]
            )
        elif is_pdf:
            parts = [
                {"inline_data": {"mime_type": "application/pdf",
                                 "data": base64.b64encode(raw_bytes).decode()}},
                {"text": prompt_text}
            ]
            response = client.models.generate_content(
                model='gemini-2.5-flash', contents=[{"parts": parts}]
            )
        else:
            try:
                import io as _io, pdfplumber
                with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                    raw_text = chr(10).join(p.extract_text() or '' for p in pdf.pages).strip()
            except Exception:
                raw_text = raw_bytes.decode('utf-8', errors='replace').strip()
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt_text + "\n\nCERTIFICATE TEXT:\n" + raw_text[:5000]
            )

        raw_out = (response.text or '').strip()
        raw_out = _re.sub(r'^```(?:json)?\s*', '', raw_out, flags=_re.MULTILINE)
        raw_out = _re.sub(r'```\s*$', '', raw_out, flags=_re.MULTILINE).strip()

        result       = _json.loads(raw_out)
        cert_name    = _v(result.get('certificate_name') or '')
        cert_staff   = _v(result.get('staff_name_on_cert') or '')
        expiry_date  = _v(result.get('expiry_date') or '')
        issue_date   = _v(result.get('issue_date') or '')
        issuing_body = _v(result.get('issuing_body') or '')

        _mark_done({
            "od_certificate_name": cert_name,
            "od_staff_name":       cert_staff,
            "od_expiry_date":      expiry_date,
            "od_issue_date":       issue_date,
            "od_issuing_body":     issuing_body,
            "od_doc_url":          doc_url,
            "od_doc_type":         od_doc.get('document_type_name', ''),
            "od_note":             "extracted successfully",
        })

        return jsonify({
            "success":            True,
            "email":              email,
            "staff_name":         full_name,
            "doc_found":          True,
            "certificate_name":   cert_name,
            "staff_name_on_cert": cert_staff,
            "expiry_date":        expiry_date,
            "issue_date":         issue_date,
            "issuing_body":       issuing_body,
            "remaining_count":    max(0, remaining_total - 1),
            "message": (
                f"Open Disclosure cert extracted for {full_name} "
                f"(expires: {expiry_date or 'unknown'}) — "
                f"{max(0, remaining_total - 1)} remaining."
            ),
        })

    except _json.JSONDecodeError:
        _mark_done({"od_note": "Gemini JSON parse error"})
        return jsonify({
            "success": False, "email": email,
            "error": "Gemini returned non-JSON",
            "remaining_count": max(0, remaining_total - 1),
        })
    except Exception as e:
        _mark_done({"od_note": f"error: {e}"})
        return jsonify({
            "success": False, "email": email,
            "error": str(e),
            "remaining_count": max(0, remaining_total - 1),
        })


# ── Export: Open Disclosure certificates to Excel ─────────────────────

@admin_bp.route('/live-staffs/export/open-disclosure-xlsx')
@admin_required
def live_staff_export_open_disclosure_xlsx():
    """Export Open Disclosure certificate details to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io as _io

        docs = list(_staffs_col().find(
            {},
            {"section_1_personal_details": 1, "email": 1,
             "od_certificate_name": 1, "od_staff_name": 1,
             "od_expiry_date": 1, "od_issue_date": 1,
             "od_issuing_body": 1, "od_fetched": 1}
        ))
        docs.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        NAVY = '1B3A6B'; GREEN = '2E9E44'; WHITE = 'FFFFFF'
        ALT  = 'EFF6FF'; WARN  = 'FFF3CD'; RED   = 'FFDDDD'

        h_font  = Font(name='Arial', bold=True, color=WHITE, size=10)
        h_fill  = PatternFill('solid', start_color=NAVY, end_color=NAVY)
        h_align = Alignment(horizontal='center', vertical='center')
        b_font  = Font(name='Arial', size=10)
        l_align = Alignment(horizontal='left',   vertical='center')
        c_align = Alignment(horizontal='center', vertical='center')
        thin    = Side(style='thin', color='CCCCCC')
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        green_b = Border(left=thin, right=thin, top=thin,
                         bottom=Side(style='medium', color=GREEN))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Open Disclosure Certificates'

        headers    = ['Sno', 'Staff Name', 'Email', 'Certificate Name',
                      'Name on Cert', 'Expiry Date', 'Issue Date', 'Issuing Body', 'Status']
        col_widths = [5, 28, 36, 35, 28, 16, 16, 30, 14]

        for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = h_font; cell.fill = h_fill
            cell.alignment = h_align; cell.border = green_b
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:I{len(docs)+1}'

        from datetime import date as _date
        today = _date.today()

        def _is_expired(expiry_str):
            if not expiry_str:
                return None
            for fmt in ('%d/%m/%Y', '%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%B %Y', '%b %Y'):
                try:
                    from datetime import datetime as _dt
                    d = _dt.strptime(expiry_str.strip(), fmt).date()
                    return d < today
                except Exception:
                    continue
            return None

        for ri, doc in enumerate(docs, start=2):
            s1       = doc.get('section_1_personal_details') or {}
            name     = _v(s1.get('full_name') or '')
            email    = _v(doc.get('email') or '')
            cert_n   = _v(doc.get('od_certificate_name') or '')
            cert_s   = _v(doc.get('od_staff_name') or '')
            expiry   = _v(doc.get('od_expiry_date') or '')
            issue    = _v(doc.get('od_issue_date') or '')
            issuer   = _v(doc.get('od_issuing_body') or '')
            fetched  = doc.get('od_fetched', False)
            expired  = _is_expired(expiry)

            if not fetched:
                status   = 'Not Checked'
                row_fill = PatternFill('solid', start_color=WARN, end_color=WARN)
            elif not cert_n:
                status   = 'No Cert Found'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is True:
                status   = 'EXPIRED'
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
            elif expired is False:
                status   = 'Valid'
                row_fill = None
            else:
                status   = 'Found'
                row_fill = None

            alt_fill = PatternFill('solid', start_color=ALT, end_color=ALT) \
                       if ri % 2 == 0 and not row_fill else None

            row_vals = [ri-1, name, email, cert_n, cert_s, expiry, issue, issuer, status]
            aligns   = [c_align, l_align, l_align, l_align, l_align,
                        c_align, c_align, l_align, c_align]

            for ci, (val, align) in enumerate(zip(row_vals, aligns), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = b_font; cell.alignment = align
                cell.border = border
                cell.fill = row_fill or alt_fill or PatternFill()

            ws.row_dimensions[ri].height = 17

        ws.cell(row=len(docs)+2, column=1,
                value=f'Total: {len(docs)}').font = Font(name='Arial', bold=True, size=9)

        buf = _io.BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="open_disclosure_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'}
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
