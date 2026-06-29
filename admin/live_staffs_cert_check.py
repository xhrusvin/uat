"""
live_staffs_cert_check.py
══════════════════════════
Independent blueprint — Certificate Validity Checker.

For each staff member, fetches all documents from XN Portal and uses
Gemini AI to extract certificate name and validity status from each document.

Saves results to live_staff_cert_checks collection.

Routes:
    GET  /live-staffs/cron/check-certificates     — cron (one staff per call)
    GET  /live-staffs/export/cert-check-xlsx       — Excel export
    GET  /live-staffs/api/cert-check-status        — check status
"""

from flask import request, jsonify, Response
from bson import ObjectId
from datetime import datetime, date
import os, re, io, json, threading

from database import db
from . import admin_bp
from admin.views import admin_required


# ── Helpers ───────────────────────────────────────────────────────────

def _v(val):
    if val is None: return ''
    return str(val).strip()

def _staffs_col():
    from flask import current_app
    return current_app.db.live_staffs

def _cert_col():
    return db.live_staff_cert_checks


# ── Certificate extraction via Gemini ─────────────────────────────────

def _extract_cert_info(doc_url, doc_type_name, dl_headers, gemini_key):
    """
    Download a document and use Gemini to extract:
    - certificate_name: what type of certificate it is
    - extracted: was info successfully extracted
    - is_valid: True/False/None
    - expiry_date: date string or None
    - notes: any relevant info
    """
    import requests as _req

    try:
        r = _req.get(doc_url, headers=dl_headers, timeout=60)
        r.raise_for_status()
        raw_bytes = r.content
        ct = r.headers.get('Content-Type', '').lower()
        ul = doc_url.lower().split('?')[0]
    except Exception as e:
        return {"extracted": False, "error": str(e)}

    raw_text = ''

    # PDF extraction
    if 'pdf' in ct or ul.endswith('.pdf'):
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
                raw_text = '\n'.join(p.extract_text() or '' for p in pdf.pages).strip()
        except Exception:
            pass

    # DOCX extraction
    if not raw_text and ('wordprocessingml' in ct or ul.endswith('.docx')):
        try:
            from docx import Document as _DDoc
            ddoc = _DDoc(io.BytesIO(raw_bytes))
            raw_text = '\n'.join(p.text for p in ddoc.paragraphs).strip()
        except Exception:
            pass

    # Gemini vision fallback for images/scanned docs
    if not raw_text and gemini_key:
        try:
            import base64 as _b64
            from google import genai as _gai
            client = _gai.Client(api_key=gemini_key)
            mime = ct if ct else 'application/pdf'
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[{"parts": [
                    {"inline_data": {"mime_type": mime,
                                     "data": _b64.b64encode(raw_bytes).decode()}},
                    {"text": "Extract all text from this document. Return plain text only."}
                ]}]
            )
            raw_text = (response.text or '').strip()
        except Exception:
            pass

    if not raw_text:
        return {"extracted": False, "error": "Could not extract text from document"}

    # Use Gemini to analyse the certificate
    if not gemini_key:
        return {"extracted": False, "error": "GEMINI_API_KEY not set"}

    try:
        from google import genai as _gai2
        client = _gai2.Client(api_key=gemini_key)
        prompt = f"""You are a document validator for an Irish healthcare staffing agency.

Analyse this document which was submitted as a "Police Clearance Certificate ( From Country Of Birth )".

Your job is to:
1. Confirm whether this document is genuinely a police clearance certificate or equivalent
2. Extract all key details
3. Determine if it is currently valid

Return ONLY a JSON object — no markdown, no explanation:
{{
  "is_police_clearance": <true if this is genuinely a police clearance/criminal background check certificate, false if not>,
  "not_police_clearance_reason": "<if false — explain exactly what the document appears to be instead>",
  "certificate_name": "<exact certificate name as shown on the document>",
  "holder_name": "<full name of the person on the certificate>",
  "issue_date": "<date issued DD/MM/YYYY or null>",
  "expiry_date": "<expiry date DD/MM/YYYY or null — many police certs expire after 6-12 months>",
  "issuing_body": "<full name of the issuing police authority, government body or agency>",
  "issuing_country": "<country that issued the certificate>",
  "is_valid": <true if is_police_clearance is true AND not expired AND appears legitimate, false otherwise, null if cannot determine>,
  "validity_reason": "<one clear sentence explaining the validity decision — mention expiry date if expired>"
}}

DOCUMENT TEXT:
{raw_text[:8000]}
"""
        response = client.models.generate_content(model='gemini-2.5-flash', contents=prompt)
        raw = re.sub(r'^```(?:json)?\s*', '', (response.text or '').strip(), flags=re.MULTILINE)
        raw = re.sub(r'```\s*$', '', raw, flags=re.MULTILINE).strip()
        result = json.loads(raw)
        result['extracted'] = True
        # Ensure is_valid is False if document is not a police clearance
        if result.get('is_police_clearance') is False:
            result['is_valid'] = False
            if not result.get('validity_reason'):
                result['validity_reason'] = result.get('not_police_clearance_reason', 'Not a police clearance certificate')
        return result
    except Exception as e:
        return {"extracted": False, "error": str(e), "raw_text_length": len(raw_text)}


# ── Cron ──────────────────────────────────────────────────────────────

@admin_bp.route('/live-staffs/cron/check-certificates', methods=['GET', 'POST'])
def live_staff_cron_check_certificates():
    """
    Cron — check all certificates for ONE staff member per call.
    Fetches documents from XN Portal, uses Gemini to validate each.
    Auth: cron_key
    """
    import requests as _req

    cron_secret = os.environ.get('CRON_SECRET', '')
    if cron_secret:
        provided = (request.args.get('cron_key') or
                    request.headers.get('X-Cron-Key', ''))
        if provided != cron_secret:
            return jsonify({"success": False, "error": "Unauthorised"}), 401

    base_url    = os.environ.get('LIVE_STAFF_URL', '').rstrip('/')
    api_key     = os.environ.get('XN_PORTAL_API_KEY', '')
    app_country = os.environ.get('XN_APP_COUNTRY', '')
    gemini_key  = os.environ.get('GEMINI_API_KEY', '')

    if not base_url:
        return jsonify({"success": False, "error": "LIVE_STAFF_URL not set"}), 500

    col      = _staffs_col()
    cert_col = _cert_col()

    # Find next staff not yet checked
    existing_ids = set(
        str(r['staff_id']) for r in cert_col.find({}, {"staff_id": 1})
        if r.get('staff_id')
    )
    pending_query = {
        "_id": {"$nin": [ObjectId(i) for i in existing_ids if len(i) == 24]}
    }
    remaining = col.count_documents(pending_query)
    staff     = col.find_one(pending_query)

    if not staff:
        return jsonify({
            "success":         True,
            "message":         "All staff certificates checked.",
            "remaining_count": 0,
        })

    staff_id  = str(staff['_id'])
    s1        = staff.get('section_1_personal_details') or {}
    full_name = _v(s1.get('full_name') or '')
    email     = _v(staff.get('email') or s1.get('email_address') or '')

    if not email:
        cert_col.insert_one({
            "staff_id":   staff_id, "staff_name": full_name,
            "email":      email, "status": "skipped",
            "note":       "No email", "checked_at": datetime.utcnow()
        })
        return jsonify({"success": True, "message": f"Skipped {full_name} — no email",
                        "remaining_count": max(0, remaining - 1)})

    # Mark as processing
    cert_col.insert_one({
        "staff_id": staff_id, "staff_name": full_name,
        "email": email, "status": "processing", "created_at": datetime.utcnow()
    })

    def _do_check():
        hdrs = {"Api-Key": api_key, "X-App-Country": app_country,
                "Content-Type": "application/json", "Accept": "application/json"}
        dl_hdrs = {k: v for k, v in hdrs.items() if k != 'Content-Type'}

        try:
            r = _req.post(f"{base_url}/ai/recruitments/user-document-list",
                          json={"email": email}, headers=hdrs, timeout=30)
            if r.status_code == 405:
                r = _req.get(f"{base_url}/ai/recruitments/user-document-list",
                             params={"email": email}, headers=hdrs, timeout=30)
            r.raise_for_status()
            api_data = r.json().get('data')
            docs = api_data if isinstance(api_data, list) else \
                   (api_data.get('documents') or [] if isinstance(api_data, dict) else [])
        except Exception as e:
            cert_col.update_one({"staff_id": staff_id, "status": "processing"},
                                {"$set": {"status": "error", "error": str(e),
                                          "checked_at": datetime.utcnow()}})
            return

        # Only check Police Clearance Certificate
        TARGET_DOC = 'police clearance certificate ( from country of birth )'

        cert_results = []
        for doc in docs:
            doc_type = _v(doc.get('document_type_name') or 'Unknown')
            url      = _v(doc.get('url') or '')
            if not url:
                continue
            if doc_type.lower().strip() != TARGET_DOC:
                continue

            info = _extract_cert_info(url, doc_type, dl_hdrs, gemini_key)
            cert_results.append({
                "document_type_name":       doc_type,
                "is_police_clearance":      info.get('is_police_clearance'),
                "not_police_clearance_reason": info.get('not_police_clearance_reason', ''),
                "certificate_name":         info.get('certificate_name', doc_type),
                "holder_name":              info.get('holder_name', ''),
                "issue_date":               info.get('issue_date', ''),
                "expiry_date":              info.get('expiry_date', ''),
                "issuing_body":             info.get('issuing_body', ''),
                "issuing_country":          info.get('issuing_country', ''),
                "is_valid":                 info.get('is_valid'),
                "validity_reason":          info.get('validity_reason', ''),
                "extracted":                info.get('extracted', False),
                "error":                    info.get('error', ''),
            })

        cert_col.update_one(
            {"staff_id": staff_id, "status": "processing"},
            {"$set": {
                "staff_name":    full_name,
                "email":         email,
                "status":        "checked",
                "certificates":  cert_results,
                "total_docs":    len(cert_results),
                "valid_count":   sum(1 for c in cert_results if c.get('is_valid') is True),
                "invalid_count": sum(1 for c in cert_results if c.get('is_valid') is False),
                "pcc_found":     len(cert_results) > 0,
                "checked_at":    datetime.utcnow(),
            }}
        )

    threading.Thread(target=_do_check, daemon=True).start()

    return jsonify({
        "success":         True,
        "staff_id":        staff_id,
        "staff_name":      full_name,
        "email":           email,
        "remaining_count": max(0, remaining - 1),
        "message":         f"Certificate check started for {full_name}",
    })


# ── Excel export ──────────────────────────────────────────────────────

@admin_bp.route('/live-staffs/export/cert-check-xlsx')
@admin_required
def live_staff_export_cert_check_xlsx():
    """Export all certificate check results to Excel — one row per certificate."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Get all staff
        all_staff = list(_staffs_col().find(
            {}, {"section_1_personal_details": 1, "email": 1, "_id": 1}
        ))
        all_staff.sort(key=lambda d: _v(
            (d.get('section_1_personal_details') or {}).get('full_name') or ''
        ).lower())

        # Build lookup of cert check results by staff_id
        cert_map = {
            d['staff_id']: d
            for d in _cert_col().find({"status": "checked"})
            if d.get('staff_id')
        }

        docs = all_staff

        NAVY = '1B3A6B'; WHITE = 'FFFFFF'; GREEN = 'D4EDDA'
        RED  = 'FFDDDD'; WARN  = 'FFF3CD'; ALT   = 'EFF6FF'

        h_font  = Font(name='Calibri', bold=True, color=WHITE, size=10)
        h_fill  = PatternFill('solid', start_color=NAVY, end_color=NAVY)
        h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        b_font  = Font(name='Calibri', size=10)
        thin    = Side(style='thin', color='CCCCCC')
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Police Clearance Certificates'

        headers    = ['Sno', 'Staff Name', 'Email',
                      'Certificate Name', 'Is Police Cert',
                      'Holder Name', 'Issue Date', 'Expiry Date',
                      'Valid', 'Validity Reason',
                      'Issuing Body', 'Issuing Country',
                      'Not PCC Reason', 'Extracted', 'Checked At']
        col_widths = [5, 28, 32, 28, 14, 22, 12, 12, 8, 40, 28, 16, 40, 10, 18]

        for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = h_font; cell.fill = h_fill
            cell.alignment = h_align; cell.border = border
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:M1'

        row_num = 2
        sno     = 0

        for staff_doc in docs:
            sno    += 1
            s1      = staff_doc.get('section_1_personal_details') or {}
            name    = _v(s1.get('full_name') or '')
            email   = _v(staff_doc.get('email') or '')
            sid     = str(staff_doc['_id'])
            checked_doc = cert_map.get(sid)

            if not checked_doc:
                # Not yet checked
                row_vals = [sno, name, email, 'Police Clearance Certificate',
                            '', '', '', 'Not Checked', 'Not yet analysed', '', '', '']
                row_fill = PatternFill('solid', start_color=WARN, end_color=WARN)
                aligns   = ['center','left','left','left','left','center',
                            'center','center','left','left','center','center']
                for ci, (val, align) in enumerate(zip(row_vals, aligns), 1):
                    cell = ws.cell(row=row_num, column=ci, value=val)
                    cell.font = b_font; cell.border = border; cell.fill = row_fill
                    cell.alignment = Alignment(horizontal=align, vertical='center')
                ws.row_dimensions[row_num].height = 16
                row_num += 1
                continue

            checked   = checked_doc.get('checked_at')
            checked_s = checked.strftime('%d %b %Y %H:%M') if checked else ''
            certs     = checked_doc.get('certificates') or []

            if not certs:
                # Checked but no PCC found
                row_vals = [sno, name, email, 'Police Clearance Certificate',
                            '', '', '', 'No PCC Found', 'No PCC document in portal',
                            '', 'Yes', checked_s]
                row_fill = PatternFill('solid', start_color=RED, end_color=RED)
                aligns   = ['center','left','left','left','left','center',
                            'center','center','left','left','center','center']
                for ci, (val, align) in enumerate(zip(row_vals, aligns), 1):
                    cell = ws.cell(row=row_num, column=ci, value=val)
                    cell.font = b_font; cell.border = border; cell.fill = row_fill
                    cell.alignment = Alignment(horizontal=align, vertical='center')
                ws.row_dimensions[row_num].height = 16
                row_num += 1
                continue

            for cert in certs:
                is_pcc    = cert.get('is_police_clearance')
                is_valid  = cert.get('is_valid')
                valid_str = 'Yes' if is_valid is True else ('No' if is_valid is False else '?')
                pcc_str   = 'Yes' if is_pcc is True else ('No' if is_pcc is False else '?')

                if is_valid is True:
                    row_fill = PatternFill('solid', start_color=GREEN, end_color=GREEN)
                elif is_valid is False:
                    row_fill = PatternFill('solid', start_color=RED, end_color=RED)
                else:
                    row_fill = PatternFill('solid', start_color=WARN, end_color=WARN)

                row_vals = [
                    sno, name, email,
                    cert.get('certificate_name', 'Police Clearance Certificate'),
                    pcc_str,
                    cert.get('holder_name', ''),
                    cert.get('issue_date', ''),
                    cert.get('expiry_date', ''),
                    valid_str,
                    cert.get('validity_reason', ''),
                    cert.get('issuing_body', ''),
                    cert.get('issuing_country', ''),
                    cert.get('not_police_clearance_reason', ''),
                    'Yes' if cert.get('extracted') else 'No',
                    checked_s,
                ]
                aligns = ['center','left','left','left','center','left',
                          'center','center','center','left','left','left',
                          'left','center','center']
                for ci, (val, align) in enumerate(zip(row_vals, aligns), 1):
                    cell = ws.cell(row=row_num, column=ci, value=val)
                    cell.font = b_font; cell.border = border; cell.fill = row_fill
                    cell.alignment = Alignment(horizontal=align, vertical='center')
                ws.row_dimensions[row_num].height = 16
                row_num += 1

        ws.cell(row=row_num, column=1,
                value=f'Total: {sno}').font = Font(name='Calibri', bold=True, size=9)

        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition":
                     f'attachment; filename="cert_check_{datetime.utcnow().strftime("%Y%m%d")}.xlsx"'}
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ── Status API ────────────────────────────────────────────────────────

@admin_bp.route('/live-staffs/api/cert-check-status')
@admin_required
def live_staff_api_cert_check_status():
    """Return summary of certificate check progress."""
    col = _cert_col()
    total    = _staffs_col().count_documents({})
    checked  = col.count_documents({"status": "checked"})
    errors   = col.count_documents({"status": "error"})
    skipped  = col.count_documents({"status": "skipped"})
    pending  = total - checked - errors - skipped

    return jsonify({
        "success":   True,
        "total":     total,
        "checked":   checked,
        "errors":    errors,
        "skipped":   skipped,
        "pending":   pending,
        "pct":       round(checked / total * 100, 1) if total else 0,
    })
