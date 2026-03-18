# booking/shift_users.py
from flask import Blueprint, render_template, request, jsonify, url_for
from bson import ObjectId
from datetime import datetime

from database import db
from .models.shift import Shift
from .models.shift_user import ShiftUser
from .models.client import Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Response
import io

shift_model     = Shift(db.shifts)
shift_user_model = ShiftUser(db.shifts_users)
client_model    = Client(db.clients)
import requests
from flask import current_app
from flask import send_file
import io

from . import bp
from admin.views import admin_required






@bp.route('/assignments')
@admin_required
def all_assignments_list():

    page        = int(request.args.get('page', 1))
    per_page    = 10
    q           = request.args.get('q', '').strip()
    date_str    = request.args.get('date', '').strip()
    avail_filter = request.args.get('avail', '').strip()

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    # ─────────────────────────────────────────────
    # STEP 1: Pre-filter shift IDs by date (hits index on shifts.date)
    # ─────────────────────────────────────────────
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, "%Y-%m-%d")
            day_start = selected_date.replace(hour=0,  minute=0,  second=0,  microsecond=0)
            day_end   = selected_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            date_match = {"date": {"$gte": day_start, "$lte": day_end}}
        except ValueError:
            date_match = {"date": {"$gte": today}}
    else:
        date_match = {"date": {"$gte": today}}

    relevant_shift_ids = db.shifts.distinct("_id", date_match)

    if not relevant_shift_ids:
        # No shifts in range → return empty immediately, no aggregation needed
        return render_template(
            'booking/assignments_list.html',
            assignments=[], page=1, pages=1, total=0, q=q,
        )

    # ─────────────────────────────────────────────
    # STEP 2: Build base pipeline — match on shifts_users first
    # ─────────────────────────────────────────────
    base_match = {"shift_id": {"$in": relevant_shift_ids}}

    # Availability filter before any joins
    if avail_filter == "unknown":
        base_match["$or"] = [
            {"availability": {"$exists": False}},
            {"availability": None},
            {"availability": {"$nin": [0, 1, 3, 4, 6]}},
        ]
    elif avail_filter in ["0", "1", "3", "4", "6"]:
        base_match["availability"] = int(avail_filter)

    pipeline = [{"$match": base_match}]

    # ─────────────────────────────────────────────
    # STEP 3: Joins (now on already-filtered small set)
    # ─────────────────────────────────────────────
    pipeline.extend([
        # Join shifts
        {
            "$lookup": {
                "from": "shifts",
                "localField": "shift_id",
                "foreignField": "_id",
                "as": "shift"
            }
        },
        {"$unwind": {"path": "$shift", "preserveNullAndEmptyArrays": True}},

        # Join users
        {
            "$lookup": {
                "from": "users",
                "localField": "user_id",
                "foreignField": "_id",
                "as": "user"
            }
        },
        {"$unwind": {"path": "$user", "preserveNullAndEmptyArrays": True}},

        # Drop records with no matching user
        {"$match": {"user._id": {"$exists": True}}},

        # Join clients (string-based client_id)
        {
            "$lookup": {
                "from": "clients",
                "let": {"client_id_str": "$shift.client_id"},
                "pipeline": [
                    {
                        "$match": {
                            "$expr": {
                                "$eq": [{"$toString": "$_id"}, "$$client_id_str"]
                            }
                        }
                    }
                ],
                "as": "client"
            }
        },
        {"$unwind": {"path": "$client", "preserveNullAndEmptyArrays": True}},
    ])

    # ─────────────────────────────────────────────
    # STEP 4: Search filter (after joins, needs joined fields)
    # ─────────────────────────────────────────────
    if q:
        pipeline.append({
            "$match": {
                "$or": [
                    {"user.name":        {"$regex": q, "$options": "i"}},
                    {"user.first_name":  {"$regex": q, "$options": "i"}},
                    {"user.last_name":   {"$regex": q, "$options": "i"}},
                    {"client.name":      {"$regex": q, "$options": "i"}},
                    {"client.county":    {"$regex": q, "$options": "i"}},
                    {"shift.shift_xn_id":{"$regex": q, "$options": "i"}},
                ]
            }
        })

    # ─────────────────────────────────────────────
    # STEP 5: $facet — single query for data + count
    # ─────────────────────────────────────────────
    projection = {
        "assignment_id": "$_id",
        "shift_id":   1,
        "user_id":    1,
        "created_at": 1,
        "availability": 1,
        "assigned_at":  1,
        "sms_sent":     1,     
        "sms_sent_at":  1,         

        "shift_name":       {"$ifNull": ["$shift.name", "—"]},
        "shift_date":       "$shift.date",
        "shift_start_time": "$shift.start_time",
        "shift_end_time":   "$shift.end_time",
        "shift_xn_id":      "$shift.shift_xn_id",

        "client_name":   {"$ifNull": ["$client.name",   "—"]},
        "client_county": {"$ifNull": ["$client.county", "—"]},

        "staff_name": {
            "$ifNull": [
                "$user.name",
                {"$trim": {"input": {"$concat": [
                    {"$ifNull": ["$user.first_name", ""]},
                    " ",
                    {"$ifNull": ["$user.last_name",  ""]}
                ]}}}
            ]
        },
        "staff_email": {"$ifNull": ["$user.email", "—"]},
        "staff_phone": {"$ifNull": ["$user.phone", "—"]},
    }

    pipeline.append({
        "$facet": {
            "data": [
                {"$sort":    {"assigned_at": -1}},
                {"$skip":    (page - 1) * per_page},
                {"$limit":   per_page},
                {"$project": projection},
            ],
            "total": [
                {"$count": "count"}
            ]
        }
    })

    # ─────────────────────────────────────────────
    # STEP 6: Execute + unpack
    # ─────────────────────────────────────────────
    result      = list(db.shifts_users.aggregate(pipeline, allowDiskUse=True))
    assignments = result[0]["data"]        if result else []
    total       = result[0]["total"][0]["count"] if result and result[0]["total"] else 0
    pages       = max(1, (total + per_page - 1) // per_page)

    # ─────────────────────────────────────────────
    # STEP 7: Format dates in Python
    # ─────────────────────────────────────────────
    for a in assignments:
        if isinstance(a.get('created_at'), datetime):
            a['created_at_formatted'] = a['created_at'].strftime('%Y-%m-%d %H:%M')
        if isinstance(a.get('shift_date'), datetime):
            a['shift_date_formatted'] = a['shift_date'].strftime('%d %b %Y')

    return render_template(
        'booking/assignments_list.html',
        assignments=assignments,
        page=page,
        pages=pages,
        total=total,
        q=q,
    )

@bp.route('/shift-users')
def shift_users_list():
    """
    Original view: list shifts with assignment summary
    """
    page = int(request.args.get('page', 1))
    per_page = 10
    q = request.args.get('q', '').strip()
    client_id_str = request.args.get('client_id')

    query = {}
    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"location": {"$regex": q, "$options": "i"}},
        ]
    if client_id_str:
        try:
            query["client_id"] = ObjectId(client_id_str)
        except:
            pass

    pipeline = [
        {"$match": query},
        {"$sort": {"date": -1, "created_at": -1}},
        {"$skip": (page - 1) * per_page},
        {"$limit": per_page},
        {"$lookup": {
            "from": "shifts_users",
            "localField": "_id",
            "foreignField": "shift_id",
            "as": "assignments"
        }},
        {"$addFields": {
            "assigned_count": {"$size": "$assignments"}
        }},
        {"$lookup": {
    "from": "clients",
    "let": {"client_id_obj": {"$toObjectId": "$shift.client_id"}},  # may fail if string is invalid
    "pipeline": [
        {"$match": {"$expr": {"$eq": ["$_id", "$$client_id_obj"]}}},
    ],
    "as": "client"
}},
{"$unwind": {"path": "$client", "preserveNullAndEmptyArrays": True}},
        {"$project": {
            "_id": 1,
            "name": 1,
            "date": 1,
            "start_time": 1,
            "end_time": 1,
            "location": 1,
            "client_name": {
            "$cond": [
                {"$eq": ["$client_id", None]},
                "No client",
                {"$ifNull": ["$client.name", "— (client deleted)"]}
            ]
        },
            "client_type": "$client.type",
            "assigned_count": 1,
            "sample_users": {"$slice": ["$assignments", 3]}
        }}
    ]

    shifts = list(db.shifts.aggregate(pipeline))

    for s in shifts:
        if s.get('start_time') and isinstance(s['start_time'], datetime):
            s['start_time_formatted'] = s['start_time'].strftime('%I:%M %p')
        if s.get('end_time') and isinstance(s['end_time'], datetime):
            s['end_time_formatted'] = s['end_time'].strftime('%I:%M %p')

    total = db.shifts.count_documents(query)
    pages = (total + per_page - 1) // per_page

    clients = list(db.clients.find({"is_active": True}).sort("name", 1).limit(300))

    return render_template(
        'booking/shift_users.html',
        shifts=shifts,
        page=page,
        pages=pages,
        total=total,
        per_page=per_page,
        q=q,
        client_id=client_id_str,
        clients=clients,
    )

from flask import jsonify, current_app, request
from bson import ObjectId
import os
import aiohttp
import asyncio
from functools import partial

# Helper to run async code in sync Flask view (you probably already have something like this)
def run_async(coro):
    loop = asyncio.get_event_loop()
    return loop.run_until_complete(coro)


@bp.route('/assignment-conversation', methods=['GET'])
def get_assignment_conversation():
    shift_id = request.args.get('shift_id')
    user_id = request.args.get('user_id')

    conv = db.shift_booking_conv.find_one(
        {"shift_id": shift_id, "user_id": user_id},
        {"_id": 0}
    )

    if not conv:
        return jsonify({"error": "No conversation found"}), 404

    # format timestamps
    if conv.get("started_at"):
        conv["started_at"] = conv["started_at"].isoformat()
    if conv.get("ended_at"):
        conv["ended_at"] = conv["ended_at"].isoformat()

    for turn in conv.get("turns", []):
        if turn.get("ts"):
            turn["ts"] = turn["ts"].isoformat()

    return jsonify({
        "conversation": conv,
        "audio_url": url_for(
            "booking.get_assignment_audio",  # ✅ correct
            shift_id=shift_id,
            user_id=user_id
        )
    })

@bp.route('/assignment-conversation/audio', methods=['GET'])
def get_assignment_audio():
    shift_id = request.args.get('shift_id')
    user_id = request.args.get('user_id')

    conv = db.shift_booking_conv.find_one(
        {"shift_id": shift_id, "user_id": user_id}
    )

    if not conv:
        return jsonify({"error": "No conversation found"}), 404

    el_conv_id = conv.get("elevenlabs_conversation_id")
    if not el_conv_id:
        return jsonify({"error": "No audio available"}), 404

    api_key = os.getenv("ELEVENLABS_API_KEY")
    url = f"https://api.elevenlabs.io/v1/convai/conversations/{el_conv_id}/audio"

    resp = requests.get(url, headers={"xi-api-key": api_key})

    if resp.status_code != 200:
        return jsonify({"error": "Failed to fetch audio"}), 400

    return send_file(
        io.BytesIO(resp.content),
        mimetype="audio/mpeg",
        as_attachment=False
    )

@bp.route('/assignments/export')
def export_assignments_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import Response
    import io

    q = request.args.get('q', '').strip()
    date_str = request.args.get('date', '').strip()
    avail_filter = request.args.get('avail', '').strip()

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    pipeline = [
        {"$lookup": {"from": "shifts", "localField": "shift_id", "foreignField": "_id", "as": "shift"}},
        {"$unwind": {"path": "$shift", "preserveNullAndEmptyArrays": True}},
        {
            "$lookup": {
                "from": "clients",
                "let": {"client_id_str": "$shift.client_id"},
                "pipeline": [{"$match": {"$expr": {"$eq": [{"$toString": "$_id"}, "$$client_id_str"]}}}],
                "as": "client"
            }
        },
        {"$unwind": {"path": "$client", "preserveNullAndEmptyArrays": True}},
        {"$lookup": {"from": "users", "localField": "user_id", "foreignField": "_id", "as": "user"}},
        {"$unwind": {"path": "$user", "preserveNullAndEmptyArrays": True}},
        {"$match": {"user._id": {"$exists": True}}},
    ]

    # if date_str:
    #     try:
    #         selected_date = datetime.strptime(date_str, "%Y-%m-%d")
    #         day_start = selected_date.replace(hour=0, minute=0, second=0, microsecond=0)
    #         day_end   = selected_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    #         pipeline.append({"$match": {"shift.date": {"$gte": day_start, "$lte": day_end}}})
    #     except ValueError:
    #         # Invalid date string — fall back to today onwards
    #         pipeline.append({"$match": {"shift.date": {"$gte": today}}})
    # else:
    #     pipeline.append({"$match": {"shift.date": {"$gte": today}}})

    if q:
        pipeline.append({"$match": {"$or": [
            {"user.name": {"$regex": q, "$options": "i"}},
            {"user.first_name": {"$regex": q, "$options": "i"}},
            {"user.last_name": {"$regex": q, "$options": "i"}},
            {"client.name": {"$regex": q, "$options": "i"}},
            {"client.county": {"$regex": q, "$options": "i"}},
            {"shift.shift_xn_id": {"$regex": q, "$options": "i"}},
        ]}})

    if avail_filter == "unknown":
        pipeline.append({"$match": {"$or": [
            {"availability": {"$exists": False}},
            {"availability": None},
            {"availability": {"$nin": [0, 1, 3, 4, 6]}}
        ]}})
    elif avail_filter in ["0", "1", "3", "4", "6"]:
        pipeline.append({"$match": {"availability": int(avail_filter)}})

    pipeline.append({"$project": {
        "availability": 1, "assigned_at": 1,
        "shift_name": {"$ifNull": ["$shift.name", "—"]},
        "shift_date": "$shift.date",
        "shift_start_time": "$shift.start_time",
        "shift_end_time": "$shift.end_time",
        "shift_xn_id": "$shift.shift_xn_id",
        "client_name": {"$ifNull": ["$client.name", "—"]},
        "client_county": {"$ifNull": ["$client.county", "—"]},
        "staff_name": {"$ifNull": ["$user.name", {"$trim": {"input": {"$concat": [
            {"$ifNull": ["$user.first_name", ""]}, " ", {"$ifNull": ["$user.last_name", ""]}
        ]}}}]},
        "staff_email": {"$ifNull": ["$user.email", "—"]},
        "staff_phone": {"$ifNull": ["$user.phone", "—"]},
    }})

    records = list(db.shifts_users.aggregate(pipeline))

    # ── Summary counts ──────────────────────────────────────────────────
    total_calls   = len(records)
    available     = sum(1 for r in records if r.get('availability') == 1)
    not_available = sum(1 for r in records if r.get('availability') == 0)
    voice_mail    = sum(1 for r in records if r.get('availability') == 3)
    not_attended  = sum(1 for r in records if r.get('availability') == 4)
    not_triggered = sum(1 for r in records if r.get('availability') == 6)
    no_response   = sum(1 for r in records if r.get('availability') not in [0, 1, 3, 4, 6])

    # ── Helper: always create fresh style objects ───────────────────────
    def make_fill(hex_color):
        return PatternFill("solid", start_color=hex_color, end_color=hex_color)

    def make_border():
        thin = Side(style="thin", color="B8CCE4")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Workbook ────────────────────────────────────────────────────────
    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    # Title
    ws.merge_cells("A1:B1")
    ws["A1"].value = "SHIFT BOOKING — ANALYSIS"
    ws["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=15)
    ws["A1"].fill = make_fill("1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # Export timestamp
    ws.merge_cells("A2:B2")
    ws["A2"].value = f"Exported: {datetime.now().strftime('%d %b %Y  %H:%M')}{'   |   Date filter: ' + date_str if date_str else ''}"
    ws["A2"].font = Font(italic=True, name="Arial", size=10, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    summary_rows = [
        ("Metric",               "Count",       "2E75B6", True),
        ("Total Calls",          total_calls,   "D6E4F0", True),
        ("Available",            available,     "E8F5E9", False),
        ("Not Available",        not_available, "FEECEC", False),
        ("Voice Mail",           voice_mail,    "FFFDE7", False),
        ("Call Not Attended",    not_attended,  "E3F2FD", False),
        ("Call Not Triggered",   not_triggered, "FFF3E0", False),
        ("No Response from User",no_response,   "F3E5F5", False),
    ]

    for i, (label, value, color, bold) in enumerate(summary_rows, start=3):
        ca = ws.cell(row=i, column=1, value=label)
        cb = ws.cell(row=i, column=2, value=value)

        for cell in (ca, cb):
            cell.fill      = make_fill(color)
            cell.font      = Font(bold=bold, name="Arial", size=11,
                                  color="FFFFFF" if color == "2E75B6" else "1F1F1F")
            cell.border    = make_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ca.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[i].height = 22

    # ════════════════════════════════════════════════════════════════════
    # Sheet 2 — Assignments
    # ════════════════════════════════════════════════════════════════════
    wd = wb.create_sheet("Assignments")

    AVAIL_MAP = {
        1: "Available",
        0: "Not Available",
        3: "Voice Mail",
        4: "Call Not Attended",
        6: "Call Not Triggered",
    }

    headers = [
        "Shift ID", "Shift Name", "Shift Date", "Start Time", "End Time",
        "Client Name", "County", "Staff Name", "Staff Email", "Staff Phone",
        "Availability", "Assigned At"
    ]
    col_widths = [14, 26, 14, 12, 12, 26, 18, 24, 30, 16, 22, 20]

    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = wd.cell(row=1, column=col_idx, value=hdr)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill      = make_fill("2E75B6")
        cell.border    = make_border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wd.column_dimensions[get_column_letter(col_idx)].width = width

    wd.row_dimensions[1].height = 28
    wd.freeze_panes = "A2"

    for row_idx, r in enumerate(records, start=2):
        row_color = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"

        avail_val   = r.get('availability')
        avail_label = AVAIL_MAP.get(avail_val, "No Response from User")

        shift_date  = r['shift_date'].strftime('%d %b %Y') if isinstance(r.get('shift_date'), datetime) else ''
        assigned_at = r['assigned_at'].strftime('%d %b %Y %H:%M') if isinstance(r.get('assigned_at'), datetime) else ''

        row_data = [
            str(r.get('shift_xn_id') or '—'),
            r.get('shift_name') or '—',
            shift_date,
            str(r.get('shift_start_time') or ''),
            str(r.get('shift_end_time') or ''),
            r.get('client_name') or '—',
            r.get('client_county') or '—',
            r.get('staff_name') or '—',
            r.get('staff_email') or '—',
            r.get('staff_phone') or '—',
            avail_label,
            assigned_at,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = wd.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = make_fill(row_color)
            cell.border    = make_border()
            cell.alignment = Alignment(vertical="center")

        wd.row_dimensions[row_idx].height = 18

    # ── Stream to browser ───────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"shift_assignments_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@bp.route('/assignment-sms-history', methods=['GET'])
def get_assignment_sms_history():
    shift_id = request.args.get('shift_id')
    user_id  = request.args.get('user_id')

    if not shift_id or not user_id:
        return jsonify({"error": "Missing shift_id or user_id"}), 400

    # ── Sent SMS logs ──────────────────────────────────────────
    sent_logs = list(db.sms_log.find(
        {"shift_id": shift_id, "user_id": user_id},
        {"_id": 0}
    ).sort("sent_at", 1))

    # ── Replies from this user for this shift ──────────────────
    replies = list(db.sms_replies.find(
        {"shift_id": shift_id, "user_id": user_id},
        {"_id": 0}
    ).sort("received_at", 1))

    # ── Serialise datetimes ────────────────────────────────────
    for log in sent_logs:
        if isinstance(log.get('sent_at'), datetime):
            log['sent_at'] = log['sent_at'].isoformat()

    for reply in replies:
        if isinstance(reply.get('received_at'), datetime):
            reply['received_at'] = reply['received_at'].isoformat()

    return jsonify({
        "sent":    sent_logs,
        "replies": replies,
    })